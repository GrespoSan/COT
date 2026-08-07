from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable
import io
import math
import re
import hashlib

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
import yfinance as yf
from PIL import Image, ImageDraw, ImageFont


# =============================================================================
# CONFIGURAZIONE PAGINA
# =============================================================================
st.set_page_config(
    page_title="COT Smart Money V6.17 — Python",
    page_icon="🛡️",
    layout="wide",
)

st.markdown(
    """
    <style>
    .block-container {padding-top: 1.25rem; padding-bottom: 2rem;}
    .cot-card {
        border: 1px solid rgba(128,128,128,.35);
        border-left: 8px solid var(--accent);
        border-radius: 10px;
        padding: 16px 18px;
        margin: 4px 0 14px 0;
        background: rgba(127,127,127,.055);
    }
    .cot-kicker {font-size:.78rem; font-weight:800; letter-spacing:.04em; opacity:.8;}
    .cot-title {font-size:1.42rem; font-weight:850; color:var(--accent); margin:.15rem 0 .45rem 0;}
    .cot-small {font-size:.93rem; line-height:1.45;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("🛡️ COT Smart Money — Python V6.17")
st.caption(
    "Due sezioni indipendenti: analisi approfondita di un singolo future e screener settimanale di tutti i mercati. "
    "Il motore seleziona automaticamente TFF per i finanziari e Disaggregated per le commodity."
)


# =============================================================================
# COSTANTI CFTC
# =============================================================================
CFTC_DATASETS = {
    "Disaggregated": "72hh-3qpy",
    "Financial": "gpe5-46if",
}

CFTC_API_BASES = (
    "https://publicreporting.cftc.gov/resource",
    "https://publicreportinghub.cftc.gov/resource",
)

TERM_OPTIONS = ["Non disponibile", "Contango", "Backwardation", "Curva piatta"]
AI_PROMPT_FILENAME = "PROMPT.TXT"
AI_SCREENER_PROMPT_FILENAME = "PROMPT_SCREENER.TXT"
ALIGNMENT_UPPER = 80.0
ALIGNMENT_LOWER = 20.0
OI_INDEX_LOOKBACK = 52
RAPID_SHIFT_EXTREME = 40.0
RAPID_SHIFT_WARNING = 20.0
COT_INDEX_SHORT_LOOKBACK = 26
COT_INDEX_LONG_LOOKBACK = 156


@dataclass(frozen=True)
class MarketSpec:
    label: str
    root: str
    group: str
    family: str  # commodity | financial
    is_fx: bool
    preferred_name: str
    search_terms: tuple[str, ...]
    yahoo_ticker: str

    @property
    def specific_report(self) -> str:
        return "Disaggregated" if self.family == "commodity" else "Financial"

    @property
    def trend_label(self) -> str:
        return "Managed Money" if self.family == "commodity" else "Leveraged Funds"

    @property
    def counter_label(self) -> str:
        if self.family == "commodity":
            return "Producer / Merchant"
        if self.is_fx:
            return "Dealer / Intermediary"
        return "Asset Manager"

    @property
    def market_family_label(self) -> str:
        if self.family == "commodity":
            return "COMMODITY"
        if self.is_fx:
            return "VALUTE"
        return "FINANZIARI"


# La lista conserva i mercati già presenti nella prima app e aggiunge alcuni
# contratti previsti dal motore Pine.
MARKETS: tuple[MarketSpec, ...] = (
    # Indici e volatilità
    MarketSpec("ES — S&P 500", "ES", "Indici", "financial", False, "S&P 500 Consolidated - CHICAGO MERCANTILE EXCHANGE", ("S&P 500", "CONSOLIDATED"), "ES=F"),
    MarketSpec("NQ — Nasdaq 100", "NQ", "Indici", "financial", False, "NASDAQ MINI - CHICAGO MERCANTILE EXCHANGE", ("NASDAQ MINI", "NASDAQ"), "NQ=F"),
    MarketSpec("YM — Dow Jones", "YM", "Indici", "financial", False, "DJIA x $5 - CHICAGO BOARD OF TRADE", ("DJIA", "$5"), "YM=F"),
    MarketSpec("RTY — Russell 2000", "RTY", "Indici", "financial", False, "RUSSELL E-MINI - CHICAGO MERCANTILE EXCHANGE", ("RUSSELL", "E-MINI"), "RTY=F"),
    MarketSpec("VX — VIX Futures", "VX", "Indici", "financial", False, "VIX FUTURES - CBOE FUTURES EXCHANGE", ("VIX",), "^VIX"),

    # Valute
    MarketSpec("DX — U.S. Dollar Index", "DX", "Valute", "financial", True, "USD INDEX - ICE FUTURES U.S.", ("USD INDEX",), "DX-Y.NYB"),
    MarketSpec("6E — Euro FX", "6E", "Valute", "financial", True, "EURO FX - CHICAGO MERCANTILE EXCHANGE", ("EURO FX",), "6E=F"),
    MarketSpec("6B — British Pound", "6B", "Valute", "financial", True, "BRITISH POUND - CHICAGO MERCANTILE EXCHANGE", ("BRITISH POUND",), "6B=F"),
    MarketSpec("6A — Australian Dollar", "6A", "Valute", "financial", True, "AUSTRALIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE", ("AUSTRALIAN DOLLAR",), "6A=F"),
    MarketSpec("6N — New Zealand Dollar", "6N", "Valute", "financial", True, "NZ DOLLAR - CHICAGO MERCANTILE EXCHANGE", ("NZ DOLLAR", "NEW ZEALAND"), "6N=F"),
    MarketSpec("6C — Canadian Dollar", "6C", "Valute", "financial", True, "CANADIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE", ("CANADIAN DOLLAR",), "6C=F"),
    MarketSpec("6S — Swiss Franc", "6S", "Valute", "financial", True, "SWISS FRANC - CHICAGO MERCANTILE EXCHANGE", ("SWISS FRANC",), "6S=F"),
    MarketSpec("6J — Japanese Yen", "6J", "Valute", "financial", True, "JAPANESE YEN - CHICAGO MERCANTILE EXCHANGE", ("JAPANESE YEN",), "6J=F"),

    # Tassi
    MarketSpec("ZQ — 30-Day Federal Funds", "ZQ", "Tassi", "financial", False, "FED FUNDS - CHICAGO BOARD OF TRADE", ("FED FUNDS",), "ZQ=F"),
    MarketSpec("ZT — U.S. Treasury 2Y", "ZT", "Tassi", "financial", False, "UST 2Y NOTE - CHICAGO BOARD OF TRADE", ("UST 2Y", "2Y NOTE"), "ZT=F"),
    MarketSpec("ZF — U.S. Treasury 5Y", "ZF", "Tassi", "financial", False, "UST 5Y NOTE - CHICAGO BOARD OF TRADE", ("UST 5Y", "5Y NOTE"), "ZF=F"),
    MarketSpec("ZN — U.S. Treasury 10Y", "ZN", "Tassi", "financial", False, "UST 10Y NOTE - CHICAGO BOARD OF TRADE", ("UST 10Y", "10Y NOTE"), "ZN=F"),
    MarketSpec("ZB — U.S. Treasury Bond", "ZB", "Tassi", "financial", False, "UST BOND - CHICAGO BOARD OF TRADE", ("UST BOND", "TREASURY BONDS"), "ZB=F"),

    # Crypto CME
    MarketSpec("BTC — Bitcoin CME", "BTC", "Crypto CME", "financial", False, "BITCOIN - CHICAGO MERCANTILE EXCHANGE", ("BITCOIN",), "BTC=F"),
    MarketSpec("ETH — Ether CME", "ETH", "Crypto CME", "financial", False, "ETHER CASH SETTLED - CHICAGO MERCANTILE EXCHANGE", ("ETHER CASH", "ETHER"), "ETH=F"),

    # Energia
    MarketSpec("CL — WTI Crude Oil", "CL", "Energia", "commodity", False, "WTI-PHYSICAL - NEW YORK MERCANTILE EXCHANGE", ("WTI-PHYSICAL", "CRUDE OIL"), "CL=F"),
    MarketSpec("RB — Gasoline RBOB", "RB", "Energia", "commodity", False, "GASOLINE RBOB - NEW YORK MERCANTILE EXCHANGE", ("GASOLINE RBOB",), "RB=F"),
    MarketSpec("NG — Natural Gas", "NG", "Energia", "commodity", False, "NAT GAS NYME - NEW YORK MERCANTILE EXCHANGE", ("NAT GAS", "NATURAL GAS"), "NG=F"),
    MarketSpec("HO — Heating Oil", "HO", "Energia", "commodity", False, "NO. 2 HEATING OIL, N.Y. HARBOR - NEW YORK MERCANTILE EXCHANGE", ("HEATING OIL",), "HO=F"),
    MarketSpec("BZ — Brent", "BZ", "Energia", "commodity", False, "BRENT LAST DAY - NEW YORK MERCANTILE EXCHANGE", ("BRENT",), "BZ=F"),

    # Metalli
    MarketSpec("GC — Gold", "GC", "Metalli", "commodity", False, "GOLD - COMMODITY EXCHANGE INC.", ("GOLD", "COMMODITY EXCHANGE"), "GC=F"),
    MarketSpec("SI — Silver", "SI", "Metalli", "commodity", False, "SILVER - COMMODITY EXCHANGE INC.", ("SILVER",), "SI=F"),
    MarketSpec("HG — Copper", "HG", "Metalli", "commodity", False, "COPPER - COMMODITY EXCHANGE INC.", ("COPPER",), "HG=F"),
    MarketSpec("PL — Platinum", "PL", "Metalli", "commodity", False, "PLATINUM - NEW YORK MERCANTILE EXCHANGE", ("PLATINUM",), "PL=F"),
    MarketSpec("PA — Palladium", "PA", "Metalli", "commodity", False, "PALLADIUM - NEW YORK MERCANTILE EXCHANGE", ("PALLADIUM",), "PA=F"),

    # Cereali e semi oleosi
    MarketSpec("ZC — Corn", "ZC", "Agricoli", "commodity", False, "CORN - CHICAGO BOARD OF TRADE", ("CORN",), "ZC=F"),
    MarketSpec("ZS — Soybeans", "ZS", "Agricoli", "commodity", False, "SOYBEANS - CHICAGO BOARD OF TRADE", ("SOYBEANS",), "ZS=F"),
    MarketSpec("ZL — Soybean Oil", "ZL", "Agricoli", "commodity", False, "SOYBEAN OIL - CHICAGO BOARD OF TRADE", ("SOYBEAN OIL",), "ZL=F"),
    MarketSpec("ZM — Soybean Meal", "ZM", "Agricoli", "commodity", False, "SOYBEAN MEAL - CHICAGO BOARD OF TRADE", ("SOYBEAN MEAL",), "ZM=F"),
    MarketSpec("ZW — Wheat SRW", "ZW", "Agricoli", "commodity", False, "WHEAT-SRW - CHICAGO BOARD OF TRADE", ("WHEAT-SRW",), "ZW=F"),
    MarketSpec("KE — Wheat HRW", "KE", "Agricoli", "commodity", False, "WHEAT-HRW - CHICAGO BOARD OF TRADE", ("WHEAT-HRW",), "KE=F"),
    MarketSpec("ZO — Oats", "ZO", "Agricoli", "commodity", False, "OATS - CHICAGO BOARD OF TRADE", ("OATS",), "ZO=F"),
    MarketSpec("ZR — Rough Rice", "ZR", "Agricoli", "commodity", False, "ROUGH RICE - CHICAGO BOARD OF TRADE", ("ROUGH RICE",), "ZR=F"),

    # Soft commodity
    MarketSpec("CC — Cocoa", "CC", "Soft", "commodity", False, "COCOA - ICE FUTURES U.S.", ("COCOA",), "CC=F"),
    MarketSpec("KC — Coffee C", "KC", "Soft", "commodity", False, "COFFEE C - ICE FUTURES U.S.", ("COFFEE C",), "KC=F"),
    MarketSpec("CT — Cotton No. 2", "CT", "Soft", "commodity", False, "COTTON NO. 2 - ICE FUTURES U.S.", ("COTTON NO. 2",), "CT=F"),
    MarketSpec("SB — Sugar No. 11", "SB", "Soft", "commodity", False, "SUGAR NO. 11 - ICE FUTURES U.S.", ("SUGAR NO. 11",), "SB=F"),
    MarketSpec("OJ — Orange Juice", "OJ", "Soft", "commodity", False, "FRZN CONCENTRATED ORANGE JUICE - ICE FUTURES U.S.", ("ORANGE JUICE",), "OJ=F"),
    MarketSpec("LBR — Lumber", "LBR", "Soft", "commodity", False, "LUMBER - CHICAGO MERCANTILE EXCHANGE", ("LUMBER",), "LBR=F"),

    # Bestiame
    MarketSpec("LE — Live Cattle", "LE", "Bestiame", "commodity", False, "LIVE CATTLE - CHICAGO MERCANTILE EXCHANGE", ("LIVE CATTLE",), "LE=F"),
    MarketSpec("HE — Lean Hogs", "HE", "Bestiame", "commodity", False, "LEAN HOGS - CHICAGO MERCANTILE EXCHANGE", ("LEAN HOGS",), "HE=F"),
    MarketSpec("GF — Feeder Cattle", "GF", "Bestiame", "commodity", False, "FEEDER CATTLE - CHICAGO MERCANTILE EXCHANGE", ("FEEDER CATTLE",), "GF=F"),
)

MARKET_BY_LABEL = {m.label: m for m in MARKETS}


# =============================================================================
# FUNZIONI GENERALI
# =============================================================================
def normalize_text(value: str) -> str:
    value = value.upper().strip()
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def escape_soql(value: str) -> str:
    return value.replace("'", "''")


def to_float(value: Any) -> float:
    if value in (None, "", "."):
        return math.nan
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def fmt_number(value: float, signed: bool = False) -> str:
    if pd.isna(value):
        return "N/A"
    template = f"{value:+,.0f}" if signed else f"{value:,.0f}"
    return template.replace(",", ".")


def fmt_pct(value: float, signed: bool = False, digits: int = 1) -> str:
    if pd.isna(value):
        return "N/A"
    sign = "+" if signed and value > 0 else ""
    return f"{sign}{value:.{digits}f}%"


def fmt_decimal(value: float, digits: int = 1) -> str:
    return "N/A" if pd.isna(value) else f"{value:.{digits}f}"


def fmt_signed_decimal(value: float, digits: int = 1) -> str:
    return "N/A" if pd.isna(value) else f"{value:+.{digits}f}"


def card(kicker: str, title: str, detail: str, accent: str) -> None:
    st.markdown(
        f"""
        <div class="cot-card" style="--accent:{accent};">
            <div class="cot-kicker">{kicker}</div>
            <div class="cot-title">{title}</div>
            <div class="cot-small">{detail}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def accent_for_state(state: str) -> str:
    text = state.upper()
    if any(token in text for token in ("LONG CONFERMATO", "RIALZISTA", "POSSIBILE MINIMO", "MIGLIORAMENTO")):
        return "#15803D"
    if any(token in text for token in ("SHORT CONFERMATO", "RIBASSISTA", "POSSIBILE MASSIMO", "DETERIORAMENTO")):
        return "#B91C1C"
    if any(token in text for token in ("RITARDO", "NON INSEGUIRE", "DIVERGENZA", "PRESSIONE", "LIQUIDAZIONE", "NET POSITION VICINA")):
        return "#C2410C"
    return "#9A6700"


def safe_pct_change(current: float, previous: float) -> float:
    if pd.isna(current) or pd.isna(previous) or previous == 0:
        return math.nan
    return (current - previous) / previous * 100.0


def historical_percentile(series: pd.Series, lookback: int) -> float:
    values = pd.to_numeric(series, errors="coerce").dropna().tail(lookback)
    if len(values) < 2:
        return math.nan
    current = values.iloc[-1]
    history = values.iloc[:-1]
    return float((history <= current).mean() * 100.0)


def last_complete_friday(today: date | None = None) -> date:
    today = today or date.today()
    weekday = today.weekday()  # lun=0 ... dom=6
    if weekday == 4:  # venerdì: evitiamo sempre la settimana potenzialmente ancora aperta
        return today - timedelta(days=7)
    if weekday < 4:
        return today - timedelta(days=weekday + 3)
    return today - timedelta(days=weekday - 4)


# =============================================================================
# ACCESSO CFTC
# =============================================================================
class CFTCError(RuntimeError):
    pass


@st.cache_data(ttl=3600, show_spinner=False)
def cftc_request(dataset_id: str, params_items: tuple[tuple[str, str], ...]) -> list[dict[str, Any]]:
    params = dict(params_items)
    errors: list[str] = []
    for base in CFTC_API_BASES:
        url = f"{base}/{dataset_id}.json"
        try:
            response = requests.get(
                url,
                params=params,
                timeout=30,
                headers={"User-Agent": "Grespo-COT-Smart-Money/1.0"},
            )
            response.raise_for_status()
            payload = response.json()
            if not isinstance(payload, list):
                errors.append(f"{base}: risposta non tabellare")
                continue
            return payload
        except (requests.RequestException, ValueError) as exc:
            errors.append(f"{base}: {exc}")
    raise CFTCError(" | ".join(errors))


def run_cftc_query(dataset_id: str, params: dict[str, Any]) -> list[dict[str, Any]]:
    normalized = tuple(sorted((str(k), str(v)) for k, v in params.items()))
    return cftc_request(dataset_id, normalized)


def candidate_score(candidate: str, spec: MarketSpec) -> float:
    cand = normalize_text(candidate)
    preferred = normalize_text(spec.preferred_name)
    ratio = SequenceMatcher(None, cand, preferred).ratio()
    keyword_bonus = sum(0.18 for term in spec.search_terms if normalize_text(term) in cand)
    return ratio + keyword_bonus


@st.cache_data(ttl=21600, show_spinner=False)
def resolve_market_name(dataset_id: str, spec: MarketSpec) -> tuple[str, str]:
    exact_params = {
        "$select": "market_and_exchange_names",
        "$where": f"market_and_exchange_names='{escape_soql(spec.preferred_name)}'",
        "$limit": 1,
    }
    exact = run_cftc_query(dataset_id, exact_params)
    if exact:
        return str(exact[0]["market_and_exchange_names"]), "corrispondenza esatta"

    candidates: dict[str, None] = {}
    for term in spec.search_terms:
        params = {
            "$select": "market_and_exchange_names",
            "$where": f"market_and_exchange_names like '%{escape_soql(term.upper())}%'",
            "$limit": 100,
        }
        try:
            for row in run_cftc_query(dataset_id, params):
                name = row.get("market_and_exchange_names")
                if name:
                    candidates[str(name)] = None
        except CFTCError:
            continue

    if not candidates:
        raise CFTCError(
            f"Mercato non trovato nel dataset {dataset_id}: {spec.preferred_name}"
        )

    best = max(candidates, key=lambda name: candidate_score(name, spec))
    return best, "corrispondenza automatica"


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_market_history(
    report_type: str,
    spec: MarketSpec,
    limit: int,
) -> tuple[list[dict[str, Any]], str, str]:
    dataset_id = CFTC_DATASETS[report_type]
    market_name, resolution = resolve_market_name(dataset_id, spec)
    params = {
        "$where": f"market_and_exchange_names='{escape_soql(market_name)}'",
        "$order": "report_date_as_yyyy_mm_dd DESC",
        "$limit": limit,
    }
    rows = run_cftc_query(dataset_id, params)
    return rows, market_name, resolution


# =============================================================================
# NORMALIZZAZIONE CAMPI CFTC
# =============================================================================
def normalized_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def resolve_field(records: Iterable[dict[str, Any]], candidates: Iterable[str]) -> str | None:
    keys: dict[str, str] = {}
    for row in list(records)[:10]:
        for key in row:
            keys.setdefault(normalized_key(key), key)

    for candidate in candidates:
        if candidate in keys.values():
            return candidate
        normalized = normalized_key(candidate)
        if normalized in keys:
            return keys[normalized]
    return None


COMMON_FIELDS = {
    "date": ("report_date_as_yyyy_mm_dd", "as_of_date_form_yyyy_mm_dd"),
    "oi": ("open_interest_all",),
    "code": ("cftc_contract_market_code",),
    "commodity": ("commodity_name", "contract_market_name"),
    "small_long": (
        "nonrept_positions_long",
        "nonrept_positions_long_all",
        "nonreportable_positions_long",
        "nonreportable_positions_long_all",
    ),
    "small_short": (
        "nonrept_positions_short",
        "nonrept_positions_short_all",
        "nonreportable_positions_short",
        "nonreportable_positions_short_all",
    ),
    "conc_long": (
        "conc_net_le_8_tdr_long",
        "conc_net_le_8_tdr_long_all",
        "conc_net_lt_8_tdr_long_all",
        "conc_net_long_all_8",
    ),
    "conc_short": (
        "conc_net_le_8_tdr_short",
        "conc_net_le_8_tdr_short_all",
        "conc_net_lt_8_tdr_short_all",
        "conc_net_short_all_8",
    ),
}

REPORT_FIELDS = {
    "Financial": {
        "trend_long": ("lev_money_positions_long", "lev_money_positions_long_all"),
        "trend_short": ("lev_money_positions_short", "lev_money_positions_short_all"),
        "asset_long": ("asset_mgr_positions_long", "asset_mgr_positions_long_all"),
        "asset_short": ("asset_mgr_positions_short", "asset_mgr_positions_short_all"),
        "dealer_long": ("dealer_positions_long", "dealer_positions_long_all"),
        "dealer_short": ("dealer_positions_short", "dealer_positions_short_all"),
        "other_long": (
            "other_rept_positions_long", "other_rept_positions_long_all",
            "other_reportable_positions_long", "other_reportable_positions_long_all",
        ),
        "other_short": (
            "other_rept_positions_short", "other_rept_positions_short_all",
            "other_reportable_positions_short", "other_reportable_positions_short_all",
        ),
    },
    "Disaggregated": {
        "trend_long": ("m_money_positions_long", "m_money_positions_long_all"),
        "trend_short": ("m_money_positions_short", "m_money_positions_short_all"),
        "producer_long": ("prod_merc_positions_long", "prod_merc_positions_long_all"),
        "producer_short": ("prod_merc_positions_short", "prod_merc_positions_short_all"),
        "swap_long": (
            "swap_positions_long", "swap_positions_long_all",
            "swap__positions_long_all",
        ),
        "swap_short": (
            "swap_positions_short", "swap_positions_short_all",
            "swap__positions_short_all",
        ),
        "other_long": (
            "other_rept_positions_long", "other_rept_positions_long_all",
            "other_reportable_positions_long", "other_reportable_positions_long_all",
        ),
        "other_short": (
            "other_rept_positions_short", "other_rept_positions_short_all",
            "other_reportable_positions_short", "other_reportable_positions_short_all",
        ),
    },
}


def build_history_df(
    records: list[dict[str, Any]],
    report_type: str,
    spec: MarketSpec,
    cot_lookback: int,
) -> tuple[pd.DataFrame, dict[str, str | None]]:
    if not records:
        return pd.DataFrame(), {}

    fields: dict[str, str | None] = {
        name: resolve_field(records, candidates) for name, candidates in COMMON_FIELDS.items()
    }
    fields.update(
        {
            name: resolve_field(records, candidates)
            for name, candidates in REPORT_FIELDS[report_type].items()
        }
    )

    if report_type == "Financial":
        if spec.is_fx:
            fields["counter_long"] = fields.get("dealer_long")
            fields["counter_short"] = fields.get("dealer_short")
        else:
            fields["counter_long"] = fields.get("asset_long")
            fields["counter_short"] = fields.get("asset_short")
    elif report_type == "Disaggregated":
        fields["counter_long"] = fields.get("producer_long")
        fields["counter_short"] = fields.get("producer_short")
    else:
        raise CFTCError(f"Tipo di report non supportato: {report_type}")

    required = ("date", "oi", "trend_long", "trend_short", "counter_long", "counter_short")
    missing = [name for name in required if not fields.get(name)]
    if missing:
        available = sorted(records[0].keys())
        raise CFTCError(
            "Campi CFTC mancanti: " + ", ".join(missing) +
            ". Campi disponibili: " + ", ".join(available[:40])
        )

    def optional_value(record: dict[str, Any], field_name: str) -> float:
        key = fields.get(field_name)
        return to_float(record.get(key or "")) if key else math.nan

    rows: list[dict[str, Any]] = []
    for record in records:
        parsed_date = pd.to_datetime(record.get(fields["date"] or ""), errors="coerce")
        if pd.isna(parsed_date):
            continue
        rows.append(
            {
                "date": parsed_date.tz_localize(None) if getattr(parsed_date, "tzinfo", None) else parsed_date,
                "oi": to_float(record.get(fields["oi"] or "")),
                "trend_long": to_float(record.get(fields["trend_long"] or "")),
                "trend_short": to_float(record.get(fields["trend_short"] or "")),
                "counter_long": to_float(record.get(fields["counter_long"] or "")),
                "counter_short": to_float(record.get(fields["counter_short"] or "")),
                "small_long": optional_value(record, "small_long"),
                "small_short": optional_value(record, "small_short"),
                "conc_long": optional_value(record, "conc_long"),
                "conc_short": optional_value(record, "conc_short"),
                "asset_long": optional_value(record, "asset_long"),
                "asset_short": optional_value(record, "asset_short"),
                "dealer_long": optional_value(record, "dealer_long"),
                "dealer_short": optional_value(record, "dealer_short"),
                "producer_long": optional_value(record, "producer_long"),
                "producer_short": optional_value(record, "producer_short"),
                "swap_long": optional_value(record, "swap_long"),
                "swap_short": optional_value(record, "swap_short"),
                "other_long": optional_value(record, "other_long"),
                "other_short": optional_value(record, "other_short"),
                "cftc_code": str(record.get(fields["code"] or "", "N/A")),
            }
        )

    df = pd.DataFrame(rows)
    if df.empty:
        return df, fields

    df = (
        df.sort_values("date")
        .drop_duplicates(subset=["date"], keep="last")
        .reset_index(drop=True)
    )
    numeric_columns = (
        "oi", "trend_long", "trend_short", "counter_long", "counter_short",
        "small_long", "small_short", "conc_long", "conc_short",
        "asset_long", "asset_short", "dealer_long", "dealer_short",
        "producer_long", "producer_short", "swap_long", "swap_short",
        "other_long", "other_short",
    )
    for column in numeric_columns:
        if column not in df:
            df[column] = math.nan
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df["trend_net"] = df["trend_long"] - df["trend_short"]
    df["counter_net"] = df["counter_long"] - df["counter_short"]
    df["small_net"] = df["small_long"] - df["small_short"]
    df["asset_net"] = df["asset_long"] - df["asset_short"]
    df["dealer_net"] = df["dealer_long"] - df["dealer_short"]
    df["producer_net"] = df["producer_long"] - df["producer_short"]
    df["swap_net"] = df["swap_long"] - df["swap_short"]
    df["other_net"] = df["other_long"] - df["other_short"]

    def rolling_cot_index(series: pd.Series, index_lookback: int) -> pd.Series:
        min_periods = min(COT_INDEX_SHORT_LOOKBACK, index_lookback)
        rolling_min = series.rolling(index_lookback, min_periods=min_periods).min()
        rolling_max = series.rolling(index_lookback, min_periods=min_periods).max()
        denominator = rolling_max - rolling_min
        result = pd.Series(
            np.where(
                denominator.ne(0),
                100.0 * (series - rolling_min) / denominator,
                50.0,
            ),
            index=series.index,
            dtype="float64",
        )
        return result.clip(lower=0.0, upper=100.0)

    # Il motore continua a usare il lookback scelto dall'utente.
    df["trend_index"] = rolling_cot_index(df["trend_net"], cot_lookback)
    df["counter_index"] = rolling_cot_index(df["counter_net"], cot_lookback)
    df["small_index"] = rolling_cot_index(df["small_net"], cot_lookback)
    df["cot_index"] = df["trend_index"]

    # Due orizzonti informativi fissi, identici a TradingView V1.5.36.
    df["trend_index_26w"] = rolling_cot_index(df["trend_net"], COT_INDEX_SHORT_LOOKBACK)
    df["trend_index_156w"] = rolling_cot_index(df["trend_net"], COT_INDEX_LONG_LOOKBACK)
    # TradingView V1.5.36 usa un Alignment contrarian fisso a 156W per tutte
    # le famiglie: categoria trend + controparte + Small Traders.
    df["counter_index_156w"] = rolling_cot_index(df["counter_net"], COT_INDEX_LONG_LOOKBACK)
    df["small_index_156w"] = rolling_cot_index(df["small_net"], COT_INDEX_LONG_LOOKBACK)
    df["cot_index_26w"] = df["trend_index_26w"]
    df["cot_index_156w"] = df["trend_index_156w"]

    def new_net_extreme_flag(series: pd.Series, horizon: int) -> pd.Series:
        # Pine confronta il valore corrente con i precedenti horizon-1 report.
        previous = series.shift(1)
        prior_high = previous.rolling(horizon - 1, min_periods=horizon - 1).max()
        prior_low = previous.rolling(horizon - 1, min_periods=horizon - 1).min()
        return pd.Series(
            np.select(
                [series.gt(prior_high), series.lt(prior_low)],
                [1.0, -1.0],
                default=0.0,
            ),
            index=series.index,
            dtype="float64",
        ).where(prior_high.notna() & prior_low.notna())

    df["trend_extreme_26w_flag"] = new_net_extreme_flag(df["trend_net"], COT_INDEX_SHORT_LOOKBACK)
    df["trend_extreme_156w_flag"] = new_net_extreme_flag(df["trend_net"], COT_INDEX_LONG_LOOKBACK)

    # Rapid Shift: variazione del COT Index della controparte in sei veri report.
    df["counter_rapid_shift_6w"] = df["counter_index"] - df["counter_index"].shift(6)

    # OI Index 52W: posizione corrente dell'Open Interest nel range annuale.
    oi_min_52 = df["oi"].rolling(OI_INDEX_LOOKBACK, min_periods=26).min()
    oi_max_52 = df["oi"].rolling(OI_INDEX_LOOKBACK, min_periods=26).max()
    oi_denominator = oi_max_52 - oi_min_52
    df["oi_index_52w"] = pd.Series(
        np.where(
            oi_denominator.ne(0),
            100.0 * (df["oi"] - oi_min_52) / oi_denominator,
            50.0,
        ),
        index=df.index,
        dtype="float64",
    ).clip(lower=0.0, upper=100.0)
    return df, fields


# =============================================================================
# PREZZO WEEKLY
# =============================================================================
@st.cache_data(ttl=21600, show_spinner=False)
def fetch_weekly_price(ticker: str) -> tuple[pd.DataFrame, str | None]:
    if not ticker.strip():
        return pd.DataFrame(), "Ticker Yahoo non impostato."
    try:
        daily = yf.download(
            ticker.strip(),
            period="5y",
            interval="1d",
            auto_adjust=False,
            progress=False,
            threads=False,
        )
    except Exception as exc:
        return pd.DataFrame(), f"Errore Yahoo Finance: {exc}"

    if daily is None or daily.empty:
        return pd.DataFrame(), "Yahoo Finance non ha restituito dati."

    if isinstance(daily.columns, pd.MultiIndex):
        close_candidates = [col for col in daily.columns if col[0] == "Close"]
        if not close_candidates:
            return pd.DataFrame(), "Colonna Close non trovata nei dati Yahoo."
        close = daily[close_candidates[0]].copy()
    elif "Close" in daily.columns:
        close = daily["Close"].copy()
    else:
        return pd.DataFrame(), "Colonna Close non trovata nei dati Yahoo."

    close.index = pd.to_datetime(close.index).tz_localize(None)
    close = pd.to_numeric(close, errors="coerce").dropna()
    weekly = close.resample("W-FRI").last().to_frame("close")
    cutoff = pd.Timestamp(last_complete_friday())
    weekly = weekly.loc[weekly.index <= cutoff].copy()
    weekly["ema21"] = weekly["close"].ewm(span=21, adjust=False).mean()
    return weekly.tail(180), None


def analyze_price(weekly: pd.DataFrame) -> dict[str, Any]:
    unavailable = {
        "available": False,
        "text": "PREZZO WEEKLY NON DISPONIBILE",
        "detail": "Manca una serie Weekly valida per la conferma del prezzo.",
        "long_confirmed": False,
        "short_confirmed": False,
        "above_ema": False,
        "below_ema": False,
        "rising": False,
        "falling": False,
        "close": math.nan,
        "previous": math.nan,
        "ema21": math.nan,
    }
    if weekly.empty or len(weekly) < 22:
        return unavailable

    current = weekly.iloc[-1]
    previous = weekly.iloc[-2]
    close = float(current["close"])
    prev_close = float(previous["close"])
    ema21 = float(current["ema21"])
    above = close > ema21
    below = close < ema21
    rising = close > prev_close
    falling = close < prev_close
    long_confirmed = above and rising
    short_confirmed = below and falling

    if long_confirmed:
        text = "CONFERMA RIALZISTA"
    elif short_confirmed:
        text = "CONFERMA RIBASSISTA"
    elif above and falling:
        text = "SOPRA EMA21, MOMENTUM IN CALO"
    elif below and rising:
        text = "SOTTO EMA21, RIMBALZO IN CORSO"
    elif above:
        text = "PREZZO SOPRA EMA21"
    elif below:
        text = "PREZZO SOTTO EMA21"
    else:
        text = "PREZZO NEUTRALE SULLA EMA21"

    return {
        "available": True,
        "text": text,
        "detail": f"Close W: {close:.4f} | EMA21 W: {ema21:.4f} | Close precedente: {prev_close:.4f}",
        "long_confirmed": long_confirmed,
        "short_confirmed": short_confirmed,
        "above_ema": above,
        "below_ema": below,
        "rising": rising,
        "falling": falling,
        "close": close,
        "previous": prev_close,
        "ema21": ema21,
    }


# =============================================================================
# MOTORE SMART MONEY
# =============================================================================
def cot_zone(index_value: float) -> str:
    """Collocazione della Net Position nel range storico, come TradingView V1.5.36."""
    if pd.isna(index_value):
        return "NON DISPONIBILE"
    if index_value >= 80:
        return "VICINO AL MASSIMO DEL RANGE STORICO"
    if index_value >= 60:
        return "FASCIA ALTA DEL RANGE STORICO"
    if index_value > 40:
        return "FASCIA CENTRALE DEL RANGE STORICO"
    if index_value > 20:
        return "FASCIA BASSA DEL RANGE STORICO"
    return "VICINO AL MINIMO DEL RANGE STORICO"


def directional_exposure(long_positions: float, short_positions: float) -> tuple[float, float]:
    """Quota Long/Short sulle sole posizioni direzionali; Spreading escluso."""
    if pd.isna(long_positions) or pd.isna(short_positions):
        return math.nan, math.nan
    total = float(long_positions) + float(short_positions)
    if total <= 0:
        return math.nan, math.nan
    return 100.0 * float(long_positions) / total, 100.0 * float(short_positions) / total


def rapid_shift_state(value: float) -> str:
    """Classifica la velocità del cambiamento del COT Index della controparte."""
    if pd.isna(value):
        return "NON DISPONIBILE"
    if value >= RAPID_SHIFT_EXTREME:
        return "RAPIDO RIALZISTA"
    if value >= RAPID_SHIFT_WARNING:
        return "IN ACCELERAZIONE"
    if value <= -RAPID_SHIFT_EXTREME:
        return "RAPIDO RIBASSISTA"
    if value <= -RAPID_SHIFT_WARNING:
        return "IN PEGGIORAMENTO"
    return "NORMALE"


def oi_index_state(value: float) -> str:
    """Descrive il livello dell'Open Interest rispetto al range degli ultimi 52 report."""
    if pd.isna(value):
        return "NON DISPONIBILE"
    if value >= 80.0:
        return "PARTECIPAZIONE MOLTO ALTA"
    if value >= 60.0:
        return "PARTECIPAZIONE ALTA"
    if value > 40.0:
        return "PARTECIPAZIONE NELLA MEDIA"
    if value > 20.0:
        return "PARTECIPAZIONE BASSA"
    return "PARTECIPAZIONE MOLTO BASSA"


def analyze_smart_money(
    df: pd.DataFrame,
    spec: MarketSpec,
    price: dict[str, Any],
    oi_threshold: float,
    cot_lookback: int,
    report_mode: str = "Compatto",
    show_all_category_flows: bool = False,
) -> dict[str, Any]:
    if df.empty or len(df) < 7:
        return {
            "available": False,
            "final_bias": "DATI NON DISPONIBILI",
            "final_detail": "I dati COT necessari non sono disponibili per questo mercato. L'indicatore non può costruire una lettura affidabile.",
            "simple_title": "DATI NON DISPONIBILI",
            "simple_detail": "I dati COT necessari non sono disponibili per questo mercato. L'indicatore non può costruire una lettura affidabile.",
            "plain_action": "NON USARE QUESTA LETTURA. VERIFICA CHE IL FUTURE DISPONGA DEI DATI COT.",
            "explanation": "Mancano una o più serie COT necessarie per confrontare i Fondi con gli altri operatori del mercato.",
            "action": "NON USARE QUESTA LETTURA. VERIFICA CHE IL FUTURE DISPONGA DEI DATI COT.",
            "reason": "Mancano una o più serie COT necessarie per confrontare i Fondi con gli altri operatori del mercato.",
            "last_report": "DATI NON DISPONIBILI",
            "structure": "DATI DELLE ULTIME 3-6 SETTIMANE NON DISPONIBILI",
            "positioning": "NON DISPONIBILE",
            "positioning_value": "COT Index 26W: N/A / 100\nCOT Index 156W: N/A / 100",
            "current_position": "DATI NON DISPONIBILI",
            "current_position_value": "Long: N/A | Short: N/A",
            "oi_quality": "OI NON DISPONIBILE",
            "concentration_state": "DATI NON DISPONIBILI",
        }

    cur = df.iloc[-1]
    prev = df.iloc[-2]
    w3 = df.iloc[-4]
    w6 = df.iloc[-7]

    trend_flow_1w = float(cur["trend_net"] - prev["trend_net"])
    counter_flow_1w = float(cur["counter_net"] - prev["counter_net"])
    trend_flow_3w = float(cur["trend_net"] - w3["trend_net"])
    counter_flow_3w = float(cur["counter_net"] - w3["counter_net"])
    trend_flow_6w = float(cur["trend_net"] - w6["trend_net"])
    counter_flow_6w = float(cur["counter_net"] - w6["counter_net"])

    trend_chg_l = float(cur["trend_long"] - prev["trend_long"])
    trend_chg_s = float(cur["trend_short"] - prev["trend_short"])
    counter_chg_l = float(cur["counter_long"] - prev["counter_long"])
    counter_chg_s = float(cur["counter_short"] - prev["counter_short"])

    def net_flow_1w(column: str) -> float:
        if column not in df.columns or pd.isna(cur.get(column, math.nan)) or pd.isna(prev.get(column, math.nan)):
            return math.nan
        return float(cur[column] - prev[column])

    asset_manager_flow_1w = net_flow_1w("asset_net")
    dealer_flow_1w = net_flow_1w("dealer_net")
    producer_flow_1w = net_flow_1w("producer_net")
    swap_flow_1w = net_flow_1w("swap_net")
    other_flow_1w = net_flow_1w("other_net")
    nonreportable_flow_1w = net_flow_1w("small_net")

    pct_delta_oi = safe_pct_change(float(cur["oi"]), float(prev["oi"]))
    pct_oi_3w = safe_pct_change(float(cur["oi"]), float(w3["oi"]))
    pct_oi_6w = safe_pct_change(float(cur["oi"]), float(w6["oi"]))
    cot_index = float(cur["cot_index"]) if not pd.isna(cur["cot_index"]) else math.nan
    cot_index_26w = float(cur["cot_index_26w"]) if "cot_index_26w" in cur and not pd.isna(cur["cot_index_26w"]) else math.nan
    cot_index_156w = float(cur["cot_index_156w"]) if "cot_index_156w" in cur and not pd.isna(cur["cot_index_156w"]) else math.nan
    extreme_26w_flag = float(cur["trend_extreme_26w_flag"]) if "trend_extreme_26w_flag" in cur and not pd.isna(cur["trend_extreme_26w_flag"]) else math.nan
    extreme_156w_flag = float(cur["trend_extreme_156w_flag"]) if "trend_extreme_156w_flag" in cur and not pd.isna(cur["trend_extreme_156w_flag"]) else math.nan

    new_max_net_26w = not pd.isna(extreme_26w_flag) and extreme_26w_flag > 0.5
    new_min_net_26w = not pd.isna(extreme_26w_flag) and extreme_26w_flag < -0.5
    new_max_net_156w = not pd.isna(extreme_156w_flag) and extreme_156w_flag > 0.5
    new_min_net_156w = not pd.isna(extreme_156w_flag) and extreme_156w_flag < -0.5
    extreme_labels = [
        label for condition, label in (
            (new_max_net_26w, "NUOVO MASSIMO NET 26W"),
            (new_min_net_26w, "NUOVO MINIMO NET 26W"),
            (new_max_net_156w, "NUOVO MASSIMO NET 156W"),
            (new_min_net_156w, "NUOVO MINIMO NET 156W"),
        ) if condition
    ]
    extreme_horizons = "\n".join(extreme_labels)
    positioning = cot_zone(cot_index_156w)
    positioning_value = (
        f"COT Index 26W: {fmt_decimal(cot_index_26w, 1)} / 100\n"
        f"COT Index 156W: {fmt_decimal(cot_index_156w, 1)} / 100"
        + (f"\n{extreme_horizons}" if extreme_horizons else "")
    )

    directional_long_pct, directional_short_pct = directional_exposure(
        float(cur["trend_long"]), float(cur["trend_short"])
    )
    position_subject = "MANAGED MONEY" if spec.family == "commodity" else "LEVERAGED FUNDS"
    if pd.isna(directional_long_pct) or pd.isna(directional_short_pct):
        current_position = "DATI NON DISPONIBILI"
    elif directional_long_pct >= 55.0:
        current_position = f"I {position_subject} SONO PREVALENTEMENTE LONG"
    elif directional_short_pct >= 55.0:
        current_position = f"I {position_subject} SONO PREVALENTEMENTE SHORT"
    else:
        current_position = f"I {position_subject} HANNO UNA POSIZIONE BILANCIATA"
    current_position_value = (
        f"Long: {fmt_pct(directional_long_pct)} | Short: {fmt_pct(directional_short_pct)}"
    )

    oi_index_52w = float(cur["oi_index_52w"]) if "oi_index_52w" in cur and not pd.isna(cur["oi_index_52w"]) else math.nan
    oi_index_52w_state = oi_index_state(oi_index_52w)

    # TradingView V1.5.36: prima identifica la direzione uniforme dei flussi 3-6W,
    # poi verifica se l'Open Interest sostiene, perde partecipazione o resta stabile.
    oi_macro_available = not pd.isna(pct_oi_3w) and not pd.isna(pct_oi_6w)
    oi_macro_up = oi_macro_available and pct_oi_3w > oi_threshold and pct_oi_6w > oi_threshold
    oi_macro_down = oi_macro_available and pct_oi_3w < -oi_threshold and pct_oi_6w < -oi_threshold
    oi_macro_stable = (
        oi_macro_available
        and abs(pct_oi_3w) <= oi_threshold
        and abs(pct_oi_6w) <= oi_threshold
    )

    oi_direction_available = not pd.isna(trend_flow_3w) and not pd.isna(trend_flow_6w)
    oi_direction_long = oi_direction_available and trend_flow_3w > 0 and trend_flow_6w > 0
    oi_direction_short = oi_direction_available and trend_flow_3w < 0 and trend_flow_6w < 0
    oi_direction = "RIALZISTA" if oi_direction_long else "RIBASSISTA" if oi_direction_short else ""

    if not oi_direction_available:
        oi_quality = "DIREZIONE 3-6W NON DISPONIBILE"
    elif not oi_direction_long and not oi_direction_short:
        oi_quality = "IL QUADRO 3-6W NON HA UNA DIREZIONE UNIFORME"
    elif oi_macro_up:
        oi_quality = f"IL MOVIMENTO {oi_direction} È SOSTENUTO"
    elif oi_macro_down:
        oi_quality = f"IL MOVIMENTO {oi_direction} PERDE PARTECIPAZIONE"
    elif oi_macro_stable:
        oi_quality = f"IL MOVIMENTO {oi_direction} HA UNA PARTECIPAZIONE STABILE"
    elif oi_macro_available:
        oi_quality = f"IL MOVIMENTO {oi_direction} NON È ANCORA CONFERMATO"
    else:
        oi_quality = "OI NON DISPONIBILE"

    oi_up = not pd.isna(pct_delta_oi) and pct_delta_oi > oi_threshold
    oi_down = not pd.isna(pct_delta_oi) and pct_delta_oi < -oi_threshold

    new_long = trend_flow_1w > 0 and trend_chg_l > 0 and oi_up
    new_short = trend_flow_1w < 0 and trend_chg_s > 0 and oi_up
    short_covering = trend_flow_1w > 0 and trend_chg_s < 0 and oi_down
    long_liquidation = trend_flow_1w < 0 and trend_chg_l < 0 and oi_down

    # TESTO SEMPLICE ULTIMO REPORT — identico a TradingView V1.5.36
    if new_long:
        last_report = "I FONDI HANNO APERTO NUOVI LONG"
    elif new_short:
        last_report = "I FONDI HANNO APERTO NUOVI SHORT"
    elif short_covering:
        last_report = "I FONDI HANNO RIDOTTO GLI SHORT"
    elif long_liquidation:
        last_report = "I FONDI HANNO RIDOTTO I LONG"
    elif trend_flow_1w > 0:
        last_report = "I FONDI HANNO COMPRATO"
    elif trend_flow_1w < 0:
        last_report = "I FONDI HANNO VENDUTO"
    else:
        last_report = "I FONDI NON HANNO CAMBIATO DIREZIONE IN MODO CHIARO"

    macro_long = trend_flow_3w > 0 and trend_flow_6w > 0
    macro_short = trend_flow_3w < 0 and trend_flow_6w < 0
    macro_recovery = trend_flow_3w > 0 and trend_flow_6w < 0
    macro_deterioration = trend_flow_3w < 0 and trend_flow_6w > 0

    def _macro_direction_text(value: float) -> str:
        if pd.isna(value):
            return "NON DISPONIBILE"
        if value > 0:
            return "VERSO IL RIALZO"
        if value < 0:
            return "VERSO IL RIBASSO"
        return "SOSTANZIALMENTE STABILE"

    structure_3w = _macro_direction_text(trend_flow_3w)
    structure_6w = _macro_direction_text(trend_flow_6w)
    structure = (
        f"ULTIME 3 SETTIMANE: {structure_3w}\n"
        f"ULTIME 6 SETTIMANE: {structure_6w}"
    )

    counter_macro_long = counter_flow_3w > 0 and counter_flow_6w > 0
    counter_macro_short = counter_flow_3w < 0 and counter_flow_6w < 0

    if spec.family == "commodity":
        if counter_flow_1w > 0:
            counter_reading = "PRODUCER MIGLIORANO LA NET POSITION"
        elif counter_flow_1w < 0:
            counter_reading = "PRODUCER AUMENTANO L'HEDGING SHORT"
        else:
            counter_reading = "PRODUCER SOSTANZIALMENTE STABILI"
    elif spec.is_fx:
        if trend_flow_1w > 0 and counter_flow_1w < 0:
            counter_reading = "DEALER CONTROPARTE DEI LONG"
        elif trend_flow_1w < 0 and counter_flow_1w > 0:
            counter_reading = "DEALER CONTROPARTE DEGLI SHORT"
        else:
            counter_reading = "DINAMICA DEALER NON CONCORDE"
    else:
        if counter_flow_1w > 0:
            counter_reading = "ASSET MANAGER AUMENTANO L'ESPOSIZIONE"
        elif counter_flow_1w < 0:
            counter_reading = "ASSET MANAGER RIDUCONO L'ESPOSIZIONE"
        else:
            counter_reading = "ASSET MANAGER STABILI"

    financial_long_alignment = spec.family == "financial" and not spec.is_fx and trend_flow_1w > 0 and counter_flow_1w > 0
    financial_short_alignment = spec.family == "financial" and not spec.is_fx and trend_flow_1w < 0 and counter_flow_1w < 0
    fx_long_alignment = spec.is_fx and trend_flow_1w > 0 and counter_flow_1w < 0
    fx_short_alignment = spec.is_fx and trend_flow_1w < 0 and counter_flow_1w > 0
    commodity_joint_long = spec.family == "commodity" and trend_flow_1w > 0 and counter_flow_1w > 0
    commodity_joint_short = spec.family == "commodity" and trend_flow_1w < 0 and counter_flow_1w < 0
    commodity_trend_long = spec.family == "commodity" and trend_flow_1w > 0 and counter_flow_1w < 0
    commodity_trend_short = spec.family == "commodity" and trend_flow_1w < 0 and counter_flow_1w > 0

    financial_long_confirmed = financial_long_alignment and macro_long and counter_macro_long
    financial_short_confirmed = financial_short_alignment and macro_short and counter_macro_short
    fx_long_confirmed = fx_long_alignment and macro_long and counter_macro_short
    fx_short_confirmed = fx_short_alignment and macro_short and counter_macro_long

    possible_bottom = spec.family == "commodity" and cot_index <= 20 and commodity_joint_long
    possible_top = spec.family == "commodity" and cot_index >= 80 and commodity_joint_short

    if possible_bottom:
        engine_bias = "POSSIBILE MINIMO ISTITUZIONALE"
        engine_detail = "Net Position dei Managed Money vicina al minimo del range storico e in miglioramento insieme ai Producer."
    elif possible_top:
        engine_bias = "POSSIBILE MASSIMO ISTITUZIONALE"
        engine_detail = "Net Position dei Managed Money vicina al massimo del range storico e flussi in peggioramento insieme ai Producer."
    elif spec.family == "commodity" and commodity_trend_long and macro_long:
        engine_bias = "TREND LONG SPECULATIVO"
        engine_detail = "Managed Money accumula mentre i Producer aumentano le coperture."
    elif spec.family == "commodity" and commodity_trend_short and macro_short:
        engine_bias = "TREND SHORT SPECULATIVO"
        engine_detail = "Managed Money aumenta la pressione ribassista mentre i Producer migliorano la posizione."
    elif spec.family == "commodity" and commodity_joint_long:
        engine_bias = "RECUPERO LONG DA CONFERMARE"
        engine_detail = "Managed Money e Producer migliorano, ma la struttura 3-6W non è ancora pienamente rialzista."
    elif spec.family == "commodity" and commodity_joint_short:
        engine_bias = "PRESSIONE SHORT IN AUMENTO"
        engine_detail = "Managed Money e Producer peggiorano insieme."
    elif fx_long_confirmed:
        engine_bias = "LONG VALUTA CONFERMATO"
        engine_detail = "Leveraged Funds rialzisti e Dealer nel normale ruolo di controparte."
    elif fx_short_confirmed:
        engine_bias = "SHORT VALUTA CONFERMATO"
        engine_detail = "Leveraged Funds ribassisti e Dealer nel normale ruolo di controparte."
    elif spec.is_fx and fx_long_alignment and macro_long:
        engine_bias = "LONG FX CON CONFERMA PARZIALE"
        engine_detail = "Struttura Leveraged rialzista, ma Dealer non ancora concordi sulle 3-6W."
    elif spec.is_fx and fx_short_alignment and macro_short:
        engine_bias = "SHORT FX CON CONFERMA PARZIALE"
        engine_detail = "Struttura Leveraged ribassista, ma Dealer non ancora concordi sulle 3-6W."
    elif spec.is_fx and fx_long_alignment:
        engine_bias = "RECUPERO LONG VALUTA"
        engine_detail = "Flusso settimanale rialzista ancora da confermare sulle 3-6W."
    elif spec.is_fx and fx_short_alignment:
        engine_bias = "PRESSIONE SHORT VALUTA"
        engine_detail = "Flusso settimanale ribassista ancora da confermare sulle 3-6W."
    elif financial_long_confirmed:
        engine_bias = "LONG ISTITUZIONALE CONFERMATO"
        engine_detail = "Leveraged Funds e Asset Manager migliorano insieme anche nella struttura 3-6W."
    elif financial_short_confirmed:
        engine_bias = "SHORT ISTITUZIONALE CONFERMATO"
        engine_detail = "Leveraged Funds e Asset Manager peggiorano insieme anche nella struttura 3-6W."
    elif financial_long_alignment and macro_long:
        engine_bias = "LONG LEVERAGED CON CONFERMA PARZIALE"
        engine_detail = "Leveraged rialzisti; Asset Manager non ancora concordi sulle 3-6W."
    elif financial_short_alignment and macro_short:
        engine_bias = "SHORT LEVERAGED CON CONFERMA PARZIALE"
        engine_detail = "Leveraged ribassisti; Asset Manager non ancora concordi sulle 3-6W."
    elif financial_long_alignment:
        engine_bias = "RECUPERO LONG ISTITUZIONALE"
        engine_detail = "Leveraged Funds e Asset Manager migliorano, ma manca conferma 3-6W."
    elif financial_short_alignment:
        engine_bias = "PRESSIONE SHORT ISTITUZIONALE"
        engine_detail = "Leveraged Funds e Asset Manager peggiorano, ma manca conferma 3-6W."
    elif trend_flow_1w > 0:
        engine_bias = "MIGLIORAMENTO SPECULATIVO"
        engine_detail = f"La Net Position dei {spec.trend_label} migliora, ma il quadro non è completo."
    elif trend_flow_1w < 0:
        engine_bias = "DETERIORAMENTO SPECULATIVO"
        engine_detail = f"La Net Position dei {spec.trend_label} peggiora, ma il quadro non è completo."
    else:
        engine_bias = "BIAS MISTO / TRANSIZIONE"
        engine_detail = "I flussi non mostrano ancora una direzione abbastanza chiara."

    confirmed_long = (spec.family == "commodity" and commodity_trend_long and macro_long) or fx_long_confirmed or financial_long_confirmed
    confirmed_short = (spec.family == "commodity" and commodity_trend_short and macro_short) or fx_short_confirmed or financial_short_confirmed
    extreme_long = not pd.isna(cot_index) and cot_index >= 80
    extreme_short = not pd.isna(cot_index) and cot_index <= 20
    crowded_long = confirmed_long and extreme_long
    crowded_short = confirmed_short and extreme_short

    full_long = confirmed_long and price["long_confirmed"]
    full_short = confirmed_short and price["short_confirmed"]
    long_price_divergence = confirmed_long and price["short_confirmed"]
    short_price_divergence = confirmed_short and price["long_confirmed"]

    partial_long = (not confirmed_long and trend_flow_1w > 0 and (commodity_joint_long or fx_long_alignment or financial_long_alignment or macro_recovery))
    partial_short = (not confirmed_short and trend_flow_1w < 0 and (commodity_joint_short or fx_short_alignment or financial_short_alignment or macro_deterioration))
    partial_long_with_price = partial_long and price["long_confirmed"]
    partial_short_with_price = partial_short and price["short_confirmed"]

    # Configurazioni parziali già estese sul COT Index del motore.
    # Servono soltanto alla gerarchia testuale di "Cosa fare", come nel Pine V1.5.36.
    partial_crowded_long = partial_long and extreme_long
    partial_crowded_short = partial_short and extreme_short

    # =========================================================================
    # POSSIBILE FUTURO CAMBIO DI REGIME — ALIGNMENT CONTRARIAN FISSO 156W
    # TradingView V1.5.36 unifica la regola per Commodity, FX e altri Financial:
    # Bull = Trend basso + Controparte alta + Small basso.
    # Bear = Trend alto + Controparte bassa + Small alto.
    # =========================================================================
    alignment_trend_label = spec.trend_label
    alignment_counter_label = (
        "Producer/Merchant" if spec.family == "commodity"
        else "Dealer" if spec.is_fx
        else "Asset Manager"
    )
    alignment_counter_detail_label = spec.counter_label
    alignment_trend_index_156w = cot_index_156w
    alignment_counter_index_156w = (
        float(cur["counter_index_156w"])
        if "counter_index_156w" in cur and not pd.isna(cur["counter_index_156w"]) else math.nan
    )
    alignment_small_index_156w = (
        float(cur["small_index_156w"])
        if "small_index_156w" in cur and not pd.isna(cur["small_index_156w"]) else math.nan
    )
    alignment_available = not any(
        pd.isna(value)
        for value in (alignment_trend_index_156w, alignment_counter_index_156w, alignment_small_index_156w)
    )
    alignment_bull_trend = alignment_available and alignment_trend_index_156w <= ALIGNMENT_LOWER
    alignment_bull_counter = alignment_available and alignment_counter_index_156w >= ALIGNMENT_UPPER
    alignment_bull_small = alignment_available and alignment_small_index_156w <= ALIGNMENT_LOWER
    alignment_bear_trend = alignment_available and alignment_trend_index_156w >= ALIGNMENT_UPPER
    alignment_bear_counter = alignment_available and alignment_counter_index_156w <= ALIGNMENT_LOWER
    alignment_bear_small = alignment_available and alignment_small_index_156w >= ALIGNMENT_UPPER
    alignment_bull_count = int(alignment_bull_trend) + int(alignment_bull_counter) + int(alignment_bull_small)
    alignment_bear_count = int(alignment_bear_trend) + int(alignment_bear_counter) + int(alignment_bear_small)
    alignment_bull_3 = alignment_available and alignment_bull_count == 3
    alignment_bear_3 = alignment_available and alignment_bear_count == 3
    alignment_bull_2 = alignment_available and alignment_bull_count == 2
    alignment_bear_2 = alignment_available and alignment_bear_count == 2

    alignment_bull_new_longs = alignment_bull_3 and trend_chg_l > 0 and trend_flow_1w > 0
    alignment_bear_new_shorts = alignment_bear_3 and trend_chg_s > 0 and trend_flow_1w < 0
    alignment_bull_longs_but_net_down = alignment_bull_3 and trend_chg_l > 0 and trend_flow_1w <= 0
    alignment_bear_shorts_but_net_up = alignment_bear_3 and trend_chg_s > 0 and trend_flow_1w >= 0
    alignment_bull_short_covering_only = alignment_bull_3 and trend_chg_s < 0 and trend_flow_1w > 0 and trend_chg_l <= 0
    alignment_bull_long_liquidation_dominant = alignment_bull_3 and trend_chg_l < 0 and trend_flow_1w < 0
    alignment_bear_long_liquidation_only = alignment_bear_3 and trend_chg_l < 0 and trend_flow_1w < 0 and trend_chg_s <= 0
    alignment_bear_short_covering_dominant = alignment_bear_3 and trend_chg_s < 0 and trend_flow_1w > 0

    alignment_bull_macro_ok = macro_long if spec.family == "commodity" else (macro_recovery or macro_long)
    alignment_bear_macro_ok = macro_short if spec.family == "commodity" else (macro_deterioration or macro_short)
    alignment_bull_regime_confirmed = alignment_bull_new_longs and price["long_confirmed"] and alignment_bull_macro_ok
    alignment_bear_regime_confirmed = alignment_bear_new_shorts and price["short_confirmed"] and alignment_bear_macro_ok
    alignment_bull_regime_developing = alignment_bull_new_longs and not alignment_bull_regime_confirmed
    alignment_bear_regime_developing = alignment_bear_new_shorts and not alignment_bear_regime_confirmed

    if not alignment_available:
        future_regime_title = "DATI ALIGNMENT 156W NON DISPONIBILI"
        future_regime_text = future_regime_title
    elif alignment_bull_regime_confirmed:
        future_regime_title = "CAMBIO DI REGIME RIALZISTA CONFERMATO"
        future_regime_text = (
            f"{future_regime_title}\nPerché? Alignment 156W rialzista 3/3 + Long {alignment_trend_label} "
            "in aumento + Net Position in miglioramento + prezzo rialzista."
        )
    elif alignment_bear_regime_confirmed:
        future_regime_title = "CAMBIO DI REGIME RIBASSISTA CONFERMATO"
        future_regime_text = (
            f"{future_regime_title}\nPerché? Alignment 156W ribassista 3/3 + Short {alignment_trend_label} "
            "in aumento + Net Position in peggioramento + prezzo ribassista."
        )
    elif alignment_bull_regime_developing:
        future_regime_title = "LONG CONTRARIAN IN SVILUPPO"
        ending = (
            "il prezzo è già rialzista, manca il completamento della struttura COT."
            if price["long_confirmed"] else "manca ancora la conferma rialzista completa del prezzo."
        )
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Long contrarian?\n"
            f"{alignment_trend_label} vicini al minimo 156W + {alignment_counter_label} vicini al massimo 156W +\n"
            f"Small Traders vicini al minimo 156W.\nPerché è in sviluppo?\n"
            f"I Long aumentano e la Net Position migliora;\n{ending}"
        )
    elif alignment_bear_regime_developing:
        future_regime_title = "SHORT CONTRARIAN IN SVILUPPO"
        ending = (
            "il prezzo è già ribassista, manca il completamento della struttura COT."
            if price["short_confirmed"] else "manca ancora la conferma ribassista completa del prezzo."
        )
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Short contrarian?\n"
            f"{alignment_trend_label} vicini al massimo 156W + {alignment_counter_label} vicini al minimo 156W +\n"
            f"Small Traders vicini al massimo 156W.\nPerché è in sviluppo?\n"
            f"Gli Short aumentano e la Net Position peggiora;\n{ending}"
        )
    elif alignment_bull_longs_but_net_down:
        future_regime_title = "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Long contrarian?\n"
            f"{alignment_trend_label} vicini al minimo 156W + {alignment_counter_label} vicini al massimo 156W +\n"
            "Small Traders vicini al minimo 156W.\nPerché non è ancora un Long in sviluppo?\n"
            "Anche se i Long sono aumentati, la Net Position continua a peggiorare:\ngli Short stanno crescendo più velocemente."
        )
    elif alignment_bear_shorts_but_net_up:
        future_regime_title = "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Short contrarian?\n"
            f"{alignment_trend_label} vicini al massimo 156W + {alignment_counter_label} vicini al minimo 156W +\n"
            "Small Traders vicini al massimo 156W.\nPerché non è ancora uno Short in sviluppo?\n"
            "Anche se gli Short sono aumentati, la Net Position continua a migliorare:\ni Long stanno crescendo più velocemente."
        )
    elif alignment_bull_short_covering_only:
        future_regime_title = "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Long contrarian?\n"
            f"{alignment_trend_label} vicini al minimo 156W + {alignment_counter_label} vicini al massimo 156W +\n"
            "Small Traders vicini al minimo 156W.\nPerché non è ancora un Long in sviluppo?\n"
            "Gli Short stanno diminuendo e la Net Position migliora, ma i Long non aumentano:\n"
            "è short covering, non ancora accumulazione Long."
        )
    elif alignment_bull_long_liquidation_dominant:
        future_regime_title = "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Long contrarian?\n"
            f"{alignment_trend_label} vicini al minimo 156W + {alignment_counter_label} vicini al massimo 156W +\n"
            "Small Traders vicini al minimo 156W.\nPerché non è ancora un Long in sviluppo?\n"
            "Long e Short stanno diminuendo, ma la Net Position peggiora:\ndomina la long liquidation, non lo short covering rialzista."
        )
    elif alignment_bear_long_liquidation_only:
        future_regime_title = "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Short contrarian?\n"
            f"{alignment_trend_label} vicini al massimo 156W + {alignment_counter_label} vicini al minimo 156W +\n"
            "Small Traders vicini al massimo 156W.\nPerché non è ancora uno Short in sviluppo?\n"
            "I Long stanno diminuendo e la Net Position peggiora, ma gli Short non aumentano:\n"
            "è long liquidation, non ancora accumulazione Short."
        )
    elif alignment_bear_short_covering_dominant:
        future_regime_title = "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Short contrarian?\n"
            f"{alignment_trend_label} vicini al massimo 156W + {alignment_counter_label} vicini al minimo 156W +\n"
            "Small Traders vicini al massimo 156W.\nPerché non è ancora uno Short in sviluppo?\n"
            "Long e Short stanno diminuendo, ma la Net Position migliora:\ndomina lo short covering, non una nuova pressione Short."
        )
    elif alignment_bull_3:
        future_regime_title = "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Long contrarian?\n"
            f"{alignment_trend_label} vicini al minimo 156W + {alignment_counter_label} vicini al massimo 156W +\n"
            "Small Traders vicini al minimo 156W.\nPerché non è ancora un Long in sviluppo?\n"
            "Mancano nuovi Long insieme a un miglioramento della Net Position."
        )
    elif alignment_bear_3:
        future_regime_title = "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE"
        future_regime_text = (
            f"{future_regime_title}\nPerché nasce il possibile Short contrarian?\n"
            f"{alignment_trend_label} vicini al massimo 156W + {alignment_counter_label} vicini al minimo 156W +\n"
            "Small Traders vicini al massimo 156W.\nPerché non è ancora uno Short in sviluppo?\n"
            "Mancano nuovi Short insieme a un peggioramento della Net Position."
        )
    elif alignment_bull_2:
        future_regime_title = "SEGNALI LONG CONTRARIAN PARZIALI"
        future_regime_text = f"{future_regime_title}\nAlignment 156W: 2/3"
    elif alignment_bear_2:
        future_regime_title = "SEGNALI SHORT CONTRARIAN PARZIALI"
        future_regime_text = f"{future_regime_title}\nAlignment 156W: 2/3"
    else:
        future_regime_title = "NESSUN CAMBIO DI REGIME CONTRARIAN EVIDENTE"
        future_regime_text = future_regime_title

    future_regime_detail = (
        "POSSIBILE CAMBIO DI REGIME — ALIGNMENT 156W\n"
        f"{alignment_trend_label}: {fmt_decimal(alignment_trend_index_156w, 1)} / 100\n"
        f"{alignment_counter_detail_label}: {fmt_decimal(alignment_counter_index_156w, 1)} / 100\n"
        f"Small Traders / Nonreportable: {fmt_decimal(alignment_small_index_156w, 1)} / 100\n"
        f"Segnali rialzisti: {alignment_bull_count}/3 | Segnali ribassisti: {alignment_bear_count}/3\n"
        f"{future_regime_text}"
        if alignment_available else ""
    )

    # Bias COT + prezzo — frasi allineate a TradingView V1.5.36
    combined_bias = engine_bias
    combined_detail = engine_detail
    if possible_bottom and price["long_confirmed"]:
        combined_bias = "I FONDI TORNANO A COMPRARE / PREZZO IN RIALZO"
    elif possible_top and price["short_confirmed"]:
        combined_bias = "I FONDI INIZIANO A VENDERE / PREZZO IN CALO"
    elif crowded_long and price["short_confirmed"]:
        combined_bias = "NET POSITION VICINA AL MASSIMO STORICO / PREZZO IN CALO"
    elif crowded_short and price["long_confirmed"]:
        combined_bias = "NET POSITION VICINA AL MINIMO STORICO / PREZZO IN RIALZO"
    elif crowded_long:
        combined_bias = "I FONDI COMPRANO, MA LA NET POSITION È VICINA AL MASSIMO STORICO"
    elif crowded_short:
        combined_bias = "I FONDI VENDONO, MA LA NET POSITION È VICINA AL MINIMO STORICO"
    elif full_long:
        combined_bias = "I FONDI COMPRANO E IL PREZZO CONFERMA"
    elif full_short:
        combined_bias = "I FONDI VENDONO E IL PREZZO CONFERMA"
    elif long_price_divergence:
        combined_bias = "I FONDI SONO LONG, MA IL PREZZO SCENDE"
    elif short_price_divergence:
        combined_bias = "I FONDI SONO SHORT, MA IL PREZZO SALE"
    elif confirmed_long:
        combined_bias = "I FONDI SONO LONG, MA IL PREZZO NON CONFERMA"
    elif confirmed_short:
        combined_bias = "I FONDI SONO SHORT, MA IL PREZZO NON CONFERMA"
    elif partial_crowded_long and price["short_confirmed"]:
        combined_bias = "NET POSITION VICINA AL MASSIMO STORICO / PREZZO IN CALO"
    elif partial_crowded_short and price["long_confirmed"]:
        combined_bias = "NET POSITION VICINA AL MINIMO STORICO / PREZZO IN RIALZO"
    elif partial_crowded_long and price["long_confirmed"]:
        combined_bias = "I FONDI COMPRANO, MA LA NET POSITION È VICINA AL MASSIMO STORICO"
    elif partial_crowded_short and price["short_confirmed"]:
        combined_bias = "I FONDI VENDONO, MA LA NET POSITION È VICINA AL MINIMO STORICO"
    elif partial_crowded_long:
        combined_bias = "I FONDI COMPRANO, MA LA NET POSITION È VICINA AL MASSIMO STORICO"
    elif partial_crowded_short:
        combined_bias = "I FONDI VENDONO, MA LA NET POSITION È VICINA AL MINIMO STORICO"
    elif partial_long_with_price:
        combined_bias = "I FONDI TORNANO A COMPRARE E IL PREZZO SALE"
    elif partial_short_with_price:
        combined_bias = "I FONDI TORNANO A VENDERE E IL PREZZO SCENDE"

    # Indicazione operativa, con priorità coerente con il Pine
    action = "ATTENDI: IL QUADRO NON È ANCORA CHIARO"
    reason = "I Fondi e gli altri operatori non mostrano ancora una direzione abbastanza chiara."
    if short_covering:
        action = "NON INSEGUIRE IL RIALZO: ATTENDI NUOVI ACQUISTI REALI"
        reason = "Il miglioramento può dipendere soprattutto dalla chiusura di vecchi Short, non da nuovi acquisti."
    elif long_liquidation:
        action = "EVITA DI COMPRARE SUBITO: ATTENDI UNA STABILIZZAZIONE"
        reason = "Il ribasso può dipendere soprattutto dalla chiusura di vecchi Long, non da nuova pressione Short."
    elif possible_bottom:
        action = "CERCA CONFERME LONG SUI SUPPORTI: EVITA SHORT TARDIVI"
        reason = "Managed Money e Producer migliorano da un estremo Short; un minimo è possibile ma va confermato."
    elif possible_top:
        action = "PROTEGGI GLI EVENTUALI LONG: CERCA CONFERME RIBASSISTE"
        reason = "Managed Money è molto carico e i flussi iniziano a peggiorare."
    elif crowded_long:
        action = "LONG CONFERMATO — NON INSEGUIRE I MASSIMI"
        reason = "Molti Fondi sono già Long; una presa di profitto può rendere la discesa rapida."
    elif crowded_short:
        action = "SHORT CONFERMATO — NON INSEGUIRE I MINIMI"
        reason = "Molti Fondi sono già Short; uno short covering può rendere il rimbalzo rapido."
    elif full_long:
        action = "PRIVILEGIA I PULLBACK LONG: EVITA SHORT CONTROTENDENZA"
        reason = "COT e prezzo Weekly confermano insieme il quadro rialzista."
    elif full_short:
        action = "PRIVILEGIA I RIMBALZI SHORT: EVITA LONG CONTROTENDENZA"
        reason = "COT e prezzo Weekly confermano insieme il quadro ribassista."
    elif long_price_divergence:
        action = "RIDUCI L'AGGRESSIVITÀ LONG: IL PREZZO NON CONFERMA"
        reason = "Il COT è rialzista, ma il prezzo Weekly è ribassista."
    elif short_price_divergence:
        action = "NON INSEGUIRE GLI SHORT: IL PREZZO NON CONFERMA"
        reason = "Il COT è ribassista, ma il prezzo Weekly è rialzista."
    elif confirmed_long:
        action = "ATTENDI CONFERMA DEL PREZZO: NON ANTICIPARE IL LONG"
        reason = "Il quadro istituzionale è rialzista, ma il prezzo Weekly non è ancora completamente concorde."
    elif confirmed_short:
        action = "ATTENDI CONFERMA DEL PREZZO: NON ANTICIPARE LO SHORT"
        reason = "Il quadro istituzionale è ribassista, ma il prezzo Weekly non è ancora completamente concorde."
    elif partial_long_with_price:
        action = "QUADRO IN RECUPERO: CERCA SOLO PULLBACK LONG CONFERMATI"
        reason = "Prezzo e flussi settimanali migliorano, ma la struttura istituzionale non è ancora completa."
    elif partial_short_with_price:
        action = "PRESSIONE IN AUMENTO: CERCA SOLO RIMBALZI SHORT CONFERMATI"
        reason = "Prezzo e flussi settimanali peggiorano, ma la struttura istituzionale non è ancora completa."
    elif macro_recovery or fx_long_alignment or financial_long_alignment or commodity_joint_long:
        action = "ATTENDI CONFERMA DEL PREZZO: NON ANTICIPARE IL LONG"
        reason = "I Fondi migliorano, ma le ultime settimane non confermano ancora un quadro rialzista completo."
    elif macro_deterioration or fx_short_alignment or financial_short_alignment or commodity_joint_short:
        action = "RIDUCI L'AGGRESSIVITÀ LONG: ATTENDI CONFERMA RIBASSISTA"
        reason = "I Fondi peggiorano, ma manca ancora una conferma Short completa."
    elif trend_flow_1w > 0:
        action = "QUADRO IN MIGLIORAMENTO: ATTENDI UN PULLBACK CONFERMATO"
        reason = "Il flusso principale migliora, ma gli altri dati non sono ancora tutti allineati."
    elif trend_flow_1w < 0:
        action = "QUADRO IN PEGGIORAMENTO: MANTIENI PRUDENZA SUI LONG"
        reason = "Il flusso principale peggiora, ma manca ancora una conferma Short completa."

    report_date = pd.Timestamp(cur["date"]).date()
    age_days = max((date.today() - report_date).days, 0)
    freshness = "FRESCO" if age_days <= 10 else "POSSIBILE RITARDO" if age_days <= 17 else "DATI DATATI"
    final_bias = combined_bias
    final_detail = combined_detail
    if age_days > 17:
        final_bias = f"{combined_bias} — DATI COT DATATI"
        action = "NESSUNA INDICAZIONE ATTUALE: ATTENDI UN NUOVO REPORT"
        reason = f"Le posizioni disponibili risalgono a {age_days} giorni fa."
    elif age_days > 10:
        final_bias = f"{combined_bias} — POSSIBILE RITARDO"
        action = f"{action} — VERIFICA L'ARRIVO DEL NUOVO REPORT"
        reason += f" Le posizioni disponibili risalgono a {age_days} giorni fa."

    conc_long_rank = historical_percentile(df["conc_long"], 156)
    conc_short_rank = historical_percentile(df["conc_short"], 156)
    concentration_available = not pd.isna(conc_long_rank) and not pd.isna(conc_short_rank)
    conc_long_high = concentration_available and conc_long_rank >= 80
    conc_short_high = concentration_available and conc_short_rank >= 80
    concentration_moderate = (
        concentration_available
        and not conc_long_high
        and not conc_short_high
        and max(conc_long_rank, conc_short_rank) >= 60
    )

    if not concentration_available:
        concentration_state = "DATI NON DISPONIBILI"
        concentration_detail = "I dati dei Top 8 trader non sono disponibili."
    elif conc_long_high and conc_short_high:
        concentration_state = "TOP 8 TRADER FORTEMENTE ESPOSTI LONG E SHORT"
        concentration_detail = (
            "I Top 8 Trader sono fortemente esposti sia Long sia Short. Se uno dei due lati riduce rapidamente "
            "le posizioni, il prezzo può muoversi con maggiore violenza."
        )
    elif conc_long_high:
        concentration_state = "TOP 8 TRADER FORTEMENTE ESPOSTI LONG"
        concentration_detail = (
            "I Top 8 Trader sono fortemente esposti Long. Se prezzo e flussi peggiorano, la chiusura di queste "
            "posizioni può accelerare la discesa."
        )
    elif conc_short_high:
        concentration_state = "TOP 8 TRADER FORTEMENTE ESPOSTI SHORT"
        concentration_detail = (
            "I Top 8 Trader sono fortemente esposti Short. Se prezzo e flussi migliorano, la chiusura di queste "
            "posizioni può rendere il rimbalzo più rapido."
        )
    elif concentration_moderate:
        concentration_state = "TOP 8 TRADER PIÙ ESPOSTI DEL NORMALE"
        concentration_detail = "L'esposizione dei Top 8 Trader è superiore alla norma, ma non è ancora estrema."
    else:
        concentration_state = "ESPOSIZIONE DEI TOP 8 TRADER NELLA NORMA"
        concentration_detail = "L'esposizione dei Top 8 Trader è nella norma."

    # =========================================================================
    # LETTURA SEMPLICE — FRASI IDENTICHE A TRADINGVIEW V1.5.36
    # =========================================================================
    short_extreme_recovery = extreme_short and trend_flow_1w > 0
    long_extreme_deterioration = extreme_long and trend_flow_1w < 0
    short_extreme_pressure = extreme_short and trend_flow_1w <= 0
    long_extreme_pressure = extreme_long and trend_flow_1w >= 0

    simple_title = "NEUTRALE / POCO CHIARO"
    simple_detail = (
        "Il quadro non è ancora abbastanza chiaro. I Fondi e gli altri operatori non mostrano tutti la stessa "
        "direzione. In questa situazione è meglio attendere una nuova conferma."
    )
    plain_action = "ATTENDI. NON CI SONO ANCORA CONDIZIONI ABBASTANZA CHIARE."
    explanation = "I dati disponibili non mostrano ancora una situazione abbastanza chiara per spiegare una direzione precisa."

    if age_days > 17:
        simple_title = "DATI COT TROPPO VECCHI"
        simple_detail = (
            "I dati COT sono troppo vecchi. Nel frattempo i Fondi potrebbero aver cambiato posizione e la situazione "
            "mostrata potrebbe non essere più valida."
        )
        plain_action = "NON USARE IL COT PER NUOVI INGRESSI. ATTENDI UN REPORT AGGIORNATO."
        explanation = (
            f"Le posizioni mostrate risalgono a {age_days} giorni fa. Un dato così vecchio non deve essere trattato "
            "come una fotografia attuale del mercato."
        )
    elif short_extreme_recovery:
        simple_title = "RECUPERO DA NET POSITION VICINA AL MINIMO STORICO"
        simple_detail = (
            "La Net Position si trova vicino al minimo del proprio range storico e nell'ultimo report ha iniziato a migliorare. "
            "Il miglioramento può derivare da nuovi Long, dalla riduzione degli Short o da entrambe le componenti. Questo può "
            "favorire un rimbalzo, ma non rappresenta ancora un segnale Long completo finché il prezzo non lo conferma."
        )
        plain_action = (
            "NON INSEGUIRE I MINIMI. CERCA EVENTUALI LONG SOLO DOPO UN PULLBACK E UNA NUOVA CONFERMA RIALZISTA."
            if price["long_confirmed"]
            else "NON APRIRE NUOVI SHORT SUI MINIMI. ATTENDI CHE IL PREZZO CONFERMI IL RECUPERO PRIMA DI VALUTARE UN LONG."
        )
        explanation = (
            f"Il COT Index di {spec.trend_label} è {fmt_decimal(cot_index, 1)} su 100. Un valore sotto 20 significa che la "
            f"Net Position si trova vicino al minimo degli ultimi {cot_lookback} report; non significa necessariamente che i "
            f"Fondi siano Net Short. Nell'ultimo report la Net Position è migliorata di {fmt_number(trend_flow_1w, signed=True)}. "
            "Sulle ultime 3 e 6 settimane il recupero non è ancora completo e il prezzo deve ancora confermare una fase "
            "rialzista più stabile."
        )
    elif long_extreme_deterioration:
        simple_title = "PEGGIORAMENTO DA NET POSITION VICINA AL MASSIMO STORICO"
        simple_detail = (
            "La Net Position si trova vicino al massimo del proprio range storico e nell'ultimo report ha iniziato a peggiorare. "
            "Il peggioramento può derivare dalla riduzione dei Long, dall'aumento degli Short o da entrambe le componenti. Questo "
            "può aumentare la pressione ribassista, ma non rappresenta ancora un segnale Short completo finché il prezzo non lo conferma."
        )
        plain_action = (
            "NON INSEGUIRE I MASSIMI. VALUTA EVENTUALI SHORT SOLO DOPO UN RIMBALZO E UNA NUOVA CONFERMA RIBASSISTA."
            if price["short_confirmed"]
            else "NON APRIRE NUOVI LONG SUI MASSIMI. ATTENDI CHE IL PREZZO CONFERMI IL PEGGIORAMENTO PRIMA DI VALUTARE UNO SHORT."
        )
        explanation = (
            f"Il COT Index di {spec.trend_label} è {fmt_decimal(cot_index, 1)} su 100. Un valore sopra 80 significa che la "
            f"Net Position si trova vicino al massimo degli ultimi {cot_lookback} report; non significa necessariamente che i "
            f"Fondi siano Net Long. Nell'ultimo report la Net Position è peggiorata di {fmt_number(trend_flow_1w, signed=True)}. "
            "Sulle ultime 3 e 6 settimane il cambiamento non è ancora completo e il prezzo deve ancora confermare una fase "
            "ribassista più stabile."
        )
    elif short_extreme_pressure or crowded_short or partial_crowded_short:
        simple_title = "SHORT CONFERMATO — NON INSEGUIRE"
        simple_detail = (
            "Il quadro recente rimane ribassista e la Net Position è già vicina al minimo del proprio range storico. La discesa "
            "può proseguire, ma l'estensione aumenta il rischio di un rimbalzo rapido. Entrare Short adesso è più rischioso, "
            "perché si potrebbe vendere vicino ai minimi del movimento."
        )
        plain_action = (
            "NON INSEGUIRE I MINIMI. ATTENDI UN RIMBALZO VERSO UNA RESISTENZA, IL POC O UN ALTRO LIVELLO IMPORTANTE. "
            "VALUTA UNO SHORT SOLO DOPO UNA NUOVA CONFERMA RIBASSISTA."
        )
        explanation = (
            f"Il COT Index di {spec.trend_label} è {fmt_decimal(cot_index, 1)} su 100. Un valore sotto 20 indica che la Net "
            f"Position è vicina al minimo degli ultimi {cot_lookback} report. Questo segnala estensione storica, non necessariamente "
            "una posizione Net Short. La direzione può restare ribassista, ma aumenta il rischio di un rimbalzo veloce."
        )
    elif long_extreme_pressure or crowded_long or partial_crowded_long:
        simple_title = "LONG CONFERMATO — NON INSEGUIRE"
        simple_detail = (
            "Il quadro recente rimane rialzista e la Net Position è già vicina al massimo del proprio range storico. Il rialzo "
            "può proseguire, ma l'estensione aumenta il rischio di una correzione rapida. Entrare Long adesso è più rischioso, "
            "perché si potrebbe comprare vicino ai massimi del movimento."
        )
        plain_action = (
            "NON INSEGUIRE I MASSIMI. ATTENDI UN PULLBACK VERSO UN SUPPORTO, IL POC O UN ALTRO LIVELLO IMPORTANTE. "
            "VALUTA UN LONG SOLO DOPO UNA NUOVA CONFERMA RIALZISTA."
        )
        explanation = (
            f"Il COT Index di {spec.trend_label} è {fmt_decimal(cot_index, 1)} su 100. Un valore sopra 80 indica che la Net "
            f"Position è vicina al massimo degli ultimi {cot_lookback} report. Questo segnala estensione storica, non necessariamente "
            "una posizione Net Long. La direzione può restare rialzista, ma aumenta il rischio di una correzione veloce."
        )
    elif short_covering:
        simple_title = "SHORT COVERING"
        simple_detail = (
            "Molti operatori stanno chiudendo vecchie posizioni Short. Per uscire devono ricomprare i contratti e questi "
            "acquisti possono far salire rapidamente il prezzo. Questo rimbalzo non significa ancora che sia iniziato un vero "
            "trend rialzista: potrebbe essere soltanto la chiusura delle vendite precedenti."
        )
        plain_action = "NON INSEGUIRE IL RIALZO. ATTENDI CHE IL PREZZO COSTRUISCA UNA VERA CONFERMA LONG."
        explanation = (
            "La Net Position dei Fondi migliora mentre l'Open Interest diminuisce. Questa combinazione suggerisce che una "
            "parte del rialzo può dipendere dalla chiusura degli Short, non necessariamente dall'ingresso di nuovi compratori."
        )
    elif long_liquidation:
        simple_title = "LIQUIDAZIONE LONG"
        simple_detail = (
            "Molti operatori stanno chiudendo vecchie posizioni Long. Per uscire devono vendere i contratti e queste vendite "
            "possono far scendere rapidamente il prezzo. Questa discesa non significa ancora che sia iniziato un vero trend "
            "ribassista: potrebbe essere soltanto la chiusura degli acquisti precedenti."
        )
        plain_action = "NON INSEGUIRE LA DISCESA. ATTENDI CHE IL PREZZO COSTRUISCA UNA VERA CONFERMA SHORT."
        explanation = (
            "La Net Position dei Fondi peggiora mentre l'Open Interest diminuisce. Questa combinazione suggerisce che una "
            "parte della discesa può dipendere dalla chiusura dei Long, non necessariamente dall'ingresso di nuovi venditori."
        )
    elif full_long:
        simple_title = "LONG CONFERMATO"
        simple_detail = (
            "Il quadro dei flussi è rialzista e il prezzo settimanale conferma la stessa direzione. Questo significa che "
            "posizionamento recente e prezzo stanno lavorando insieme. Il rischio principale è entrare dopo una salita già troppo estesa."
        )
        plain_action = "CERCA EVENTUALI LONG SUI PULLBACK. NON COMPRARE DOPO UNA SALITA GIÀ MOLTO ESTESA."
        explanation = (
            "I flussi dei Fondi sono positivi sia nell'ultimo report sia nella struttura delle ultime 3 e 6 settimane. Anche "
            "il prezzo settimanale è sopra la propria EMA e sta salendo. La direzione dei Fondi e quella del prezzo sono quindi concordi."
        )
    elif full_short:
        simple_title = "SHORT CONFERMATO"
        simple_detail = (
            "Il quadro dei flussi è ribassista e il prezzo settimanale conferma la stessa direzione. Questo significa che "
            "posizionamento recente e prezzo stanno lavorando insieme. Il rischio principale è entrare dopo una discesa già troppo estesa."
        )
        plain_action = "CERCA EVENTUALI SHORT DOPO UN RIMBALZO. NON VENDERE DOPO UNA DISCESA GIÀ MOLTO ESTESA."
        explanation = (
            "I flussi dei Fondi sono negativi sia nell'ultimo report sia nella struttura delle ultime 3 e 6 settimane. Anche "
            "il prezzo settimanale è sotto la propria EMA e sta scendendo. La direzione dei Fondi e quella del prezzo sono quindi concordi."
        )
    elif long_price_divergence:
        simple_title = "COT LONG, MA IL PREZZO NON CONFERMA"
        simple_detail = (
            "I flussi COT restano orientati al rialzo, ma il prezzo settimanale sta scendendo. Le due letture non sono d'accordo. "
            "Il posizionamento recente può anticipare un recupero, ma il prezzo potrebbe continuare a scendere prima di confermarlo."
        )
        plain_action = "NON ANTICIPARE IL LONG. ATTENDI CHE IL PREZZO TORNI A SALIRE E CONFERMI IL POSIZIONAMENTO DEI FONDI."
        explanation = (
            "I dati COT mostrano un orientamento rialzista, mentre l'ultima settimana chiusa non conferma questa direzione. "
            "Quando prezzo e Fondi divergono, il timing resta incerto."
        )
    elif short_price_divergence:
        simple_title = "COT SHORT, MA IL PREZZO NON CONFERMA"
        simple_detail = (
            "I flussi COT restano orientati al ribasso, ma il prezzo settimanale sta salendo. Le due letture non sono d'accordo. "
            "Il posizionamento recente può anticipare una nuova discesa, ma il prezzo potrebbe continuare a salire prima di confermarla."
        )
        plain_action = "NON ANTICIPARE LO SHORT. ATTENDI CHE IL PREZZO TORNI A SCENDERE E CONFERMI IL POSIZIONAMENTO DEI FONDI."
        explanation = (
            "I dati COT mostrano un orientamento ribassista, mentre l'ultima settimana chiusa non conferma questa direzione. "
            "Quando prezzo e Fondi divergono, il timing resta incerto."
        )
    elif confirmed_long:
        simple_title = "LONG IN COSTRUZIONE"
        simple_detail = (
            "I Fondi sono orientati Long e la struttura COT è già coerente, ma il prezzo settimanale non ha ancora "
            "completato la conferma rialzista. Il quadro resta favorevole ai Long, ma il timing non è ancora completo."
        )
        plain_action = "I FONDI SONO ORIENTATI LONG. ATTENDI LA CONFERMA DEL PREZZO."
        explanation = (
            "I Fondi stanno mantenendo una direzione rialzista sia nell'ultimo report sia nelle ultime settimane, "
            "ma il prezzo settimanale non ha ancora confermato pienamente il rialzo."
        )
    elif confirmed_short:
        simple_title = "SHORT IN COSTRUZIONE"
        simple_detail = (
            "I Fondi sono orientati Short e la struttura COT è già coerente, ma il prezzo settimanale non ha ancora "
            "completato la conferma ribassista. Il quadro resta favorevole agli Short, ma il timing non è ancora completo."
        )
        plain_action = "I FONDI SONO ORIENTATI SHORT. ATTENDI LA CONFERMA DEL PREZZO."
        explanation = (
            "I Fondi stanno mantenendo una direzione ribassista sia nell'ultimo report sia nelle ultime settimane, "
            "ma il prezzo settimanale non ha ancora confermato pienamente la discesa."
        )
    elif partial_long:
        simple_title = "LONG IN COSTRUZIONE"
        simple_detail = (
            "I Fondi stanno migliorando, ma il recupero non è ancora completo. Una parte dei dati è diventata positiva, mentre "
            "la struttura delle settimane precedenti o il prezzo non confermano ancora pienamente. Potrebbe essere l'inizio di "
            "una fase rialzista, ma è ancora presto per considerarla confermata."
        )
        if price["long_confirmed"]:
            if spec.is_fx:
                plain_action = (
                    "IL PREZZO È GIÀ RIALZISTA. ATTENDI CHE I LEVERAGED FUNDS CONFERMINO IL RIALZO ANCHE NELLA STRUTTURA 3-6W. POI VALUTA EVENTUALI LONG SUI PULLBACK." if not macro_long
                    else "IL PREZZO È GIÀ RIALZISTA E I LEVERAGED FUNDS SONO RIALZISTI A 3-6W. ATTENDI CHE I DEALER/INTERMEDIARY CONFERMINO IL QUADRO NELL'ULTIMO REPORT. POI VALUTA EVENTUALI LONG SUI PULLBACK." if not fx_long_alignment
                    else "IL PREZZO È GIÀ RIALZISTA E I LEVERAGED FUNDS SONO RIALZISTI A 3-6W. ATTENDI CHE I DEALER/INTERMEDIARY CONFERMINO IL QUADRO 3-6W MUOVENDOSI VERSO LO SHORT. POI VALUTA EVENTUALI LONG SUI PULLBACK." if not counter_macro_short
                    else "IL PREZZO È GIÀ RIALZISTA. ATTENDI IL COMPLETAMENTO DELLA CONFERMA COT PRIMA DI VALUTARE EVENTUALI LONG SUI PULLBACK."
                )
            elif spec.family == "financial":
                plain_action = (
                    "IL PREZZO È GIÀ RIALZISTA. ATTENDI CHE I LEVERAGED FUNDS CONFERMINO IL RIALZO ANCHE NELLA STRUTTURA 3-6W. POI VALUTA EVENTUALI LONG SUI PULLBACK." if not macro_long
                    else "IL PREZZO È GIÀ RIALZISTA E I LEVERAGED FUNDS SONO RIALZISTI A 3-6W. ATTENDI CHE GLI ASSET MANAGER CONFERMINO IL QUADRO NELL'ULTIMO REPORT. POI VALUTA EVENTUALI LONG SUI PULLBACK." if not financial_long_alignment
                    else "IL PREZZO È GIÀ RIALZISTA E I LEVERAGED FUNDS SONO RIALZISTI A 3-6W. ATTENDI CHE ANCHE GLI ASSET MANAGER CONFERMINO IL QUADRO 3-6W. POI VALUTA EVENTUALI LONG SUI PULLBACK." if not counter_macro_long
                    else "IL PREZZO È GIÀ RIALZISTA. ATTENDI IL COMPLETAMENTO DELLA CONFERMA COT PRIMA DI VALUTARE EVENTUALI LONG SUI PULLBACK."
                )
            else:
                plain_action = "IL PREZZO È GIÀ RIALZISTA. ATTENDI CHE IL QUADRO COT SI COMPLETI PRIMA DI VALUTARE EVENTUALI LONG SUI PULLBACK."
        else:
            plain_action = "ATTENDI LA CONFERMA DEL PREZZO. VALUTA UN LONG SOLO DOPO UN PULLBACK E UN NUOVO SEGNALE RIALZISTA."
        explanation = (
            f"La Net Position dei Fondi è migliorata di {fmt_number(trend_flow_1w, signed=True)} nell'ultimo report. Le variazioni "
            f"a 3 e 6 settimane sono rispettivamente {fmt_number(trend_flow_3w, signed=True)} e "
            f"{fmt_number(trend_flow_6w, signed=True)}. Il miglioramento recente non ha quindi ancora trasformato completamente "
            "la struttura precedente."
        )
    elif partial_short:
        simple_title = "SHORT IN COSTRUZIONE"
        simple_detail = (
            "I Fondi stanno peggiorando, ma la fase ribassista non è ancora completa. Una parte dei dati è diventata negativa, "
            "mentre la struttura delle settimane precedenti o il prezzo non confermano ancora pienamente. Potrebbe essere "
            "l'inizio di una fase debole, ma è ancora presto per considerarla confermata."
        )
        if price["short_confirmed"]:
            if spec.is_fx:
                plain_action = (
                    "IL PREZZO È GIÀ RIBASSISTA. ATTENDI CHE I LEVERAGED FUNDS CONFERMINO IL RIBASSO ANCHE NELLA STRUTTURA 3-6W. POI VALUTA EVENTUALI SHORT SUI RIMBALZI." if not macro_short
                    else "IL PREZZO È GIÀ RIBASSISTA E I LEVERAGED FUNDS SONO RIBASSISTI A 3-6W. ATTENDI CHE I DEALER/INTERMEDIARY CONFERMINO IL QUADRO NELL'ULTIMO REPORT. POI VALUTA EVENTUALI SHORT SUI RIMBALZI." if not fx_short_alignment
                    else "IL PREZZO È GIÀ RIBASSISTA E I LEVERAGED FUNDS SONO RIBASSISTI A 3-6W. ATTENDI CHE I DEALER/INTERMEDIARY CONFERMINO IL QUADRO 3-6W MUOVENDOSI VERSO IL LONG. POI VALUTA EVENTUALI SHORT SUI RIMBALZI." if not counter_macro_long
                    else "IL PREZZO È GIÀ RIBASSISTA. ATTENDI IL COMPLETAMENTO DELLA CONFERMA COT PRIMA DI VALUTARE EVENTUALI SHORT SUI RIMBALZI."
                )
            elif spec.family == "financial":
                plain_action = (
                    "IL PREZZO È GIÀ RIBASSISTA. ATTENDI CHE I LEVERAGED FUNDS CONFERMINO IL RIBASSO ANCHE NELLA STRUTTURA 3-6W. POI VALUTA EVENTUALI SHORT SUI RIMBALZI." if not macro_short
                    else "IL PREZZO È GIÀ RIBASSISTA E I LEVERAGED FUNDS SONO RIBASSISTI A 3-6W. ATTENDI CHE GLI ASSET MANAGER CONFERMINO IL QUADRO NELL'ULTIMO REPORT. POI VALUTA EVENTUALI SHORT SUI RIMBALZI." if not financial_short_alignment
                    else "IL PREZZO È GIÀ RIBASSISTA E I LEVERAGED FUNDS SONO RIBASSISTI A 3-6W. ATTENDI CHE ANCHE GLI ASSET MANAGER CONFERMINO IL QUADRO 3-6W. POI VALUTA EVENTUALI SHORT SUI RIMBALZI." if not counter_macro_short
                    else "IL PREZZO È GIÀ RIBASSISTA. ATTENDI IL COMPLETAMENTO DELLA CONFERMA COT PRIMA DI VALUTARE EVENTUALI SHORT SUI RIMBALZI."
                )
            else:
                plain_action = "IL PREZZO È GIÀ RIBASSISTA. ATTENDI CHE IL QUADRO COT SI COMPLETI PRIMA DI VALUTARE EVENTUALI SHORT SUI RIMBALZI."
        else:
            plain_action = "ATTENDI LA CONFERMA DEL PREZZO. VALUTA UNO SHORT SOLO DOPO UN RIMBALZO E UN NUOVO SEGNALE RIBASSISTA."
        explanation = (
            f"La Net Position dei Fondi è cambiata di {fmt_number(trend_flow_1w, signed=True)} nell'ultimo report. Le variazioni "
            f"a 3 e 6 settimane sono rispettivamente {fmt_number(trend_flow_3w, signed=True)} e "
            f"{fmt_number(trend_flow_6w, signed=True)}. Il peggioramento recente non ha quindi ancora trasformato completamente "
            "la struttura precedente."
        )
    elif price["long_confirmed"]:
        simple_title = "IL PREZZO SALE, MA IL COT NON CONFERMA ANCORA"
        simple_detail = (
            "Il prezzo settimanale sta salendo, ma i Fondi non mostrano ancora una conferma altrettanto chiara. Il rialzo può "
            "continuare, ma per ora manca una piena conferma dei Fondi."
        )
        plain_action = "NON INSEGUIRE IL RIALZO. ATTENDI CHE ANCHE I DATI COT MIGLIORINO."
        explanation = (
            "Il prezzo è sopra la EMA settimanale e sta salendo, mentre i flussi dei Fondi restano misti. Il movimento è "
            "sostenuto dal prezzo, ma non ancora da una lettura COT completa."
        )
    elif price["short_confirmed"]:
        simple_title = "IL PREZZO SCENDE, MA IL COT NON CONFERMA ANCORA"
        simple_detail = (
            "Il prezzo settimanale sta scendendo, ma i Fondi non mostrano ancora una conferma altrettanto chiara. La discesa "
            "può continuare, ma per ora manca una piena conferma dei Fondi."
        )
        plain_action = "NON INSEGUIRE LA DISCESA. ATTENDI CHE ANCHE I DATI COT PEGGIORINO."
        explanation = (
            "Il prezzo è sotto la EMA settimanale e sta scendendo, mentre i flussi dei Fondi restano misti. Il movimento è "
            "sostenuto dal prezzo, ma non ancora da una lettura COT completa."
        )

    # =========================================================================
    # SINTESI CHIARA DEL QUADRO — allineata alla V1.5.36
    # L'esposizione attuale deriva dalle percentuali Long/Short direzionali;
    # il COT Index descrive separatamente la collocazione nel range storico.
    # =========================================================================
    exposure_strong_long = not pd.isna(directional_long_pct) and directional_long_pct >= 70.0
    exposure_strong_short = not pd.isna(directional_short_pct) and directional_short_pct >= 70.0
    exposure_long_side = not pd.isna(directional_long_pct) and directional_long_pct >= 55.0
    exposure_short_side = not pd.isna(directional_short_pct) and directional_short_pct >= 55.0
    exposure_balanced = (
        not pd.isna(directional_long_pct)
        and not pd.isna(directional_short_pct)
        and not exposure_long_side
        and not exposure_short_side
    )

    if exposure_strong_long:
        exposure_summary = f"I {spec.trend_label} rimangono fortemente esposti Long"
    elif exposure_long_side:
        exposure_summary = f"I {spec.trend_label} rimangono prevalentemente esposti Long"
    elif exposure_strong_short:
        exposure_summary = f"I {spec.trend_label} rimangono fortemente esposti Short"
    elif exposure_short_side:
        exposure_summary = f"I {spec.trend_label} rimangono prevalentemente esposti Short"
    else:
        exposure_summary = f"I {spec.trend_label} mantengono una posizione bilanciata"

    if trend_chg_l < 0 and trend_chg_s > 0:
        last_change_summary = "hanno ridotto l'esposizione rialzista e aperto nuovi Short."
    elif trend_chg_l > 0 and trend_chg_s < 0:
        last_change_summary = "hanno aumentato l'esposizione rialzista e ridotto gli Short."
    elif trend_chg_l < 0 and trend_chg_s < 0:
        last_change_summary = "hanno ridotto sia i Long sia gli Short."
    elif trend_chg_l > 0 and trend_chg_s > 0:
        last_change_summary = "hanno aumentato sia i Long sia gli Short."
    elif trend_chg_l < 0:
        last_change_summary = "hanno ridotto i Long."
    elif trend_chg_l > 0:
        last_change_summary = "hanno aumentato i Long."
    elif trend_chg_s > 0:
        last_change_summary = "hanno aperto nuovi Short."
    elif trend_chg_s < 0:
        last_change_summary = "hanno ridotto gli Short."
    else:
        last_change_summary = "non hanno modificato in modo significativo Long e Short."

    last_change_opposes_exposure = (
        (exposure_long_side and trend_flow_1w < 0)
        or (exposure_short_side and trend_flow_1w > 0)
    )
    conjunction = "; nell'ultimo report " if exposure_balanced else ", ma nell'ultimo report " if last_change_opposes_exposure else " e nell'ultimo report "
    funds_summary = exposure_summary + conjunction + last_change_summary

    counter_subject = (
        "I Producer/Merchant" if spec.family == "commodity"
        else "I Dealer/Intermediary" if spec.is_fx
        else "Gli Asset Manager"
    )
    counter_opposite = (
        (trend_flow_1w > 0 and counter_flow_1w < 0)
        or (trend_flow_1w < 0 and counter_flow_1w > 0)
    )
    if pd.isna(counter_flow_1w):
        counter_summary = f"{counter_subject} non dispongono di dati sufficienti per il confronto."
    elif counter_flow_1w > 0:
        counter_summary = f"{counter_subject} si sono {'invece ' if counter_opposite else ''}mossi verso il Long."
    elif counter_flow_1w < 0:
        counter_summary = f"{counter_subject} si sono {'invece ' if counter_opposite else ''}mossi verso lo Short."
    else:
        counter_summary = f"{counter_subject} non hanno mostrato un cambiamento netto significativo."

    if not price.get("available"):
        price_summary = "La conferma del prezzo settimanale non è disponibile."
    elif price.get("long_confirmed") and trend_flow_1w < 0:
        price_summary = "Il prezzo settimanale rimane rialzista, quindi il possibile indebolimento COT non è ancora confermato dal prezzo."
    elif price.get("long_confirmed") and trend_flow_1w > 0:
        price_summary = "Il prezzo settimanale rimane rialzista e conferma il miglioramento del posizionamento COT."
    elif price.get("short_confirmed") and trend_flow_1w > 0:
        price_summary = "Il prezzo settimanale rimane ribassista, quindi il miglioramento COT non è ancora confermato dal prezzo."
    elif price.get("short_confirmed") and trend_flow_1w < 0:
        price_summary = "Il prezzo settimanale rimane ribassista e conferma l'indebolimento del posizionamento COT."
    else:
        price_summary = "Il prezzo settimanale non fornisce ancora una conferma chiara del quadro COT."

    if pd.isna(cot_index_156w):
        historical_range_summary = "La collocazione storica della Net Position non è disponibile."
    elif cot_index_156w >= 80:
        historical_range_summary = "Rispetto alle ultime 156W, la loro Net Position si trova vicino al massimo del range storico."
    elif cot_index_156w >= 60:
        historical_range_summary = "Rispetto alle ultime 156W, la loro Net Position si trova nella fascia alta del range storico."
    elif cot_index_156w > 40:
        historical_range_summary = "Rispetto alle ultime 156W, la loro Net Position si trova nella fascia centrale del range storico."
    elif cot_index_156w > 20:
        historical_range_summary = "Rispetto alle ultime 156W, la loro Net Position si trova nella fascia bassa del range storico."
    else:
        historical_range_summary = "Rispetto alle ultime 156W, la loro Net Position si trova vicino al minimo del range storico."

    if alignment_bull_regime_confirmed:
        future_regime_summary = (
            f"Il cambio di regime rialzista ha ora una conferma iniziale: l'Alignment è 3/3, i {alignment_trend_label} "
            "stanno aumentando i Long, la Net Position migliora e il prezzo conferma il rialzo."
        )
    elif alignment_bear_regime_confirmed:
        future_regime_summary = (
            f"Il cambio di regime ribassista ha ora una conferma iniziale: l'Alignment è 3/3, i {alignment_trend_label} "
            "stanno aumentando gli Short, la Net Position peggiora e il prezzo conferma il ribasso."
        )
    elif alignment_bull_regime_developing:
        future_regime_summary = (
            f"Il COT mostra un possibile Long contrarian 3/3: i {alignment_trend_label} stanno aumentando i Long e la Net Position sta migliorando. "
            + ("Il prezzo è già rialzista; manca ancora il completamento della struttura COT 3-6W." if price["long_confirmed"] else "Il cambio di regime non è ancora confermato finché il prezzo non completa la conferma rialzista.")
        )
    elif alignment_bear_regime_developing:
        future_regime_summary = (
            f"Il COT mostra un possibile Short contrarian 3/3: i {alignment_trend_label} stanno aumentando gli Short e la Net Position sta peggiorando. "
            + ("Il prezzo è già ribassista; manca ancora il completamento della struttura COT 3-6W." if price["short_confirmed"] else "Il cambio di regime non è ancora confermato finché il prezzo non completa la conferma ribassista.")
        )
    elif alignment_bull_longs_but_net_down:
        future_regime_summary = f"Il COT mostra un Alignment rialzista 3/3 e i {alignment_trend_label} hanno aumentato i Long, ma gli Short sono cresciuti di più e la Net Position continua a peggiorare. Il possibile Long contrarian resta quindi in costruzione, non ancora in sviluppo."
    elif alignment_bear_shorts_but_net_up:
        future_regime_summary = f"Il COT mostra un Alignment ribassista 3/3 e i {alignment_trend_label} hanno aumentato gli Short, ma i Long sono cresciuti di più e la Net Position continua a migliorare. Il possibile Short contrarian resta quindi in costruzione, non ancora in sviluppo."
    elif alignment_bull_short_covering_only:
        future_regime_summary = "Il COT mostra un Alignment rialzista 3/3. Gli Short diminuiscono e la Net Position migliora, ma i Long non aumentano: è short covering, quindi cresce il rischio di rimbalzo ma non è ancora un cambio di regime Long."
    elif alignment_bull_long_liquidation_dominant:
        future_regime_summary = "Il COT mostra un Alignment rialzista 3/3, ma Long e Short stanno diminuendo e la Net Position peggiora. Domina la long liquidation: il possibile Long contrarian resta in costruzione."
    elif alignment_bear_long_liquidation_only:
        future_regime_summary = "Il COT mostra un Alignment ribassista 3/3. I Long diminuiscono e la Net Position peggiora, ma gli Short non aumentano: è long liquidation, quindi cresce il rischio di discesa ma non è ancora un cambio di regime Short."
    elif alignment_bear_short_covering_dominant:
        future_regime_summary = "Il COT mostra un Alignment ribassista 3/3, ma Long e Short stanno diminuendo e la Net Position migliora. Domina lo short covering: il possibile Short contrarian resta in costruzione."
    elif alignment_bull_3:
        future_regime_summary = f"Il COT sta preparando un possibile Long contrarian: {alignment_trend_label}, {alignment_counter_label} e Small Traders sono allineati 3/3 sugli estremi 156W. Per passare a Long contrarian in sviluppo servono nuovi Long della categoria seguita insieme a un miglioramento della Net Position, poi la conferma rialzista del prezzo."
    elif alignment_bear_3:
        future_regime_summary = f"Il COT sta preparando un possibile Short contrarian: {alignment_trend_label}, {alignment_counter_label} e Small Traders sono allineati 3/3 sugli estremi 156W. Per passare a Short contrarian in sviluppo servono nuovi Short della categoria seguita insieme a un peggioramento della Net Position, poi la conferma ribassista del prezzo."
    elif alignment_bull_2:
        future_regime_summary = "Sono presenti segnali Long contrarian parziali 2/3, ma non sono ancora sufficienti per parlare di cambio di regime."
    elif alignment_bear_2:
        future_regime_summary = "Sono presenti segnali Short contrarian parziali 2/3, ma non sono ancora sufficienti per parlare di cambio di regime."
    else:
        future_regime_summary = ""

    clear_summary = f"{funds_summary} {counter_summary} {historical_range_summary}{(' ' + future_regime_summary) if future_regime_summary else ''} {price_summary}"

    action_confirmed_long_view = full_long or long_extreme_pressure or crowded_long or partial_crowded_long
    action_confirmed_short_view = full_short or short_extreme_pressure or crowded_short or partial_crowded_short
    action_developing_long_view = short_extreme_recovery or short_covering or long_price_divergence or partial_long
    action_developing_short_view = long_extreme_deterioration or long_liquidation or short_price_divergence or partial_short
    if action_confirmed_long_view:
        action_view_intro = "IN FUNZIONE DELLA VIEW RIALZISTA,"
    elif action_confirmed_short_view:
        action_view_intro = "IN FUNZIONE DELLA VIEW RIBASSISTA,"
    elif action_developing_long_view:
        action_view_intro = "IN FUNZIONE DI UNA POSSIBILE VIEW RIALZISTA,"
    elif action_developing_short_view:
        action_view_intro = "IN FUNZIONE DI UNA POSSIBILE VIEW RIBASSISTA,"
    elif price.get("long_confirmed"):
        action_view_intro = "CON PREZZO RIALZISTA MA VIEW COT NON ANCORA CONFERMATA,"
    elif price.get("short_confirmed"):
        action_view_intro = "CON PREZZO RIBASSISTA MA VIEW COT NON ANCORA CONFERMATA,"
    else:
        action_view_intro = "CON UNA VIEW ANCORA NON CONFERMATA,"

    followup_detail = simple_detail
    removable_prefixes = (
        "Molti Fondi sono ancora posizionati Short, ma nell'ultimo report hanno iniziato a ridurre questa posizione. ",
        "Molti Fondi sono ancora posizionati Long, ma nell'ultimo report hanno iniziato a ridurre questa posizione. ",
        "Il quadro rimane ribassista, ma troppi operatori, soprattutto Fondi speculativi, si sono già posizionati nella stessa direzione Short. ",
        "Il quadro rimane rialzista, ma troppi operatori, soprattutto Fondi speculativi, si sono già posizionati nella stessa direzione Long. ",
        "Il quadro è rialzista. I Fondi stanno aumentando o mantenendo le posizioni Long e il prezzo settimanale conferma la stessa direzione. Questo significa che posizionamento e prezzo stanno lavorando insieme. ",
        "Il quadro è ribassista. I Fondi stanno aumentando o mantenendo le posizioni Short e il prezzo settimanale conferma la stessa direzione. Questo significa che posizionamento e prezzo stanno lavorando insieme. ",
        "I Fondi restano orientati Long, ma il prezzo settimanale sta scendendo. Le due letture non sono d'accordo. ",
        "I Fondi restano orientati Short, ma il prezzo settimanale sta salendo. Le due letture non sono d'accordo. ",
        "I Fondi stanno migliorando, ma il recupero non è ancora completo. ",
        "I Fondi stanno peggiorando, ma la fase ribassista non è ancora completa. ",
        "Il prezzo settimanale sta salendo, ma i Fondi non mostrano ancora una conferma altrettanto chiara. ",
        "Il prezzo settimanale sta scendendo, ma i Fondi non mostrano ancora una conferma altrettanto chiara. ",
    )
    for prefix in removable_prefixes:
        followup_detail = followup_detail.replace(prefix, "")

    if age_days <= 17:
        simple_detail = f"{clear_summary}\n\n{followup_detail}".strip()
        if alignment_bull_regime_confirmed:
            plain_action = "CAMBIO DI REGIME RIALZISTA CONFERMATO. CERCA EVENTUALI LONG SUI PULLBACK E NON INSEGUIRE IL PREZZO."
        elif alignment_bear_regime_confirmed:
            plain_action = "CAMBIO DI REGIME RIBASSISTA CONFERMATO. CERCA EVENTUALI SHORT SUI RIMBALZI E NON INSEGUIRE IL PREZZO."
        elif alignment_bull_regime_developing:
            plain_action = (
                f"LONG CONTRARIAN IN SVILUPPO. {alignment_trend_label.upper()} AUMENTANO I LONG, LA NET POSITION MIGLIORA E IL PREZZO È GIÀ RIALZISTA. ATTENDI CHE IL MIGLIORAMENTO SI ESTENDA ALLA STRUTTURA COT 3-6W PRIMA DI CONSIDERARE IL REGIME CONFERMATO."
                if price["long_confirmed"] else
                f"LONG CONTRARIAN IN SVILUPPO. {alignment_trend_label.upper()} AUMENTANO I LONG E LA NET POSITION MIGLIORA. NON INSEGUIRE NUOVI SHORT E ATTENDI LA CONFERMA RIALZISTA DEL PREZZO PRIMA DI VALUTARE LONG."
            )
        elif alignment_bear_regime_developing:
            plain_action = (
                f"SHORT CONTRARIAN IN SVILUPPO. {alignment_trend_label.upper()} AUMENTANO GLI SHORT, LA NET POSITION PEGGIORA E IL PREZZO È GIÀ RIBASSISTA. ATTENDI CHE IL PEGGIORAMENTO SI ESTENDA ALLA STRUTTURA COT 3-6W PRIMA DI CONSIDERARE IL REGIME CONFERMATO."
                if price["short_confirmed"] else
                f"SHORT CONTRARIAN IN SVILUPPO. {alignment_trend_label.upper()} AUMENTANO GLI SHORT E LA NET POSITION PEGGIORA. NON INSEGUIRE NUOVI LONG E ATTENDI LA CONFERMA RIBASSISTA DEL PREZZO PRIMA DI VALUTARE SHORT."
            )
        elif alignment_bull_longs_but_net_down:
            plain_action = "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE.\nI LONG SONO AUMENTATI, MA LA NET POSITION PEGGIORA\nPERCHÉ GLI SHORT CRESCONO DI PIÙ.\nNON INSEGUIRE NUOVI SHORT.\nATTENDI CHE NUOVI LONG E MIGLIORAMENTO DELLA NET POSITION\nSI PRESENTINO INSIEME, POI LA CONFERMA RIALZISTA DEL PREZZO."
        elif alignment_bear_shorts_but_net_up:
            plain_action = "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE.\nGLI SHORT SONO AUMENTATI, MA LA NET POSITION MIGLIORA\nPERCHÉ I LONG CRESCONO DI PIÙ.\nNON INSEGUIRE NUOVI LONG.\nATTENDI CHE NUOVI SHORT E PEGGIORAMENTO DELLA NET POSITION\nSI PRESENTINO INSIEME, POI LA CONFERMA RIBASSISTA DEL PREZZO."
        elif alignment_bull_short_covering_only:
            plain_action = "SHORT COVERING IN CORSO. LA NET POSITION MIGLIORA, MA I LONG NON AUMENTANO. AUMENTA IL RISCHIO DI RIMBALZO, MA NON È ANCORA UN CAMBIO DI REGIME LONG. ATTENDI NUOVI LONG."
        elif alignment_bull_long_liquidation_dominant:
            plain_action = "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE. LONG E SHORT DIMINUISCONO, MA LA NET POSITION PEGGIORA: DOMINA LA LONG LIQUIDATION. NON TRATTARLO COME SHORT COVERING RIALZISTA. ATTENDI NUOVI LONG E MIGLIORAMENTO DELLA NET POSITION."
        elif alignment_bear_long_liquidation_only:
            plain_action = "LONG LIQUIDATION IN CORSO. LA NET POSITION PEGGIORA, MA GLI SHORT NON AUMENTANO. AUMENTA IL RISCHIO DI DISCESA, MA NON È ANCORA UN CAMBIO DI REGIME SHORT. ATTENDI NUOVI SHORT."
        elif alignment_bear_short_covering_dominant:
            plain_action = "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE. LONG E SHORT DIMINUISCONO, MA LA NET POSITION MIGLIORA: DOMINA LO SHORT COVERING. NON TRATTARLO COME NUOVA PRESSIONE SHORT. ATTENDI NUOVI SHORT E PEGGIORAMENTO DELLA NET POSITION."
        elif alignment_bull_3:
            plain_action = f"POSSIBILE LONG CONTRARIAN IN COSTRUZIONE. NON INSEGUIRE NUOVI SHORT. ATTENDI CHE {alignment_trend_label.upper()} AUMENTINO I LONG E CHE LA NET POSITION MIGLIORI; POI ATTENDI LA CONFERMA RIALZISTA DEL PREZZO."
        elif alignment_bear_3:
            plain_action = f"POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE. NON INSEGUIRE NUOVI LONG. ATTENDI CHE {alignment_trend_label.upper()} AUMENTINO GLI SHORT E CHE LA NET POSITION PEGGIORI; POI ATTENDI LA CONFERMA RIBASSISTA DEL PREZZO."
        elif alignment_bull_2:
            plain_action = "SEGNALI LONG CONTRARIAN PARZIALI 2/3. MANTIENI PRUDENZA SUGLI SHORT E ATTENDI UN EVENTUALE ALIGNMENT 3/3 O UNA CONFERMA DEL PREZZO."
        elif alignment_bear_2:
            plain_action = "SEGNALI SHORT CONTRARIAN PARZIALI 2/3. MANTIENI PRUDENZA SUI LONG E ATTENDI UN EVENTUALE ALIGNMENT 3/3 O UNA CONFERMA DEL PREZZO."
        else:
            plain_action = f"{action_view_intro} {plain_action}"

    # Confronto FX Leveraged Funds / Asset Manager, solo nei dettagli completi.
    fx_asset_manager_comparison = ""
    if spec.is_fx and report_mode == "Completo":
        asset_net = float(cur.get("asset_net", math.nan))
        asset_net_prev = float(prev.get("asset_net", math.nan))
        asset_flow = asset_net - asset_net_prev if not pd.isna(asset_net) and not pd.isna(asset_net_prev) else math.nan
        trend_net = float(cur["trend_net"])
        if pd.isna(asset_net) or pd.isna(asset_flow):
            fx_side = "DATI NON DISPONIBILI"
            fx_flow = "CONFRONTO NON DISPONIBILE"
        else:
            fx_side = (
                "ENTRAMBE LE CATEGORIE SONO NET LONG" if trend_net > 0 and asset_net > 0
                else "ENTRAMBE LE CATEGORIE SONO NET SHORT" if trend_net < 0 and asset_net < 0
                else "LE CATEGORIE RESTANO SU LATI OPPOSTI" if trend_net * asset_net < 0
                else "POSIZIONAMENTO ATTUALE MISTO"
            )
            fx_flow = (
                "LE DUE CATEGORIE MIGLIORANO IL POSIZIONAMENTO" if trend_flow_1w > 0 and asset_flow > 0
                else "LE DUE CATEGORIE PEGGIORANO IL POSIZIONAMENTO" if trend_flow_1w < 0 and asset_flow < 0
                else "LEVERAGED FUNDS MIGLIORANO, ASSET MANAGER NON CONFERMANO" if trend_flow_1w > 0 and asset_flow <= 0
                else "LEVERAGED FUNDS PEGGIORANO, ASSET MANAGER NON CONFERMANO" if trend_flow_1w < 0 and asset_flow >= 0
                else "FLUSSI SETTIMANALI SOSTANZIALMENTE STABILI"
            )
        fx_asset_manager_comparison = (
            "\n\nCONFRONTO LEVERAGED FUNDS / ASSET MANAGER\n"
            f"Leveraged Funds\nLong: {fmt_number(float(cur['trend_long']))} | Short: {fmt_number(float(cur['trend_short']))}\n"
            f"Net: {fmt_number(float(cur['trend_net']), signed=True)} | Variazione 1W: {fmt_number(trend_flow_1w, signed=True)}\n"
            f"Asset Manager\nLong: {fmt_number(float(cur.get('asset_long', math.nan)))} | Short: {fmt_number(float(cur.get('asset_short', math.nan)))}\n"
            f"Net: {fmt_number(asset_net, signed=True)} | Variazione 1W: {fmt_number(asset_flow, signed=True)}\n"
            f"{fx_side}\n{fx_flow}"
        )

    category_flows: list[tuple[str, float]] = []
    if spec.family == "commodity":
        category_flows = [
            (spec.trend_label, trend_flow_1w),
            (spec.counter_label, counter_flow_1w),
            ("Swap Dealer", swap_flow_1w),
            ("Other Reportables", other_flow_1w),
            ("Nonreportable", nonreportable_flow_1w),
        ]
    elif spec.is_fx:
        category_flows = [
            (spec.trend_label, trend_flow_1w),
            (spec.counter_label, counter_flow_1w),
            ("Asset Manager", asset_manager_flow_1w),
            ("Other Reportables", other_flow_1w),
            ("Nonreportable", nonreportable_flow_1w),
        ]
    else:
        category_flows = [
            (spec.trend_label, trend_flow_1w),
            (spec.counter_label, counter_flow_1w),
            ("Dealer / Intermediary", dealer_flow_1w),
            ("Other Reportables", other_flow_1w),
            ("Nonreportable", nonreportable_flow_1w),
        ]

    all_category_flows_text = ""
    if show_all_category_flows and report_mode == "Completo":
        all_category_flows_text = "\n\nFLUSSI DI TUTTE LE CATEGORIE — ULTIMO REPORT\n" + "\n".join(
            f"{label}: {fmt_number(value, signed=True)}" for label, value in category_flows
        )

    full_diagnosis = (
        "PERCHÉ IL REPORT DÀ QUESTA LETTURA\n"
        f"{explanation}\n\n"
        "DATI DEI FONDI OSSERVATI\n"
        f"{spec.trend_label}\n"
        f"Ultimo report: {fmt_number(trend_flow_1w, signed=True)}\n"
        f"Ultime 3 settimane: {fmt_number(trend_flow_3w, signed=True)}\n"
        f"Ultime 6 settimane: {fmt_number(trend_flow_6w, signed=True)}\n"
        f"COT Index 26W: {fmt_decimal(cot_index_26w, 1)} su 100\n"
        f"COT Index 156W: {fmt_decimal(cot_index_156w, 1)} su 100\n"
        f"COT Index motore ({cot_lookback}W): {fmt_decimal(cot_index, 1)} su 100"
        + (f"\n{extreme_horizons}" if extreme_horizons else "")
        + "\n\nESPOSIZIONE DIREZIONALE DEI FONDI\n"
        f"Long: {fmt_pct(directional_long_pct)} | Short: {fmt_pct(directional_short_pct)}\n"
        "(Sono escluse le eventuali posizioni Spreading)"
        + (f"\n\n{future_regime_detail}" if future_regime_detail else "")
        + "\n\nALTRI OPERATORI OSSERVATI\n"
        f"{spec.counter_label}\n"
        f"Ultimo report: {fmt_number(counter_flow_1w, signed=True)}\n"
        f"Ultime 3 settimane: {fmt_number(counter_flow_3w, signed=True)}\n"
        f"Ultime 6 settimane: {fmt_number(counter_flow_6w, signed=True)}"
        f"{fx_asset_manager_comparison}{all_category_flows_text}\n\n"
        "PREZZO SETTIMANALE\n"
        f"{price.get('text', 'N/A')}\n"
        f"Chiusura: {fmt_decimal(float(price.get('close', math.nan)), 4)}\n"
        f"EMA21: {fmt_decimal(float(price.get('ema21', math.nan)), 4)}\n\n"
        "CONCENTRAZIONE DEI TOP 8 TRADER\n"
        f"{concentration_state}\n"
        f"Top 8 Long: {fmt_pct(float(cur['conc_long']))} | Percentile storico: {fmt_pct(conc_long_rank)}\n"
        f"Top 8 Short: {fmt_pct(float(cur['conc_short']))} | Percentile storico: {fmt_pct(conc_short_rank)}\n"
        f"{concentration_detail}"
    )

    if 10 < age_days <= 17:
        simple_detail += " Attenzione: il report COT potrebbe essere in ritardo e la situazione attuale potrebbe essere già cambiata."
        plain_action += " VERIFICA PRIMA SE È DISPONIBILE UN REPORT PIÙ RECENTE."

    # Le chiavi storiche restano disponibili per non rompere screener, export e AI.
    final_detail = simple_detail
    action = plain_action
    reason = explanation

    return {
        "available": True,
        "report_date": report_date,
        "previous_date": pd.Timestamp(prev["date"]).date(),
        "age_days": age_days,
        "freshness": freshness,
        "cftc_code": str(cur.get("cftc_code", "N/A")),
        "oi": float(cur["oi"]),
        "pct_delta_oi": pct_delta_oi,
        "pct_oi_3w": pct_oi_3w,
        "pct_oi_6w": pct_oi_6w,
        "oi_macro_available": oi_macro_available,
        "oi_macro_up": oi_macro_up,
        "oi_macro_down": oi_macro_down,
        "oi_macro_stable": oi_macro_stable,
        "oi_direction": oi_direction,
        "oi_direction_available": oi_direction_available,
        "oi_quality": oi_quality,
        "oi_report": f"{oi_quality}\nOI Index 52W: {fmt_decimal(oi_index_52w, 1)} | {oi_index_52w_state}",
        "oi_index_52w": oi_index_52w,
        "oi_index_52w_state": oi_index_52w_state,
        "trend_long": float(cur["trend_long"]),
        "trend_short": float(cur["trend_short"]),
        "counter_long": float(cur["counter_long"]),
        "counter_short": float(cur["counter_short"]),
        "trend_chg_l": trend_chg_l,
        "trend_chg_s": trend_chg_s,
        "counter_chg_l": counter_chg_l,
        "counter_chg_s": counter_chg_s,
        "trend_flow_1w": trend_flow_1w,
        "trend_flow_3w": trend_flow_3w,
        "trend_flow_6w": trend_flow_6w,
        "counter_flow_1w": counter_flow_1w,
        "counter_flow_3w": counter_flow_3w,
        "counter_flow_6w": counter_flow_6w,
        "cot_index": cot_index,
        "cot_index_26w": cot_index_26w,
        "cot_index_156w": cot_index_156w,
        "positioning": positioning,
        "positioning_value": positioning_value,
        "extreme_horizons": extreme_horizons,
        "directional_long_pct": directional_long_pct,
        "directional_short_pct": directional_short_pct,
        "current_position": current_position,
        "current_position_value": current_position_value,
        "last_report": last_report,
        "structure": structure,
        "structure_3w": structure_3w,
        "structure_6w": structure_6w,
        "counter_reading": counter_reading,
        "engine_bias": engine_bias,
        "engine_detail": engine_detail,
        "combined_bias": combined_bias,
        "final_bias": final_bias,
        "final_detail": final_detail,
        "simple_title": simple_title,
        "simple_detail": simple_detail,
        "plain_action": plain_action,
        "explanation": explanation,
        "action": action,
        "reason": reason,
        "new_long": new_long,
        "new_short": new_short,
        "short_covering": short_covering,
        "long_liquidation": long_liquidation,
        "confirmed_long": confirmed_long,
        "confirmed_short": confirmed_short,
        "partial_long": partial_long,
        "partial_short": partial_short,
        "partial_long_with_price": partial_long_with_price,
        "partial_short_with_price": partial_short_with_price,
        "crowded_long": crowded_long,
        "crowded_short": crowded_short,
        "partial_crowded_long": partial_crowded_long,
        "partial_crowded_short": partial_crowded_short,
        "alignment_available": alignment_available,
        "alignment_trend_index_156w": alignment_trend_index_156w,
        "alignment_counter_index_156w": alignment_counter_index_156w,
        "alignment_small_index_156w": alignment_small_index_156w,
        "alignment_bull_count": alignment_bull_count,
        "alignment_bear_count": alignment_bear_count,
        "alignment_bull_3": alignment_bull_3,
        "alignment_bear_3": alignment_bear_3,
        "alignment_bull_2": alignment_bull_2,
        "alignment_bear_2": alignment_bear_2,
        "alignment_bull_new_longs": alignment_bull_new_longs,
        "alignment_bear_new_shorts": alignment_bear_new_shorts,
        "alignment_bull_longs_but_net_down": alignment_bull_longs_but_net_down,
        "alignment_bear_shorts_but_net_up": alignment_bear_shorts_but_net_up,
        "alignment_bull_short_covering_only": alignment_bull_short_covering_only,
        "alignment_bull_long_liquidation_dominant": alignment_bull_long_liquidation_dominant,
        "alignment_bear_long_liquidation_only": alignment_bear_long_liquidation_only,
        "alignment_bear_short_covering_dominant": alignment_bear_short_covering_dominant,
        "alignment_bull_macro_ok": alignment_bull_macro_ok,
        "alignment_bear_macro_ok": alignment_bear_macro_ok,
        "alignment_bull_regime_confirmed": alignment_bull_regime_confirmed,
        "alignment_bear_regime_confirmed": alignment_bear_regime_confirmed,
        "alignment_bull_regime_developing": alignment_bull_regime_developing,
        "alignment_bear_regime_developing": alignment_bear_regime_developing,
        "future_regime_title": future_regime_title,
        "future_regime_text": future_regime_text,
        "future_regime_detail": future_regime_detail,
        "future_regime_summary": future_regime_summary,
        "conc_long": float(cur["conc_long"]) if not pd.isna(cur["conc_long"]) else math.nan,
        "conc_short": float(cur["conc_short"]) if not pd.isna(cur["conc_short"]) else math.nan,
        "conc_long_rank": conc_long_rank,
        "conc_short_rank": conc_short_rank,
        "concentration_available": concentration_available,
        "concentration_moderate": concentration_moderate,
        "concentration_state": concentration_state,
        "concentration_detail": concentration_detail,
        "clear_summary": clear_summary,
        "historical_range_summary": historical_range_summary,
        "action_view_intro": action_view_intro,
        "full_diagnosis": full_diagnosis,
        "category_flows": category_flows,
        "fx_asset_manager_comparison": fx_asset_manager_comparison,
    }


# =============================================================================
# COT ALIGNMENT MAP — CONTESTO STRUTTURALE
# =============================================================================
def alignment_zone(value: float) -> str:
    if pd.isna(value):
        return "N/A"
    if value >= ALIGNMENT_UPPER:
        return "ESTREMO ALTO"
    if value <= ALIGNMENT_LOWER:
        return "ESTREMO BASSO"
    if value > 50:
        return "SOPRA LA MEDIA"
    if value < 50:
        return "SOTTO LA MEDIA"
    return "NEUTRO"


def analyze_alignment_map(df: pd.DataFrame, spec: MarketSpec) -> dict[str, Any]:
    """Alignment contrarian fisso 156W, unificato come TradingView V1.5.36."""
    unavailable = {
        "available": False,
        "lookback": COT_INDEX_LONG_LOOKBACK,
        "speculative_index": math.nan,
        "counterparty_index": math.nan,
        "small_index": math.nan,
        "rapid_shift_6w": math.nan,
        "rapid_shift_state": "NON DISPONIBILE",
        "bull_score": 0,
        "bear_score": 0,
        "state": "DATI NON DISPONIBILI",
        "description": "Le serie necessarie all'Alignment contrarian 156W non sono disponibili.",
    }
    required = {"trend_index_156w", "counter_index_156w", "small_index_156w"}
    if df.empty or not required.issubset(df.columns):
        return unavailable

    cur = df.iloc[-1]
    speculative_index = float(cur["trend_index_156w"]) if not pd.isna(cur["trend_index_156w"]) else math.nan
    counterparty_index = float(cur["counter_index_156w"]) if not pd.isna(cur["counter_index_156w"]) else math.nan
    small_index = float(cur["small_index_156w"]) if not pd.isna(cur["small_index_156w"]) else math.nan
    rapid_shift_6w = float(cur["counter_rapid_shift_6w"]) if "counter_rapid_shift_6w" in cur and not pd.isna(cur["counter_rapid_shift_6w"]) else math.nan
    rapid_state = rapid_shift_state(rapid_shift_6w)
    if any(pd.isna(value) for value in (speculative_index, counterparty_index, small_index)):
        return unavailable

    # V1.5.36: stessa regola contrarian per tutte le famiglie.
    bull_speculative = speculative_index <= ALIGNMENT_LOWER
    bull_counterparty = counterparty_index >= ALIGNMENT_UPPER
    bull_small = small_index <= ALIGNMENT_LOWER
    bear_speculative = speculative_index >= ALIGNMENT_UPPER
    bear_counterparty = counterparty_index <= ALIGNMENT_LOWER
    bear_small = small_index >= ALIGNMENT_UPPER

    bull_score = int(bull_speculative) + int(bull_counterparty) + int(bull_small)
    bear_score = int(bear_speculative) + int(bear_counterparty) + int(bear_small)

    if bull_score == 3:
        state = "ALLINEAMENTO RIALZISTA 3/3"
    elif bear_score == 3:
        state = "ALLINEAMENTO RIBASSISTA 3/3"
    elif bull_score == 2 and bull_score > bear_score:
        state = "ALLINEAMENTO RIALZISTA 2/3"
    elif bear_score == 2 and bear_score > bull_score:
        state = "ALLINEAMENTO RIBASSISTA 2/3"
    elif bull_score == bear_score and bull_score >= 2:
        state = "SEGNALI MISTI"
    else:
        state = "NESSUN ALLINEAMENTO"

    trend_name = spec.trend_label
    counter_name = "Producer/Merchant" if spec.family == "commodity" else "Dealer" if spec.is_fx else "Asset Manager"
    if bull_score == 3:
        description = f"{trend_name} vicini al minimo 156W; {counter_name} vicini al massimo 156W; Small Traders vicini al minimo 156W."
    elif bear_score == 3:
        description = f"{trend_name} vicini al massimo 156W; {counter_name} vicini al minimo 156W; Small Traders vicini al massimo 156W."
    elif bull_score == 2 and bull_score > bear_score:
        description = "Segnali Long contrarian parziali 2/3: contesto interessante, ma non sufficiente per parlare di cambio di regime."
    elif bear_score == 2 and bear_score > bull_score:
        description = "Segnali Short contrarian parziali 2/3: contesto interessante, ma non sufficiente per parlare di cambio di regime."
    else:
        description = "Non emerge un Alignment contrarian 156W completo."

    return {
        "available": True,
        "lookback": COT_INDEX_LONG_LOOKBACK,
        "speculative_index": speculative_index,
        "counterparty_index": counterparty_index,
        "small_index": small_index,
        "speculative_zone": alignment_zone(speculative_index),
        "counterparty_zone": alignment_zone(counterparty_index),
        "small_zone": alignment_zone(small_index),
        "rapid_shift_6w": rapid_shift_6w,
        "rapid_shift_state": rapid_state,
        "bull_score": bull_score,
        "bear_score": bear_score,
        "state": state,
        "description": description,
    }


# =============================================================================
# TERM STRUCTURE MANUALE
# =============================================================================
def read_term_csv(source: Any) -> dict[str, str]:
    try:
        frame = pd.read_csv(source)
    except Exception:
        return {}
    required = {"root", "term_structure"}
    if not required.issubset(frame.columns):
        return {}
    result: dict[str, str] = {}
    for _, row in frame.iterrows():
        root = str(row["root"]).strip().upper()
        value = str(row["term_structure"]).strip()
        if root and value in TERM_OPTIONS:
            result[root] = value
    return result


def bundled_term_defaults() -> dict[str, str]:
    path = Path(__file__).with_name("term_structure.csv")
    return read_term_csv(path) if path.exists() else {}


def load_operational_prompt() -> tuple[str, str]:
    """Carica le istruzioni AI da PROMPT.TXT senza memorizzarle in cache."""
    prompt_path = Path(__file__).with_name(AI_PROMPT_FILENAME)
    try:
        prompt_text = prompt_path.read_text(encoding="utf-8").strip()
        if prompt_text:
            return prompt_text, AI_PROMPT_FILENAME
    except (OSError, UnicodeError):
        pass

    fallback = (
        "Scrivi in italiano con frasi semplici. Dichiara prima se il mercato è Long, "
        "Short, in costruzione oppure poco chiaro. Poi spiega la qualità dei flussi e "
        "indica quali conferme attendere. Non inventare dati o livelli mancanti."
    )
    return fallback, "prompt interno di emergenza"


# =============================================================================
# GRAFICI
# =============================================================================
def plot_cot_index(df: pd.DataFrame, label: str) -> go.Figure:
    chart_df = df.dropna(subset=["cot_index"]).tail(180)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=chart_df["date"], y=chart_df["cot_index"], mode="lines", name=label))
    fig.add_hrect(y0=80, y1=100, opacity=0.10, line_width=0)
    fig.add_hrect(y0=0, y1=20, opacity=0.10, line_width=0)
    fig.add_hline(y=80, line_dash="dash")
    fig.add_hline(y=50, line_dash="dot")
    fig.add_hline(y=20, line_dash="dash")
    fig.update_layout(title=f"COT Index — {label}", height=390, margin=dict(l=20, r=20, t=55, b=20), yaxis_range=[0, 100])
    return fig


def plot_net_positions(df: pd.DataFrame, trend_label: str, counter_label: str) -> go.Figure:
    chart_df = df.tail(180)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=chart_df["date"], y=chart_df["trend_net"], mode="lines", name=trend_label))
    fig.add_trace(go.Scatter(x=chart_df["date"], y=chart_df["counter_net"], mode="lines", name=counter_label))
    fig.add_hline(y=0, line_dash="dot")
    fig.update_layout(title="Net Position", height=390, margin=dict(l=20, r=20, t=55, b=20), legend=dict(orientation="h"))
    return fig


def plot_weekly_price(weekly: pd.DataFrame, ticker: str) -> go.Figure:
    chart_df = weekly.tail(104)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=chart_df.index, y=chart_df["close"], mode="lines", name="Close Weekly"))
    fig.add_trace(go.Scatter(x=chart_df.index, y=chart_df["ema21"], mode="lines", name="EMA21 Weekly"))
    fig.update_layout(title=f"Prezzo Weekly — {ticker}", height=390, margin=dict(l=20, r=20, t=55, b=20), legend=dict(orientation="h"))
    return fig



def secret_value(name: str, default: str) -> str:
    try:
        value = st.secrets.get(name, default)
        return str(value) if value not in (None, "") else default
    except Exception:
        return default


# =============================================================================
# SCREENER — CLASSIFICAZIONE, PUNTEGGIO ED EXPORT EXCEL
# =============================================================================
def screener_flow_type(smart: dict[str, Any]) -> str:
    if smart.get("new_long"):
        return "NUOVI LONG"
    if smart.get("new_short"):
        return "NUOVI SHORT"
    if smart.get("short_covering"):
        return "SHORT COVERING"
    if smart.get("long_liquidation"):
        return "LIQUIDAZIONE LONG"
    if smart.get("trend_flow_1w", 0) > 0:
        return "MIGLIORAMENTO MISTO"
    if smart.get("trend_flow_1w", 0) < 0:
        return "PEGGIORAMENTO MISTO"
    return "FLUSSO NEUTRALE"


SCREENER_STATUS_OPTIONS = [
    "LONG IN COSTRUZIONE",
    "LONG CONFERMATO",
    "LONG CONFERMATO — NON INSEGUIRE",
    "NEUTRALE / POCO CHIARO",
    "SHORT IN COSTRUZIONE",
    "SHORT CONFERMATO",
    "SHORT CONFERMATO — NON INSEGUIRE",
    "DATI COT DATATI — NON UTILIZZARE",
]


def screener_status(
    smart: dict[str, Any],
    price: dict[str, Any],
    alignment: dict[str, Any],
) -> tuple[str, str]:
    """Classificazione pubblica rigorosa, coerente con la gerarchia TradingView V1.5.36."""
    # 0) I dati troppo vecchi non possono produrre uno Stato operativo, anche se le
    #    vecchie serie storiche mostrano una direzione apparentemente chiara.
    if int(smart.get("age_days", 0) or 0) > 17:
        return "DATI COT DATATI — NON UTILIZZARE", "NEUTRALE"

    # 1) Configurazione COT già confermata, estrema e con prezzo concorde: non inseguire.
    if smart.get("crowded_long") and price.get("long_confirmed"):
        return "LONG CONFERMATO — NON INSEGUIRE", "LONG"
    if smart.get("crowded_short") and price.get("short_confirmed"):
        return "SHORT CONFERMATO — NON INSEGUIRE", "SHORT"

    # 2) Conferma completa. Un cambio di regime 156W diventa direzione principale
    #    soltanto quando è realmente confermato (Alignment 3/3 + flusso + struttura + prezzo).
    if smart.get("alignment_bull_regime_confirmed"):
        return "LONG CONFERMATO", "LONG"
    if smart.get("alignment_bear_regime_confirmed"):
        return "SHORT CONFERMATO", "SHORT"
    if smart.get("confirmed_long") and price.get("long_confirmed"):
        return "LONG CONFERMATO", "LONG"
    if smart.get("confirmed_short") and price.get("short_confirmed"):
        return "SHORT CONFERMATO", "SHORT"

    # 3) Un regime contrarian IN SVILUPPO può diventare Stato principale in costruzione.
    #    Un semplice 3/3 grezzo resta invece un warning separato nella colonna Regime 156W:
    #    non deve ribaltare da solo la direzione corrente dello screener.
    if smart.get("alignment_bull_regime_developing"):
        return "LONG IN COSTRUZIONE", "LONG"
    if smart.get("alignment_bear_regime_developing"):
        return "SHORT IN COSTRUZIONE", "SHORT"

    # 4) Struttura COT istituzionale già confermata ma prezzo non ancora concorde.
    if smart.get("confirmed_long"):
        return "LONG IN COSTRUZIONE", "LONG"
    if smart.get("confirmed_short"):
        return "SHORT IN COSTRUZIONE", "SHORT"

    # 5) Configurazioni istituzionali parziali: usa le condizioni
    #    smart_partial_long / smart_partial_short del motore TradingView V1.5.36.
    if smart.get("partial_long"):
        return "LONG IN COSTRUZIONE", "LONG"
    if smart.get("partial_short"):
        return "SHORT IN COSTRUZIONE", "SHORT"

    # 6) Persistenza direzionale forte, ma senza conferma istituzionale completa.
    #    Questo NON equivale a "confermato": serve soltanto a non classificare come
    #    neutrale un mercato in cui 1W, 3W, 6W e prezzo Weekly puntano tutti dalla
    #    stessa parte. È molto più severo del vecchio fallback V6.12/V6.13.
    flow_1w = float(smart.get("trend_flow_1w", 0) or 0)
    flow_3w = float(smart.get("trend_flow_3w", 0) or 0)
    flow_6w = float(smart.get("trend_flow_6w", 0) or 0)
    persistent_long = flow_1w > 0 and flow_3w > 0 and flow_6w > 0 and bool(price.get("long_confirmed"))
    persistent_short = flow_1w < 0 and flow_3w < 0 and flow_6w < 0 and bool(price.get("short_confirmed"))
    if persistent_long and not persistent_short:
        return "LONG IN COSTRUZIONE", "LONG"
    if persistent_short and not persistent_long:
        return "SHORT IN COSTRUZIONE", "SHORT"

    return "NEUTRALE / POCO CHIARO", "NEUTRALE"


def regime_156w_stage(smart: dict[str, Any]) -> str:
    if smart.get("alignment_bull_regime_confirmed"):
        return "CAMBIO DI REGIME RIALZISTA CONFERMATO"
    if smart.get("alignment_bear_regime_confirmed"):
        return "CAMBIO DI REGIME RIBASSISTA CONFERMATO"
    if smart.get("alignment_bull_regime_developing"):
        return "LONG CONTRARIAN IN SVILUPPO"
    if smart.get("alignment_bear_regime_developing"):
        return "SHORT CONTRARIAN IN SVILUPPO"
    if smart.get("alignment_bull_3"):
        return "POSSIBILE LONG CONTRARIAN IN COSTRUZIONE"
    if smart.get("alignment_bear_3"):
        return "POSSIBILE SHORT CONTRARIAN IN COSTRUZIONE"
    if smart.get("alignment_bull_2"):
        return "SEGNALI LONG CONTRARIAN PARZIALI 2/3"
    if smart.get("alignment_bear_2"):
        return "SEGNALI SHORT CONTRARIAN PARZIALI 2/3"
    if not smart.get("alignment_available"):
        return "DATI ALIGNMENT 156W NON DISPONIBILI"
    return "NESSUN CAMBIO DI REGIME CONTRARIAN EVIDENTE"


def calculate_screener_score(
    smart: dict[str, Any],
    price: dict[str, Any],
    alignment: dict[str, Any],
    oi_threshold: float,
) -> dict[str, Any]:
    """Score di qualità 0-100, separato dallo Stato e coerente con la direzione.

    V6.16 mantiene le correzioni dello Score V6.15 e completa la coerenza tra Stato, Regime 156W e qualità del dato:
    1) un flusso opposto alla Direzione non viene premiato;
    2) i mercati davvero NEUTRALI non ricevono punti solo perché l'ultimo report è forte;
    3) l'Open Interest 1W non viene contato due volte: NUOVI LONG/SHORT lo incorpora
       già, quindi lo Score OI usa la partecipazione 3-6W;
    4) una persistenza coerente 1W+3W+6W+prezzo può essere "IN COSTRUZIONE",
       ma con Score motore ridotto finché manca la conferma istituzionale completa.
    """
    status, direction = screener_status(smart, price, alignment)
    flow_type = screener_flow_type(smart)

    # STADIO DEL SETUP: misura la maturità, non decide la direzione.
    # I setup "in costruzione" nati soltanto da persistenza 1W+3W+6W+prezzo
    # ricevono un punteggio motore inferiore perché manca ancora la conferma
    # istituzionale della controparte richiesta dal motore specifico di mercato.
    persistence_only_long = (
        status == "LONG IN COSTRUZIONE"
        and not smart.get("confirmed_long")
        and not smart.get("partial_long")
        and not smart.get("alignment_bull_regime_developing")
    )
    persistence_only_short = (
        status == "SHORT IN COSTRUZIONE"
        and not smart.get("confirmed_short")
        and not smart.get("partial_short")
        and not smart.get("alignment_bear_regime_developing")
    )
    persistence_only = persistence_only_long or persistence_only_short

    if "CONFERMATO" in status:
        motor_score = 30
    elif "IN COSTRUZIONE" in status and persistence_only:
        motor_score = 12
    elif "IN COSTRUZIONE" in status:
        motor_score = 18
    else:
        motor_score = 0

    # FLUSSO ULTIMO REPORT: premia solo ciò che è coerente con la Direzione.
    # NUOVI LONG/SHORT includono già la conferma dell'OI 1W nel motore; per questo
    # l'OI 1W non riceve un secondo bonus separato più avanti.
    if direction == "LONG":
        flow_score = {
            "NUOVI LONG": 20,
            "MIGLIORAMENTO MISTO": 10,
            "SHORT COVERING": 5,
            "NUOVI SHORT": -20,
            "PEGGIORAMENTO MISTO": -10,
            "LIQUIDAZIONE LONG": -8,
        }.get(flow_type, 0)
    elif direction == "SHORT":
        flow_score = {
            "NUOVI SHORT": 20,
            "PEGGIORAMENTO MISTO": 10,
            "LIQUIDAZIONE LONG": 5,
            "NUOVI LONG": -20,
            "MIGLIORAMENTO MISTO": -10,
            "SHORT COVERING": -8,
        }.get(flow_type, 0)
    else:
        flow_score = 0

    # STRUTTURA 3-6W: conferma o contraddice la Direzione dello screener.
    flow_3w = float(smart.get("trend_flow_3w", 0) or 0)
    flow_6w = float(smart.get("trend_flow_6w", 0) or 0)
    if direction == "LONG":
        if flow_3w > 0 and flow_6w > 0:
            structure_score = 18
        elif flow_3w > 0 or flow_6w > 0:
            structure_score = 8
        elif flow_3w < 0 and flow_6w < 0:
            structure_score = -8
        else:
            structure_score = 0
        relevant_alignment = int(alignment.get("bull_score", 0))
        opposite_alignment = int(alignment.get("bear_score", 0))
        if smart.get("alignment_bull_regime_confirmed"):
            regime_score = 15
        elif smart.get("alignment_bull_regime_developing"):
            regime_score = 12
        elif smart.get("alignment_bull_3"):
            regime_score = 8
        elif smart.get("alignment_bull_2"):
            regime_score = 3
        else:
            regime_score = 0
    elif direction == "SHORT":
        if flow_3w < 0 and flow_6w < 0:
            structure_score = 18
        elif flow_3w < 0 or flow_6w < 0:
            structure_score = 8
        elif flow_3w > 0 and flow_6w > 0:
            structure_score = -8
        else:
            structure_score = 0
        relevant_alignment = int(alignment.get("bear_score", 0))
        opposite_alignment = int(alignment.get("bull_score", 0))
        if smart.get("alignment_bear_regime_confirmed"):
            regime_score = 15
        elif smart.get("alignment_bear_regime_developing"):
            regime_score = 12
        elif smart.get("alignment_bear_3"):
            regime_score = 8
        elif smart.get("alignment_bear_2"):
            regime_score = 3
        else:
            regime_score = 0
    else:
        structure_score = 0
        relevant_alignment = max(int(alignment.get("bull_score", 0)), int(alignment.get("bear_score", 0)))
        opposite_alignment = relevant_alignment
        regime_score = 0

    # Un Alignment opposto è un warning; un 2/3 da solo non crea una direzione.
    if opposite_alignment == 3 and opposite_alignment > relevant_alignment:
        regime_score -= 6
    elif opposite_alignment == 2 and opposite_alignment > relevant_alignment:
        regime_score -= 3
    alignment_score = regime_score  # alias mantenuto per compatibilità con export/storico.

    # PREZZO WEEKLY: conferma esterna al COT.
    if direction == "LONG":
        if price.get("long_confirmed"):
            price_score = 15
        elif price.get("above_ema"):
            price_score = 8
        elif price.get("short_confirmed"):
            price_score = -10
        else:
            price_score = 0
    elif direction == "SHORT":
        if price.get("short_confirmed"):
            price_score = 15
        elif price.get("below_ema"):
            price_score = 8
        elif price.get("long_confirmed"):
            price_score = -10
        else:
            price_score = 0
    else:
        price_score = 0

    # OPEN INTEREST: usa SOLO la partecipazione 3-6W per evitare il doppio conteggio.
    # L'OI Index 52W resta informativo, esattamente come nel motore TradingView.
    oi_quality = str(smart.get("oi_quality", "") or "")
    if direction == "LONG":
        if "RIALZISTA È SOSTENUTO" in oi_quality:
            oi_score = 10
        elif "RIALZISTA HA UNA PARTECIPAZIONE STABILE" in oi_quality:
            oi_score = 5
        elif "RIALZISTA PERDE PARTECIPAZIONE" in oi_quality:
            oi_score = -6
        elif "RIBASSISTA È SOSTENUTO" in oi_quality:
            oi_score = -10
        elif "RIBASSISTA HA UNA PARTECIPAZIONE STABILE" in oi_quality:
            oi_score = -5
        elif "RIBASSISTA PERDE PARTECIPAZIONE" in oi_quality:
            oi_score = 3
        elif "NON HA UNA DIREZIONE UNIFORME" in oi_quality:
            oi_score = -3
        else:
            oi_score = 0
    elif direction == "SHORT":
        if "RIBASSISTA È SOSTENUTO" in oi_quality:
            oi_score = 10
        elif "RIBASSISTA HA UNA PARTECIPAZIONE STABILE" in oi_quality:
            oi_score = 5
        elif "RIBASSISTA PERDE PARTECIPAZIONE" in oi_quality:
            oi_score = -6
        elif "RIALZISTA È SOSTENUTO" in oi_quality:
            oi_score = -10
        elif "RIALZISTA HA UNA PARTECIPAZIONE STABILE" in oi_quality:
            oi_score = -5
        elif "RIALZISTA PERDE PARTECIPAZIONE" in oi_quality:
            oi_score = 3
        elif "NON HA UNA DIREZIONE UNIFORME" in oi_quality:
            oi_score = -3
        else:
            oi_score = 0
    else:
        oi_score = 0

    # RISCHI / QUALITÀ DEL DATO. La concentrazione resta una fragilità, non una direzione.
    penalty = 0
    if "NON INSEGUIRE" in status:
        penalty -= 8
    elif smart.get("concentration_moderate"):
        penalty -= 3

    age_days = int(smart.get("age_days", 0) or 0)
    stale_hard_stop = age_days > 17
    if stale_hard_stop:
        penalty -= 100  # un dato obsoleto non deve mai salire nella classifica operativa
    elif age_days > 10:
        penalty -= 10

    total = max(0, min(100, motor_score + flow_score + structure_score + regime_score + price_score + oi_score + penalty))
    if total >= 80:
        quality = "MOLTO ALTA"
    elif total >= 65:
        quality = "ALTA"
    elif total >= 50:
        quality = "MEDIA"
    else:
        quality = "BASSA"

    return {
        "Stato": status,
        "Direzione": direction,
        "Tipo flusso": flow_type,
        "Qualità": quality,
        "Score": total,
        "Score motore": motor_score,
        "Score flusso": flow_score,
        "Score struttura": structure_score,
        "Score Alignment": alignment_score,
        "Score Regime 156W": regime_score,
        "Score prezzo": price_score,
        "Score OI": oi_score,
        "Penalità": penalty,
    }


def analyze_market_for_screener(
    spec: MarketSpec,
    cot_lookback: int,
    oi_threshold: float,
) -> tuple[dict[str, Any], pd.DataFrame]:
    history_limit = max(cot_lookback + 12, 180)
    rows, market_name, resolution = fetch_market_history(spec.specific_report, spec, history_limit)
    history_df, _ = build_history_df(rows, spec.specific_report, spec, cot_lookback)
    weekly_price, price_error = fetch_weekly_price(spec.yahoo_ticker)
    price = analyze_price(weekly_price)
    smart = analyze_smart_money(history_df, spec, price, oi_threshold, cot_lookback, "Compatto", False)
    alignment = analyze_alignment_map(history_df, spec)
    if not smart.get("available"):
        raise CFTCError(smart.get("final_detail", "Analisi non disponibile."))

    scoring = calculate_screener_score(smart, price, alignment, oi_threshold)
    matching_alignment = alignment.get("bull_score", 0) if scoring["Direzione"] == "LONG" else alignment.get("bear_score", 0) if scoring["Direzione"] == "SHORT" else max(alignment.get("bull_score", 0), alignment.get("bear_score", 0))
    price_confirmed = (
        scoring["Direzione"] == "LONG" and price.get("long_confirmed")
    ) or (
        scoring["Direzione"] == "SHORT" and price.get("short_confirmed")
    )

    row = {
        "Strumento": spec.label,
        "Root": spec.root,
        "Gruppo": spec.group,
        "Famiglia COT": spec.market_family_label,
        "Report": spec.specific_report,
        "Data COT": smart["report_date"].isoformat(),
        **scoring,
        "COT Index": round(float(smart["cot_index"]), 1) if not pd.isna(smart["cot_index"]) else math.nan,
        "COT Index 26W": round(float(smart["cot_index_26w"]), 1) if not pd.isna(smart["cot_index_26w"]) else math.nan,
        "COT Index 156W": round(float(smart["cot_index_156w"]), 1) if not pd.isna(smart["cot_index_156w"]) else math.nan,
        "Esposizione Long %": smart["directional_long_pct"],
        "Esposizione Short %": smart["directional_short_pct"],
        "Posizione attuale": smart["current_position"],
        "Posizionamento": smart["positioning"],
        "Flow 1W": smart["trend_flow_1w"],
        "Flow 3W": smart["trend_flow_3w"],
        "Flow 6W": smart["trend_flow_6w"],
        "Flow controparte 1W": smart["counter_flow_1w"],
        "Rapid Shift 6W": alignment.get("rapid_shift_6w", math.nan),
        "Stato Rapid Shift": alignment.get("rapid_shift_state", "NON DISPONIBILE"),
        "Variazione OI %": smart["pct_delta_oi"],
        "Partecipazione OI 3-6W": smart.get("oi_quality", "OI NON DISPONIBILE"),
        "OI Index 52W": smart.get("oi_index_52w", math.nan),
        "Stato OI 52W": smart.get("oi_index_52w_state", "NON DISPONIBILE"),
        "Prezzo Weekly": price["text"],
        "Prezzo confermato": bool(price_confirmed),
        "Alignment rialzista": int(alignment.get("bull_score", 0)),
        "Alignment ribassista": int(alignment.get("bear_score", 0)),
        "Alignment utile": int(matching_alignment),
        "Stato Alignment": alignment.get("state", "DATI NON DISPONIBILI"),
        "Regime 156W": regime_156w_stage(smart),
        "Dettaglio Regime 156W": smart.get("future_regime_text", ""),
        "Regime rialzista confermato": bool(smart.get("alignment_bull_regime_confirmed")),
        "Regime ribassista confermato": bool(smart.get("alignment_bear_regime_confirmed")),
        "Regime Long in sviluppo": bool(smart.get("alignment_bull_regime_developing")),
        "Regime Short in sviluppo": bool(smart.get("alignment_bear_regime_developing")),
        "Speculativi Index": alignment.get("speculative_index", math.nan),
        "Controparte Index": alignment.get("counterparty_index", math.nan),
        "Small Traders Index": alignment.get("small_index", math.nan),
        "Concentrazione Top 8": smart["concentration_state"],
        "Bias motore": smart["final_bias"],
        "Indicazione": smart["action"],
        "Motivazione": smart["reason"],
        "Mercato CFTC": market_name,
        "Risoluzione nome": resolution,
        "Ticker Yahoo": spec.yahoo_ticker,
        "Errore prezzo": price_error or "",
    }
    return row, history_df


def screener_signature(frame: pd.DataFrame) -> str:
    if frame.empty:
        return "empty"
    stable = frame.sort_values("Root").fillna("").astype(str)
    payload = stable.to_csv(index=False).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()[:20]


def build_screener_excel(results: pd.DataFrame, errors: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    ordered = results.sort_values(["Score", "Strumento"], ascending=[False, True]).reset_index(drop=True)
    display_columns = [
        "Strumento", "Stato", "Direzione", "Qualità", "Score", "Tipo flusso",
        "COT Index", "COT Index 26W", "COT Index 156W", "Posizione attuale", "Posizionamento", "Esposizione Long %", "Esposizione Short %",
        "Alignment rialzista", "Alignment ribassista", "Regime 156W", "Rapid Shift 6W",
        "Variazione OI %", "Partecipazione OI 3-6W", "OI Index 52W", "Prezzo Weekly", "Concentrazione Top 8", "Data COT",
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ordered[display_columns].to_excel(writer, sheet_name="Classifica generale", index=False)
        ordered[(ordered["Direzione"] == "LONG") & (ordered["Score"] >= 50)].to_excel(writer, sheet_name="Long interessanti", index=False)
        ordered[(ordered["Direzione"] == "SHORT") & (ordered["Score"] >= 50)].to_excel(writer, sheet_name="Short interessanti", index=False)
        ordered[ordered["Alignment utile"] >= 2].to_excel(writer, sheet_name="Alignment 2-3", index=False)
        ordered[ordered["Regime 156W"].str.contains("CONFERMATO", na=False)].to_excel(writer, sheet_name="Regime confermato", index=False)
        ordered[ordered["Regime 156W"].str.contains("IN SVILUPPO", na=False)].to_excel(writer, sheet_name="Contrarian sviluppo", index=False)
        ordered[ordered["Regime 156W"].str.contains("IN COSTRUZIONE", na=False)].to_excel(writer, sheet_name="Contrarian 3-3", index=False)
        ordered[ordered["Rapid Shift 6W"] >= RAPID_SHIFT_EXTREME].to_excel(writer, sheet_name="Rapid Shift rialzista", index=False)
        ordered[ordered["Rapid Shift 6W"] <= -RAPID_SHIFT_EXTREME].to_excel(writer, sheet_name="Rapid Shift ribassista", index=False)
        ordered[(ordered["OI Index 52W"] >= 80) | (ordered["OI Index 52W"] <= 20)].to_excel(writer, sheet_name="OI Index estremi", index=False)
        ordered[ordered["Stato"].str.contains("NON INSEGUIRE", na=False)].to_excel(writer, sheet_name="Non inseguire", index=False)
        ordered.to_excel(writer, sheet_name="Dati completi", index=False)
        errors.to_excel(writer, sheet_name="Errori", index=False)

        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column_cells in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells[:200]:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, min(len(value), 45))
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                ws.column_dimensions[column_letter].width = max(10, min(max_length + 2, 42))
            headers = {cell.value: cell.column for cell in ws[1]}
            if "Score" in headers and ws.max_row >= 2:
                col = get_column_letter(headers["Score"])
                ws.conditional_formatting.add(
                    f"{col}2:{col}{ws.max_row}",
                    ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=60, mid_color="FFEB84", end_type="num", end_value=100, end_color="63BE7B"),
                )
            for pct_header in ("Variazione OI %", "Rapid Shift 6W", "OI Index 52W"):
                if pct_header in headers:
                    col = get_column_letter(headers[pct_header])
                    for cell in ws[col][1:]:
                        cell.number_format = '0.00'

    return output.getvalue()



def _table_font(size: int, bold: bool = False) -> ImageFont.ImageFont:
    """Carica un font leggibile, con fallback al font predefinito di Pillow."""
    regular_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    )
    bold_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    )
    for candidate in bold_candidates if bold else regular_candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def _wrap_image_text(
    draw: ImageDraw.ImageDraw,
    text: Any,
    font: ImageFont.ImageFont,
    max_width: int,
) -> list[str]:
    """Divide il testo in righe che rientrano nella larghezza della cella."""
    value = "" if pd.isna(text) else str(text)
    if not value:
        return [""]

    lines: list[str] = []
    for paragraph in value.splitlines() or [value]:
        words = paragraph.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            width = draw.textbbox((0, 0), candidate, font=font)[2]
            if width <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines or [""]


def build_screener_jpg(frame: pd.DataFrame, export_label: str = "Tabella completa") -> bytes:
    """Crea un'immagine JPG della porzione selezionata della tabella visibile."""
    image_frame = frame.copy()
    if image_frame.empty:
        image_frame = pd.DataFrame({"Risultato": ["Nessun mercato visibile con i filtri correnti."]})

    preferred_widths = {
        "Posizione": 120,
        "Strumento": 270,
        "Stato": 360,
        "Qualità": 135,
        "Score": 105,
        "Tipo flusso": 205,
        "COT Index": 125,
        "COT Index 26W": 135,
        "COT Index 156W": 145,
        "Esposizione Long %": 160,
        "Esposizione Short %": 165,
        "Alignment rialzista": 145,
        "Alignment ribassista": 145,
        "Regime 156W": 430,
        "Rapid Shift 6W": 155,
        "Variazione OI %": 145,
        "Partecipazione OI 3-6W": 330,
        "OI Index 52W": 145,
        "Prezzo Weekly": 285,
        "Concentrazione Top 8": 590,
        "Data COT": 145,
    }
    columns = list(image_frame.columns)
    column_widths = [preferred_widths.get(column, 190) for column in columns]

    title_font = _table_font(38, bold=True)
    subtitle_font = _table_font(23)
    header_font = _table_font(22, bold=True)
    body_font = _table_font(21)

    margin = 28
    title_height = 100
    header_padding = 13
    cell_padding_x = 12
    cell_padding_y = 10
    line_height = 28
    table_width = sum(column_widths)
    canvas_width = table_width + margin * 2

    probe = Image.new("RGB", (canvas_width, 100), "white")
    probe_draw = ImageDraw.Draw(probe)

    wrapped_headers: list[list[str]] = []
    for column, width in zip(columns, column_widths):
        wrapped_headers.append(
            _wrap_image_text(probe_draw, column, header_font, width - 2 * cell_padding_x)
        )
    header_lines = max(len(lines) for lines in wrapped_headers)
    header_height = max(58, header_lines * line_height + 2 * header_padding)

    wrapped_rows: list[list[list[str]]] = []
    row_heights: list[int] = []
    for _, row in image_frame.iterrows():
        wrapped_cells: list[list[str]] = []
        max_lines = 1
        for column, width in zip(columns, column_widths):
            value = row[column]
            if column == "Score" and not pd.isna(value):
                value = f"{float(value):.0f}"
            elif column in ("COT Index", "COT Index 26W", "COT Index 156W", "Esposizione Long %", "Esposizione Short %") and not pd.isna(value):
                value = f"{float(value):.1f}"
            elif column == "Variazione OI %" and not pd.isna(value):
                value = f"{float(value):+.2f}"
            elif column == "Rapid Shift 6W" and not pd.isna(value):
                value = f"{float(value):+.1f}"
            elif column == "OI Index 52W" and not pd.isna(value):
                value = f"{float(value):.1f}"
            lines = _wrap_image_text(probe_draw, value, body_font, width - 2 * cell_padding_x)
            wrapped_cells.append(lines)
            max_lines = max(max_lines, len(lines))
        wrapped_rows.append(wrapped_cells)
        row_heights.append(max(52, max_lines * line_height + 2 * cell_padding_y))

    footer_height = 45
    canvas_height = title_height + header_height + sum(row_heights) + footer_height + margin * 2
    image = Image.new("RGB", (canvas_width, canvas_height), "#F5F7FA")
    draw = ImageDraw.Draw(image)

    draw.text((margin, margin), f"COT Screener — {export_label}", font=title_font, fill="#111827")
    subtitle = f"Righe esportate: {len(frame)}   |   Generato il {date.today().isoformat()}"
    draw.text((margin, margin + 51), subtitle, font=subtitle_font, fill="#4B5563")

    x = margin
    y = margin + title_height
    for width, lines in zip(column_widths, wrapped_headers):
        draw.rectangle((x, y, x + width, y + header_height), fill="#1F2937", outline="#4B5563", width=1)
        text_y = y + (header_height - len(lines) * line_height) / 2
        for line in lines:
            draw.text((x + cell_padding_x, text_y), line, font=header_font, fill="white")
            text_y += line_height
        x += width

    y += header_height
    for row_index, (wrapped_cells, row_height) in enumerate(zip(wrapped_rows, row_heights)):
        x = margin
        row_fill = "#FFFFFF" if row_index % 2 == 0 else "#EEF2F7"
        for column, width, lines in zip(columns, column_widths, wrapped_cells):
            fill = row_fill
            if column == "Stato":
                joined = " ".join(lines)
                if "NON INSEGUIRE" in joined:
                    fill = "#FEF3C7"
                elif "LONG CONFERMATO" in joined:
                    fill = "#DCFCE7"
                elif "SHORT CONFERMATO" in joined:
                    fill = "#FEE2E2"
                elif "IN COSTRUZIONE" in joined:
                    fill = "#E0F2FE"
            draw.rectangle((x, y, x + width, y + row_height), fill=fill, outline="#CBD5E1", width=1)
            text_y = y + cell_padding_y
            for line in lines:
                draw.text((x + cell_padding_x, text_y), line, font=body_font, fill="#111827")
                text_y += line_height
            x += width
        y += row_height

    draw.text(
        (margin, canvas_height - margin - 25),
        "Lo Score ordina la qualità complessiva e non rappresenta un segnale automatico di ingresso.",
        font=subtitle_font,
        fill="#4B5563",
    )

    output = io.BytesIO()
    image.save(output, format="JPEG", quality=92, optimize=True, subsampling=0)
    return output.getvalue()


def call_ai_provider(provider: str, model: str, prompt: str) -> str:
    if provider == "Google Gemini":
        from google import genai

        api_key = secret_value("GEMINI_API_KEY", "")
        if not api_key:
            raise KeyError("GEMINI_API_KEY")
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(model=model, contents=prompt)
        return response.text or "Nessuna risposta ricevuta da Gemini."

    from groq import Groq

    api_key = secret_value("GROQ_API_KEY", "")
    if not api_key:
        raise KeyError("GROQ_API_KEY")
    client = Groq(api_key=api_key)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "system",
                "content": (
                    "Sei un analista COT prudente. Usa esclusivamente i dati della classifica, "
                    "non inventare livelli tecnici e non trasformare il COT in un segnale immediato di ingresso. "
                    "Il COT Index descrive la collocazione della Net Position nel range storico e non indica da solo "
                    "che i Fondi siano Net Long o Net Short. Rispetta obbligatoriamente la colonna Classificazione AI: "
                    "ogni strumento deve comparire in una sola sezione operativa e non può essere riclassificato."
                ),
            },
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
    )
    return response.choices[0].message.content or "Nessuna risposta ricevuta da Groq."


def load_screener_prompt_template() -> str:
    prompt_path = Path(__file__).with_name(AI_SCREENER_PROMPT_FILENAME)
    try:
        text = prompt_path.read_text(encoding="utf-8").strip()
        if text:
            return text
    except (OSError, UnicodeError):
        pass
    return (
        "Analizza i migliori {TOP_N} risultati dello screener COT. "
        "Spiega in italiano semplice perché sono in cima alla classifica, senza modificare lo Score "
        "e senza inventare livelli tecnici.\n\nDATI SCREENER\n{DATI_SCREENER}"
    )



def screener_ai_classification(status: str) -> str:
    """Assegna una sezione AI esclusiva in base allo Stato deterministico."""
    mapping = {
        "LONG CONFERMATO": "OPPORTUNITÀ LONG CONFERMATA",
        "SHORT CONFERMATO": "OPPORTUNITÀ SHORT CONFERMATA",
        "LONG IN COSTRUZIONE": "DA MONITORARE — LONG IN COSTRUZIONE",
        "SHORT IN COSTRUZIONE": "DA MONITORARE — SHORT IN COSTRUZIONE",
        "LONG CONFERMATO — NON INSEGUIRE": "NON INSEGUIRE — LONG ESTESO",
        "SHORT CONFERMATO — NON INSEGUIRE": "NON INSEGUIRE — SHORT ESTESO",
        "NEUTRALE / POCO CHIARO": "POCO CHIARO — NESSUN VANTAGGIO OPERATIVO",
        "DATI COT DATATI — NON UTILIZZARE": "DATI DATATI — NON UTILIZZARE OPERATIVAMENTE",
    }
    return mapping.get(str(status), "POCO CHIARO — NESSUN VANTAGGIO OPERATIVO")


def build_screener_ai_prompt(top_rows: pd.DataFrame, top_n: int) -> str:
    ai_rows = top_rows.copy()
    ai_rows["Classificazione AI"] = ai_rows["Stato"].map(screener_ai_classification)
    compact_columns = [
        "Strumento", "Stato", "Classificazione AI", "Direzione", "Qualità", "Score", "Tipo flusso",
        "COT Index", "COT Index 26W", "COT Index 156W", "Posizionamento", "Posizione attuale", "Esposizione Long %", "Esposizione Short %",
        "Flow 1W", "Flow 3W", "Flow 6W", "Rapid Shift 6W", "Stato Rapid Shift",
        "Variazione OI %", "Partecipazione OI 3-6W", "OI Index 52W", "Stato OI 52W",
        "Alignment rialzista", "Alignment ribassista", "Regime 156W", "Dettaglio Regime 156W", "Prezzo Weekly", "Concentrazione Top 8",
        "Indicazione", "Motivazione",
    ]
    table_text = ai_rows[compact_columns].to_csv(index=False)
    template = load_screener_prompt_template()
    return (
        template.replace("{TOP_N}", str(top_n))
        .replace("{DATI_SCREENER}", table_text)
        .strip()
    )


def render_screener() -> None:
    st.header("COT Screener — tutti i mercati")
    st.caption(
        "La scansione parte soltanto su richiesta. Ogni mercato usa lo stesso motore Smart Money, "
        "la stessa Alignment Map e la stessa conferma prezzo dell'analisi singola."
    )

    with st.sidebar:
        st.header("Impostazioni Screener")
        all_groups = sorted({m.group for m in MARKETS})
        selected_groups = st.multiselect("Famiglie da analizzare", all_groups, default=all_groups)
        available_markets = [m for m in MARKETS if m.group in selected_groups]
        selected_labels = st.multiselect(
            "Mercati",
            [m.label for m in available_markets],
            default=[m.label for m in available_markets],
            help="Puoi ridurre la selezione per velocizzare la prima prova.",
        )
        cot_lookback = st.selectbox(
            "COT Index lookback Screener",
            [26, 52, 156, 260],
            index=2,
            format_func=lambda x: f"{x} settimane",
            key="screener_lookback",
        )
        oi_threshold = st.number_input(
            "Soglia Open Interest Screener (%)",
            min_value=0.0,
            max_value=10.0,
            value=0.5,
            step=0.1,
            key="screener_oi_threshold",
        )
        run_scan = st.button("Avvia analisi di tutti i mercati selezionati", type="primary", width="stretch")
        st.caption("La prima scansione completa può richiedere alcuni minuti. I dati vengono poi memorizzati nella cache.")

    if run_scan:
        selected_specs = [MARKET_BY_LABEL[label] for label in selected_labels]
        if not selected_specs:
            st.warning("Seleziona almeno un mercato.")
        else:
            rows: list[dict[str, Any]] = []
            errors: list[dict[str, str]] = []
            progress = st.progress(0.0, text="Avvio screener...")
            status_box = st.empty()
            for index, spec in enumerate(selected_specs, start=1):
                status_box.info(f"Analisi {index}/{len(selected_specs)} — {spec.label}")
                try:
                    row, _ = analyze_market_for_screener(spec, int(cot_lookback), float(oi_threshold))
                    rows.append(row)
                except Exception as exc:
                    errors.append({"Strumento": spec.label, "Root": spec.root, "Errore": str(exc)})
                progress.progress(index / len(selected_specs), text=f"Completati {index}/{len(selected_specs)} mercati")
            progress.empty()
            status_box.empty()

            results_df = pd.DataFrame(rows)
            errors_df = pd.DataFrame(errors, columns=["Strumento", "Root", "Errore"])
            if not results_df.empty:
                results_df = results_df.sort_values(["Score", "Strumento"], ascending=[False, True]).reset_index(drop=True)
                results_df.insert(0, "Posizione", range(1, len(results_df) + 1))
            st.session_state["screener_results"] = results_df
            st.session_state["screener_errors"] = errors_df
            st.session_state["screener_signature"] = screener_signature(results_df)
            for key in ("screener_ai_answer", "screener_ai_context"):
                st.session_state.pop(key, None)

    results_df = st.session_state.get("screener_results", pd.DataFrame())
    errors_df = st.session_state.get("screener_errors", pd.DataFrame(columns=["Strumento", "Root", "Errore"]))
    if results_df.empty:
        st.info("Premi il pulsante nella sidebar per creare la prima classifica.")
        return

    st.success(f"Screener completato: {len(results_df)} mercati analizzati; {len(errors_df)} errori o dati mancanti.")

    metric1, metric2, metric3, metric4, metric5 = st.columns(5)
    metric1.metric("Mercati analizzati", len(results_df))
    metric2.metric("Score ≥ 65", int((results_df["Score"] >= 65).sum()))
    metric3.metric("Long", int((results_df["Direzione"] == "LONG").sum()))
    metric4.metric("Short", int((results_df["Direzione"] == "SHORT").sum()))
    metric5.metric("Regime 156W confermato", int(results_df["Regime 156W"].str.contains("CONFERMATO", na=False).sum()))

    st.subheader("Filtri classifica")
    f1, f2, f3, f4 = st.columns(4)
    with f1:
        direction_filter = st.selectbox("Direzione", ["TUTTI", "LONG", "SHORT", "NEUTRALE"])
    with f2:
        min_score = st.slider("Score minimo", 0, 100, 0, 5)
    with f3:
        min_alignment = st.selectbox("Alignment 156W minimo", [0, 1, 2, 3], index=0, format_func=lambda x: f"{x}/3")
    with f4:
        price_only = st.toggle("Solo prezzo confermato", value=False)

    # Mostra sempre tutti gli stati possibili, anche quando uno stato non è presente
    # nei risultati della scansione corrente.
    status_options = SCREENER_STATUS_OPTIONS
    selected_statuses = st.multiselect("Stati", status_options, default=status_options)
    exclude_weak = st.toggle("Escludi short covering, liquidazione Long e flussi neutrali", value=False)

    extra1, extra2, extra3 = st.columns(3)
    with extra1:
        rapid_filter = st.selectbox(
            "Filtro Rapid Shift 6W",
            ["TUTTI", "RAPIDO RIALZISTA ≥ +40", "RAPIDO RIBASSISTA ≤ -40", "QUALSIASI MOVIMENTO RAPIDO: ≥ +40 O ≤ -40"],
        )
    with extra2:
        oi_index_filter = st.selectbox(
            "Filtro OI Index 52W",
            ["TUTTI", "MOLTO ALTO ≥ 80", "MOLTO BASSO ≤ 20", "ALTO ≥ 60", "BASSO ≤ 40"],
        )
    with extra3:
        regime_filter = st.selectbox(
            "Filtro cambio di regime 156W",
            ["TUTTI", "CONFERMATO", "IN SVILUPPO", "POSSIBILE 3/3 IN COSTRUZIONE", "SEGNALI 2/3", "NESSUNO"],
        )

    st.caption(
        "Rapid Shift 6W: ≥ +40 seleziona solo accelerazioni rialziste; ≤ -40 solo accelerazioni ribassiste; "
        "l'opzione 'qualsiasi movimento rapido' include entrambe le direzioni."
    )

    filtered = results_df.copy()
    if direction_filter != "TUTTI":
        filtered = filtered[filtered["Direzione"] == direction_filter]
    filtered = filtered[filtered["Score"] >= min_score]
    filtered = filtered[filtered["Alignment utile"] >= min_alignment]
    filtered = filtered[filtered["Stato"].isin(selected_statuses)]
    if price_only:
        filtered = filtered[filtered["Prezzo confermato"]]
    if exclude_weak:
        filtered = filtered[~filtered["Tipo flusso"].isin(["SHORT COVERING", "LIQUIDAZIONE LONG", "FLUSSO NEUTRALE"])]
    if rapid_filter == "RAPIDO RIALZISTA ≥ +40":
        filtered = filtered[filtered["Rapid Shift 6W"] >= RAPID_SHIFT_EXTREME]
    elif rapid_filter == "RAPIDO RIBASSISTA ≤ -40":
        filtered = filtered[filtered["Rapid Shift 6W"] <= -RAPID_SHIFT_EXTREME]
    elif rapid_filter == "QUALSIASI MOVIMENTO RAPIDO: ≥ +40 O ≤ -40":
        filtered = filtered[filtered["Rapid Shift 6W"].abs() >= RAPID_SHIFT_EXTREME]
    if oi_index_filter == "MOLTO ALTO ≥ 80":
        filtered = filtered[filtered["OI Index 52W"] >= 80.0]
    elif oi_index_filter == "MOLTO BASSO ≤ 20":
        filtered = filtered[filtered["OI Index 52W"] <= 20.0]
    elif oi_index_filter == "ALTO ≥ 60":
        filtered = filtered[filtered["OI Index 52W"] >= 60.0]
    elif oi_index_filter == "BASSO ≤ 40":
        filtered = filtered[filtered["OI Index 52W"] <= 40.0]

    if regime_filter == "CONFERMATO":
        filtered = filtered[filtered["Regime 156W"].str.contains("CONFERMATO", na=False)]
    elif regime_filter == "IN SVILUPPO":
        filtered = filtered[filtered["Regime 156W"].str.contains("IN SVILUPPO", na=False)]
    elif regime_filter == "POSSIBILE 3/3 IN COSTRUZIONE":
        filtered = filtered[filtered["Regime 156W"].str.contains("IN COSTRUZIONE", na=False)]
    elif regime_filter == "SEGNALI 2/3":
        filtered = filtered[filtered["Regime 156W"].str.contains("2/3", na=False)]
    elif regime_filter == "NESSUNO":
        filtered = filtered[filtered["Regime 156W"] == "NESSUN CAMBIO DI REGIME CONTRARIAN EVIDENTE"]

    visible_columns = [
        "Posizione", "Strumento", "Stato", "Qualità", "Score", "Tipo flusso",
        "COT Index 26W", "COT Index 156W", "Alignment rialzista", "Alignment ribassista", "Regime 156W", "Rapid Shift 6W",
        "Variazione OI %", "Partecipazione OI 3-6W", "OI Index 52W", "Prezzo Weekly", "Concentrazione Top 8", "Data COT",
    ]
    display_df = filtered[visible_columns].copy()
    display_df["Alignment rialzista"] = display_df["Alignment rialzista"].astype(int).astype(str) + "/3"
    display_df["Alignment ribassista"] = display_df["Alignment ribassista"].astype(int).astype(str) + "/3"
    st.dataframe(
        display_df,
        width="stretch",
        hide_index=True,
        column_config={
            "Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d"),
            "COT Index 26W": st.column_config.NumberColumn("COT 26W", format="%.1f"),
            "COT Index 156W": st.column_config.NumberColumn("COT 156W", format="%.1f"),
            "Rapid Shift 6W": st.column_config.NumberColumn("Rapid Shift 6W", format="%+.1f"),
            "Variazione OI %": st.column_config.NumberColumn("Δ OI %", format="%+.2f"),
            "OI Index 52W": st.column_config.NumberColumn("OI Index 52W", format="%.1f"),
            "Alignment rialzista": "Bull",
            "Alignment ribassista": "Bear",
            "Regime 156W": st.column_config.TextColumn("Regime 156W", width="large"),
            "Concentrazione Top 8": st.column_config.TextColumn("Concentrazione Top 8", width="large"),
        },
    )
    st.caption(f"Risultati visibili: {len(filtered)} su {len(results_df)}. Lo Score ordina la qualità complessiva, non è un segnale automatico di ingresso.")

    with st.expander("Come viene costruito lo Score", expanded=False):
        st.write(
            "Lo Score misura la qualità del setup nella Direzione indicata dallo Stato. Premia soltanto i flussi coerenti, la struttura 3–6W, "
            "il regime 156W, il prezzo Weekly e la partecipazione OI 3–6W. Un flusso opposto viene penalizzato. L’OI 1W non viene contato due volte: "
            "è già incluso nella classificazione NUOVI LONG/NUOVI SHORT. I mercati NEUTRALI non ricevono bonus direzionali e i dati molto obsoleti hanno Score 0."
        )
        component_columns = [
            "Strumento", "Score", "Score motore", "Score flusso", "Score struttura",
            "Score Regime 156W", "Score prezzo", "Score OI", "Penalità",
        ]
        st.dataframe(filtered[component_columns], width="stretch", hide_index=True)

    excel_bytes = build_screener_excel(results_df, errors_df)

    # Le tre immagini rispettano i filtri correnti. Top 5 e Top 10 sono
    # semplicemente le prime righe della classifica attualmente visibile.
    jpg_top5_bytes = build_screener_jpg(display_df.head(5), "Top 5 risultati visibili")
    jpg_top10_bytes = build_screener_jpg(display_df.head(10), "Top 10 risultati visibili")
    jpg_total_bytes = build_screener_jpg(display_df, "Tabella completa visibile")

    export_col1, export_col2, export_col3, export_col4 = st.columns(4)
    with export_col1:
        st.download_button(
            "Scarica Screener Excel",
            data=excel_bytes,
            file_name=f"cot_screener_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    with export_col2:
        st.download_button(
            "Scarica JPG Top 5",
            data=jpg_top5_bytes,
            file_name=f"cot_screener_top5_{date.today().isoformat()}.jpg",
            mime="image/jpeg",
            width="stretch",
        )
    with export_col3:
        st.download_button(
            "Scarica JPG Top 10",
            data=jpg_top10_bytes,
            file_name=f"cot_screener_top10_{date.today().isoformat()}.jpg",
            mime="image/jpeg",
            width="stretch",
        )
    with export_col4:
        st.download_button(
            "Scarica JPG Totale",
            data=jpg_total_bytes,
            file_name=f"cot_screener_totale_{date.today().isoformat()}.jpg",
            mime="image/jpeg",
            width="stretch",
        )

    if not errors_df.empty:
        with st.expander(f"Errori o dati mancanti ({len(errors_df)})", expanded=False):
            st.dataframe(errors_df, width="stretch", hide_index=True)

    st.divider()
    st.subheader("Spiegazione AI dei migliori risultati")
    st.caption("L'AI interviene una sola volta, dopo la classifica, e non modifica lo Score deterministico.")

    ai1, ai2, ai3 = st.columns([1, 1, 1])
    with ai1:
        top_n = st.radio("Numero di strumenti", [5, 10], horizontal=True, key="screener_ai_top_n")
    with ai2:
        provider = st.selectbox("Provider AI Screener", ["Google Gemini", "Groq"], key="screener_ai_provider")
    with ai3:
        if provider == "Google Gemini":
            model = st.text_input("Modello Gemini Screener", value=secret_value("GEMINI_MODEL", "gemini-3.5-flash"), key="screener_ai_model_gemini")
        else:
            model = st.text_input("Modello Groq Screener", value=secret_value("GROQ_MODEL", "openai/gpt-oss-120b"), key="screener_ai_model_groq")

    ai_source = filtered if not filtered.empty else results_df
    top_rows = ai_source.sort_values(["Score", "Strumento"], ascending=[False, True]).head(int(top_n))
    ai_context = f"{st.session_state.get('screener_signature', '')}|{top_n}|{provider}|{model}|" + ",".join(top_rows["Root"].astype(str))

    if st.button(f"Spiega con AI i primi {min(int(top_n), len(top_rows))} risultati", type="primary", width="stretch"):
        if top_rows.empty:
            st.warning("Nessun risultato disponibile con i filtri correnti.")
        else:
            prompt = build_screener_ai_prompt(top_rows, min(int(top_n), len(top_rows)))
            with st.spinner(f"Interrogo {provider}..."):
                try:
                    answer = call_ai_provider(provider, model, prompt)
                    st.session_state["screener_ai_answer"] = answer
                    st.session_state["screener_ai_context"] = ai_context
                except KeyError as exc:
                    st.error(f"Chiave API mancante nei Secrets di Streamlit: {exc}.")
                except ModuleNotFoundError as exc:
                    st.error(f"Dipendenza AI non installata: {exc}.")
                except Exception as exc:
                    st.error(f"Errore durante la comunicazione con l'AI: {exc}")

    if st.session_state.get("screener_ai_answer") and st.session_state.get("screener_ai_context") == ai_context:
        st.markdown("### Risposta dell'AI sullo Screener")
        st.write(st.session_state["screener_ai_answer"])

    st.divider()
    st.warning(
        "Lo screener serve a stabilire quali mercati meritano un approfondimento. "
        "Prima di operare apri sempre il singolo strumento e verifica il prezzo sul grafico."
    )




def render_single_analysis() -> None:
    # =============================================================================
    # SIDEBAR
    # =============================================================================
    st.header("Analisi singolo strumento")
    st.caption("Approfondimento completo di un solo future, con interrogazione AI facoltativa e risposta soltanto a video.")

    with st.sidebar:
        st.header("Impostazioni")
        group = st.selectbox("Famiglia", ["Tutti"] + sorted({m.group for m in MARKETS}))
        filtered = [m for m in MARKETS if group == "Tutti" or m.group == group]
        selected_label = st.selectbox("Mercato", [m.label for m in filtered])
        spec = MARKET_BY_LABEL[selected_label]

        cot_lookback = st.selectbox("COT Index lookback", [26, 52, 156, 260], index=2, format_func=lambda x: f"{x} settimane")
        oi_threshold = st.number_input("Soglia Open Interest (%)", min_value=0.0, max_value=10.0, value=0.5, step=0.1)
        yahoo_ticker = st.text_input("Ticker prezzo Yahoo", value=spec.yahoo_ticker)

        st.divider()
        report_mode = st.selectbox(
            "Contenuto Report",
            ["Compatto", "Completo"],
            index=0,
            help="Compatto mostra il quadro generale. Completo aggiunge i dettagli numerici della lettura.",
        )
        show_all_category_flows = st.toggle(
            "Mostra flussi di tutte le categorie",
            value=False,
            disabled=report_mode != "Completo",
            help="Aggiunge la variazione settimanale delle categorie non già usate dal motore principale.",
        )
        show_debug = st.toggle("Mostra diagnostica campi", value=False)

        st.divider()
        st.subheader("Term Structure")
        if spec.family != "commodity":
            term_structure = "Non applicabile"
            term_usage_status = "NON APPLICABILE"
            st.info("NON APPLICABILE")
            st.caption("Non viene utilizzata per indici, valute, tassi, volatilità e crypto CME.")
        else:
            term_usage_status = "OPZIONALE"
            st.info("OPZIONALE")
            st.caption("Non entra nel responso Smart Money né nell'Alignment Map. Puoi lasciarla su Non disponibile.")

            term_defaults = bundled_term_defaults()
            uploaded_term_file = st.file_uploader("Carica CSV opzionale", type=["csv"])
            if uploaded_term_file is not None:
                term_defaults.update(read_term_csv(uploaded_term_file))

            default_term = term_defaults.get(spec.root, "Non disponibile")
            default_index = TERM_OPTIONS.index(default_term) if default_term in TERM_OPTIONS else 0
            term_structure = st.selectbox("Curva M1–M2", TERM_OPTIONS, index=default_index)
            st.caption("Per salvare un valore permanente modifica term_structure.csv nel repository.")


    # =============================================================================
    # RECUPERO DATI
    # =============================================================================
    history_limit = max(cot_lookback + 12, 180)

    try:
        with st.spinner("Recupero report CFTC e prezzo Weekly..."):
            specific_rows, specific_market_name, specific_resolution = fetch_market_history(
                spec.specific_report, spec, history_limit
            )
            specific_df, specific_fields = build_history_df(
                specific_rows, spec.specific_report, spec, cot_lookback
            )

            weekly_price, price_error = fetch_weekly_price(yahoo_ticker)
    except CFTCError as exc:
        st.error(f"Errore CFTC: {exc}")
        st.stop()
    except Exception as exc:
        st.error(f"Errore inatteso durante il recupero dei dati: {exc}")
        st.stop()

    price_analysis = analyze_price(weekly_price)
    smart = analyze_smart_money(
        specific_df, spec, price_analysis, float(oi_threshold), cot_lookback,
        report_mode, show_all_category_flows,
    )
    alignment = analyze_alignment_map(specific_df, spec)

    if not smart.get("available"):
        st.error(smart.get("final_detail", "Analisi non disponibile."))
        st.stop()


    # Una risposta AI appartiene sempre a un solo strumento e a una sola data COT.
    # Se cambia il mercato oppure arriva un nuovo report, la vecchia interrogazione viene eliminata.
    current_ai_data_context = f"{spec.root}|{smart['report_date'].isoformat()}"
    if st.session_state.get("_cot_ai_data_context") != current_ai_data_context:
        for state_key in ("cot_ai_answer", "cot_ai_context", "cot_ai_question"):
            st.session_state.pop(state_key, None)
        st.session_state["_cot_ai_data_context"] = current_ai_data_context


    # =============================================================================
    # INTESTAZIONE DATI
    # =============================================================================
    st.success(
        f"{spec.label} | Report {spec.specific_report} | posizioni al {smart['report_date'].strftime('%d/%m/%Y')} "
        f"| mercato CFTC: {specific_market_name} ({specific_resolution})."
    )
    if price_error:
        st.warning(price_error)

    col1, col2, col3, col4, col5, col6 = st.columns(6)
    col1.metric("COT Index 26W", fmt_decimal(smart["cot_index_26w"], 1), smart.get("extreme_horizons", "") or None)
    col2.metric("COT Index 156W", fmt_decimal(smart["cot_index_156w"], 1), smart["positioning"])
    col3.metric("Open Interest", fmt_number(smart["oi"]), fmt_pct(smart["pct_delta_oi"], signed=True, digits=2))
    col4.metric(
        "OI Index 52W",
        fmt_decimal(smart["oi_index_52w"], 1),
        smart["oi_index_52w_state"],
    )
    col5.metric(f"Flow {spec.trend_label} 1W", fmt_number(smart["trend_flow_1w"], signed=True), f"3W {fmt_number(smart['trend_flow_3w'], signed=True)}")
    col6.metric(f"Flow {spec.counter_label} 1W", fmt_number(smart["counter_flow_1w"], signed=True), f"3W {fmt_number(smart['counter_flow_3w'], signed=True)}")


    # =============================================================================
    # RESPONSO PRINCIPALE
    # =============================================================================
    st.header("1. Responso Smart Money")
    main_accent = accent_for_state(smart["simple_title"])
    structure_html = smart["structure"].replace("\n", "<br>")
    positioning_html = smart["positioning_value"].replace("\n", "<br>")
    current_position_html = smart["current_position_value"].replace("\n", "<br>")
    oi_report_html = smart["oi_report"].replace("\n", "<br>")

    # Nota didattica: il possibile cambio di regime 156W è un modulo anticipatore
    # separato dalla direzione principale. Serve a evitare che un 2/3 o un 3/3
    # grezzo venga interpretato come un'inversione già avvenuta.
    future_regime_title = str(smart.get("future_regime_title", ""))
    if "IN COSTRUZIONE" in future_regime_title:
        future_regime_teaching = (
            "<br><br><b>Come leggerlo:</b> questo è un segnale anticipatore separato dalla direzione principale. "
            "Non modifica il responso Smart Money finché non passa almeno a <b>IN SVILUPPO</b>."
        )
    elif "PARZIALI" in future_regime_title:
        future_regime_teaching = (
            "<br><br><b>Come leggerlo:</b> sono segnali contrarian ancora incompleti. "
            "Non modificano la direzione principale e non indicano ancora un cambio di regime."
        )
    elif "IN SVILUPPO" in future_regime_title:
        future_regime_teaching = (
            "<br><br><b>Come leggerlo:</b> il possibile cambio di regime ha compiuto un passo ulteriore, "
            "ma non è ancora un cambio di regime confermato."
        )
    elif "CONFERMATO" in future_regime_title:
        future_regime_teaching = (
            "<br><br><b>Come leggerlo:</b> il possibile cambio di regime ha ora una conferma iniziale "
            "coerente con flussi, struttura COT e prezzo Weekly."
        )
    else:
        future_regime_teaching = ""

    future_regime_html = smart["future_regime_text"].replace(chr(10), "<br>") + future_regime_teaching

    card(
        "SMART MONEY REPORT — " + spec.trend_label,
        smart["simple_title"],
        (
            f"<b>Come sono attualmente posizionati i Fondi?</b><br>{smart['current_position']}<br>{current_position_html}<br><br>"
            f"<b>Cosa hanno fatto i Fondi nell'ultimo report?</b><br>{smart['last_report']}<br><br>"
            f"<b>Come è cambiato il posizionamento dei Fondi?</b><br>{structure_html}<br><br>"
            f"<b>Quanto è estremo il loro posizionamento?</b><br>{smart['positioning']}<br>{positioning_html}<br><br>"
            f"<b>Il movimento è sostenuto dagli operatori (OI 3–6 W)?</b><br>{oi_report_html}<br><br>"
            f"<b>I Top 8 Trader sono molto esposti?</b><br>{smart['concentration_state']}<br><br>"
            f"<b>Si sta preparando un possibile cambio di regime?</b><br>{future_regime_html}"
        ),
        main_accent,
    )

    left, middle, right = st.columns([1.15, 0.95, 1.15])
    with left:
        card("LETTURA SEMPLICE", smart["simple_title"], smart["simple_detail"], main_accent)
    with middle:
        participation_accent = (
            "#15803D" if "È SOSTENUTO" in smart["oi_quality"]
            else "#C2410C" if "PERDE PARTECIPAZIONE" in smart["oi_quality"]
            else "#6B7280" if "PARTECIPAZIONE STABILE" in smart["oi_quality"] or "NON HA UNA DIREZIONE UNIFORME" in smart["oi_quality"]
            else "#9A6700"
        )
        card(
            "IL MOVIMENTO È SOSTENUTO DAGLI OPERATORI (OI 3–6 W)?",
            smart["oi_quality"],
            (
                f"OI Index 52W: {fmt_decimal(smart['oi_index_52w'], 1)} | {smart['oi_index_52w_state']}<br><br>"
                f"<b>Variazione OI 3W:</b> {fmt_pct(smart['pct_oi_3w'], signed=True, digits=2)}<br>"
                f"<b>Variazione OI 6W:</b> {fmt_pct(smart['pct_oi_6w'], signed=True, digits=2)}"
            ),
            participation_accent,
        )
    with right:
        card("COSA FARE", smart["plain_action"], smart["explanation"], main_accent)

    if report_mode == "Completo":
        with st.expander("Dettagli dell'analisi", expanded=True):
            st.text(smart["full_diagnosis"])
            if spec.family == "commodity":
                st.markdown("**SCADENZE DEI FUTURE**")
                st.write(f"Term Structure manuale: {term_structure}")


    # =============================================================================
    # QUADRO TECNICO
    # =============================================================================
    st.header("2. Quadro tecnico")
    tech_rows = [
        {"Voce": "Famiglia mercato", "Valore": spec.market_family_label},
        {"Voce": "Report utilizzato", "Valore": spec.specific_report},
        {"Voce": "Categoria trend", "Valore": spec.trend_label},
        {"Voce": "Controparte", "Valore": spec.counter_label},
        {"Voce": "Data posizioni COT", "Valore": smart["report_date"].strftime("%d/%m/%Y")},
        {"Voce": "Codice CFTC", "Valore": smart["cftc_code"]},
        {"Voce": f"Δ {spec.trend_label} Long", "Valore": fmt_number(smart["trend_chg_l"], signed=True)},
        {"Voce": f"Δ {spec.trend_label} Short", "Valore": fmt_number(smart["trend_chg_s"], signed=True)},
        {"Voce": f"Δ {spec.counter_label} Long", "Valore": fmt_number(smart["counter_chg_l"], signed=True)},
        {"Voce": f"Δ {spec.counter_label} Short", "Valore": fmt_number(smart["counter_chg_s"], signed=True)},
        {"Voce": "Come sono attualmente posizionati i Fondi?", "Valore": f"{smart['current_position']} | {smart['current_position_value']}"},
        {"Voce": f"COT Index {cot_lookback}W usato dal motore", "Valore": fmt_decimal(smart["cot_index"], 1)},
        {"Voce": "COT Index 26W", "Valore": fmt_decimal(smart["cot_index_26w"], 1)},
        {"Voce": "COT Index 156W", "Valore": fmt_decimal(smart["cot_index_156w"], 1)},
        {"Voce": "Quanto è estremo il loro posizionamento?", "Valore": f"{smart['positioning']} | {smart['extreme_horizons'] or 'Nessun nuovo estremo Net'}"},
        {"Voce": "Esposizione direzionale Long / Short", "Valore": f"{fmt_pct(smart['directional_long_pct'])} / {fmt_pct(smart['directional_short_pct'])}"},
        {
            "Voce": "Il movimento è sostenuto dagli operatori? (OI 3–6 W)",
            "Valore": smart["oi_quality"],
        },
        {"Voce": "Variazione Open Interest 3W", "Valore": fmt_pct(smart["pct_oi_3w"], signed=True, digits=2)},
        {"Voce": "Variazione Open Interest 6W", "Valore": fmt_pct(smart["pct_oi_6w"], signed=True, digits=2)},
        {"Voce": "OI Index 52W — livello partecipazione", "Valore": f"{fmt_decimal(smart['oi_index_52w'], 1)} — {smart['oi_index_52w_state']}"},
        {"Voce": "Rapid Shift controparte 6W", "Valore": f"{fmt_signed_decimal(alignment.get('rapid_shift_6w', math.nan), 1)} — {alignment.get('rapid_shift_state', 'NON DISPONIBILE')}"},
        {"Voce": "Uso Term Structure", "Valore": term_usage_status},
        {"Voce": "Term Structure", "Valore": term_structure},
        {"Voce": "Top 8 Net Long", "Valore": f"{fmt_pct(smart['conc_long'])} | Pctl {fmt_pct(smart['conc_long_rank'])}"},
        {"Voce": "Top 8 Net Short", "Valore": f"{fmt_pct(smart['conc_short'])} | Pctl {fmt_pct(smart['conc_short_rank'])}"},
        {"Voce": "I Top 8 Trader sono molto esposti?", "Valore": smart["concentration_state"]},
        {"Voce": "Lettura Top 8 Trader", "Valore": smart["concentration_detail"]},
        {"Voce": "Si sta preparando un possibile cambio di regime?", "Valore": smart["future_regime_title"]},
        {"Voce": "Alignment 156W — segnali rialzisti / ribassisti", "Valore": f"{smart['alignment_bull_count']}/3 / {smart['alignment_bear_count']}/3"},
    ]
    st.dataframe(pd.DataFrame(tech_rows), width="stretch", hide_index=True)

    flow_rows = pd.DataFrame(
        [
            {"Orizzonte": "1W", spec.trend_label: smart["trend_flow_1w"], spec.counter_label: smart["counter_flow_1w"]},
            {"Orizzonte": "3W", spec.trend_label: smart["trend_flow_3w"], spec.counter_label: smart["counter_flow_3w"]},
            {"Orizzonte": "6W", spec.trend_label: smart["trend_flow_6w"], spec.counter_label: smart["counter_flow_6w"]},
        ]
    )
    st.dataframe(flow_rows, width="stretch", hide_index=True)


    # =============================================================================
    # COT ALIGNMENT MAP
    # =============================================================================
    st.header("3. COT Alignment Map")
    st.caption(
        "Contesto contrarian fisso a 156 settimane, come TradingView V1.5.36. La regola è la stessa per Commodity, FX e altri Financial. "
        "Il 2/3 è soltanto parziale; il 3/3 prepara il setup, mentre flussi, Net Position, struttura 3–6W e prezzo determinano lo stadio del possibile cambio di regime."
    )

    if alignment["available"]:
        a1, a2, a3 = st.columns(3)
        a1.metric(
            spec.trend_label,
            fmt_decimal(alignment["speculative_index"], 1),
            alignment["speculative_zone"],
        )
        a2.metric(
            spec.counter_label,
            fmt_decimal(alignment["counterparty_index"], 1),
            alignment["counterparty_zone"],
        )
        a3.metric(
            "Nonreportable / Small Traders",
            fmt_decimal(alignment["small_index"], 1),
            alignment["small_zone"],
        )

        st.metric(
            f"Rapid Shift {spec.counter_label} 6W",
            fmt_signed_decimal(alignment["rapid_shift_6w"], 1),
            alignment["rapid_shift_state"],
            help="Variazione del COT Index della controparte rispetto a sei report settimanali fa. È informativa e non modifica lo Score.",
        )

        score_left, score_right = st.columns(2)
        score_left.metric("Allineamento rialzista", f"{alignment['bull_score']}/3")
        score_right.metric("Allineamento ribassista", f"{alignment['bear_score']}/3")

        card(
            "CONTESTO COT — ALIGNMENT MAP",
            alignment["state"],
            alignment["description"] +
            "<br><b>Regola:</b> Alignment Map prepara il contesto; Net Position conferma il flusso; il prezzo attiva l'operazione.",
            accent_for_state(alignment["state"]),
        )
    else:
        st.warning(
            "COT Alignment Map non disponibile: il dataset non ha restituito una serie valida "
            "per Nonreportable / Small Traders. Il responso Smart Money resta comunque utilizzabile."
        )


    # =============================================================================
    # TERM STRUCTURE
    # =============================================================================
    st.subheader("Term Structure")
    if term_usage_status == "NON APPLICABILE":
        st.info("NON APPLICABILE — questo mercato non richiede l'inserimento della curva M1–M2.")
    else:
        st.info("OPZIONALE — non modifica il responso Smart Money né il COT Alignment Map.")

    if spec.family == "commodity":
        if term_structure == "Backwardation":
            st.success("Backwardation: il contratto vicino quota sopra il successivo. È un'informazione aggiuntiva sulla curva, non un segnale COT.")
        elif term_structure == "Contango":
            st.warning("Contango: il contratto successivo quota sopra il vicino. È un'informazione aggiuntiva sulla curva.")
        elif term_structure == "Curva piatta":
            st.info("Curva piatta: differenza M1–M2 non significativa.")
        else:
            st.info("Valore non impostato. L'analisi Smart Money e l'Alignment Map restano complete.")


    # =============================================================================
    # GRAFICI
    # =============================================================================
    st.header("4. Grafici")
    chart_col1, chart_col2 = st.columns(2)
    with chart_col1:
        st.plotly_chart(plot_cot_index(specific_df, spec.trend_label), width="stretch")
    with chart_col2:
        st.plotly_chart(plot_net_positions(specific_df, spec.trend_label, spec.counter_label), width="stretch")

    if not weekly_price.empty:
        st.plotly_chart(plot_weekly_price(weekly_price, yahoo_ticker), width="stretch")


    # =============================================================================
    # INTERROGAZIONE AI
    # =============================================================================
    st.header("5. Analisi e interrogazione AI")
    st.caption(
        "L'AI interpreta il motore Smart Money, la qualità dei flussi e il COT Alignment Map già calcolati. "
        "Non modifica i dati e non trasforma il COT in un segnale immediato di ingresso."
    )


    def secret_value(name: str, default: str) -> str:
        try:
            value = st.secrets.get(name, default)
            return str(value) if value not in (None, "") else default
        except Exception:
            return default


    def build_ai_prompt(user_question: str, operational_prompt: str) -> tuple[str, str]:
        """Costruisce il prompt AI in due modalità automatiche.

        - Campo domanda vuoto: applica integralmente PROMPT.TXT.
        - Campo domanda compilato: ignora PROMPT.TXT e risponde soltanto alla domanda specifica.
        """
        custom_question = user_question.strip()
        use_custom_question = bool(custom_question)

        if use_custom_question:
            instruction_block = f"""
    MODALITÀ: DOMANDA SPECIFICA

    Ignora completamente il prompt operativo preimpostato contenuto in {AI_PROMPT_FILENAME}.
    Non generare automaticamente la lettura completa né il post social.
    Rispondi esclusivamente e direttamente alla domanda dell'utente riportata sotto.
    Usa soltanto i dati strutturati della dashboard e mantieni una risposta semplice, concreta e coerente con il responso deterministico.

    DOMANDA DELL'UTENTE
    {custom_question}
    """.strip()
            request_mode = "Domanda specifica — PROMPT.TXT ignorato"
        else:
            instruction_block = f"""
    MODALITÀ: ANALISI COMPLETA PREIMPOSTATA

    ISTRUZIONI OPERATIVE CARICATE DA {AI_PROMPT_FILENAME}

    {operational_prompt}

    RICHIESTA
    Applica integralmente il prompt operativo preimpostato e produci sia la lettura completa sia la versione breve per il post social.
    """.strip()
            request_mode = f"Analisi completa — {AI_PROMPT_FILENAME}"

        prompt = f"""
    {instruction_block}

    ==================================================
    FONTE DATI DISPONIBILE NELLA DASHBOARD PYTHON
    ==================================================

    In questa interrogazione non è allegato automaticamente uno screenshot.
    Usa come fonte principale esclusivamente i dati strutturati riportati sotto.
    La dashboard calcola automaticamente Smart Money Report, tabella tecnica e COT Alignment Map.
    Non calcola ancora POC, supporti o resistenze.
    Per questi elementi scrivi "dato non chiaramente leggibile" e non inventare valori o livelli.
    Quando un trigger numerico non è disponibile, descrivi soltanto la condizione necessaria.

    MERCATO
    - Strumento: {spec.label}
    - Famiglia: {spec.market_family_label}
    - Report CFTC: {spec.specific_report}
    - Categoria principale: {spec.trend_label}
    - Controparte: {spec.counter_label}
    - Data posizioni: {smart['report_date'].strftime('%d/%m/%Y')}
    - Data report precedente: {smart['previous_date'].strftime('%d/%m/%Y')}

    POSIZIONAMENTO E FLUSSI
    - COT Index motore ({cot_lookback}W): {smart['cot_index']:.1f}
    - COT Index 26W: {fmt_decimal(smart['cot_index_26w'], 1)}
    - COT Index 156W: {fmt_decimal(smart['cot_index_156w'], 1)} — {smart['positioning']}
    - Posizione attuale dei Fondi: {smart['current_position']} — {smart['current_position_value']}
    - Nuovi estremi Net: {smart['extreme_horizons'] or 'NESSUNO'}
    - Open Interest: {smart['oi']:.0f}
    - Variazione Open Interest 1W: {smart['pct_delta_oi']:+.2f}%
    - Sostegno del movimento OI 3-6W: {smart['oi_quality']}
    - Variazione Open Interest 3W: {fmt_pct(smart['pct_oi_3w'], signed=True, digits=2)}
    - Variazione Open Interest 6W: {fmt_pct(smart['pct_oi_6w'], signed=True, digits=2)}
    - OI Index 52W: {fmt_decimal(smart['oi_index_52w'], 1)} — {smart['oi_index_52w_state']}
    - Delta {spec.trend_label} Long: {smart['trend_chg_l']:+.0f}
    - Delta {spec.trend_label} Short: {smart['trend_chg_s']:+.0f}
    - Flusso {spec.trend_label}: 1W {smart['trend_flow_1w']:+.0f}, 3W {smart['trend_flow_3w']:+.0f}, 6W {smart['trend_flow_6w']:+.0f}
    - Delta {spec.counter_label} Long: {smart['counter_chg_l']:+.0f}
    - Delta {spec.counter_label} Short: {smart['counter_chg_s']:+.0f}
    - Flusso {spec.counter_label}: 1W {smart['counter_flow_1w']:+.0f}, 3W {smart['counter_flow_3w']:+.0f}, 6W {smart['counter_flow_6w']:+.0f}
    - Concentrazione Top 8: {smart['concentration_state']}
    - Spiegazione Top 8: {smart['concentration_detail']}
    - Top 8 Long: {fmt_pct(smart['conc_long'])}, percentile {fmt_pct(smart['conc_long_rank'])}
    - Top 8 Short: {fmt_pct(smart['conc_short'])}, percentile {fmt_pct(smart['conc_short_rank'])}

    COT ALIGNMENT MAP
    - Disponibile: {"SÌ" if alignment['available'] else "NO"}
    - {spec.trend_label}: {fmt_decimal(alignment['speculative_index'], 1)} — {alignment.get('speculative_zone', 'N/A')}
    - {spec.counter_label}: {fmt_decimal(alignment['counterparty_index'], 1)} — {alignment.get('counterparty_zone', 'N/A')}
    - Nonreportable / Small Traders: {fmt_decimal(alignment['small_index'], 1)} — {alignment.get('small_zone', 'N/A')}
    - Rapid Shift {spec.counter_label} 6W: {fmt_signed_decimal(alignment.get('rapid_shift_6w', math.nan), 1)} — {alignment.get('rapid_shift_state', 'NON DISPONIBILE')}
    - Allineamento rialzista: {alignment['bull_score']}/3
    - Allineamento ribassista: {alignment['bear_score']}/3
    - Stato: {alignment['state']}
    - Lettura: {alignment['description']}

    POSSIBILE CAMBIO DI REGIME — ALIGNMENT 156W
    - Stato: {smart['future_regime_title']}
    - Dettaglio: {smart['future_regime_text']}
    - {spec.trend_label} 156W: {fmt_decimal(smart['alignment_trend_index_156w'], 1)} / 100
    - {spec.counter_label} 156W: {fmt_decimal(smart['alignment_counter_index_156w'], 1)} / 100
    - Small Traders / Nonreportable 156W: {fmt_decimal(smart['alignment_small_index_156w'], 1)} / 100
    - Segnali rialzisti: {smart['alignment_bull_count']}/3
    - Segnali ribassisti: {smart['alignment_bear_count']}/3

    PREZZO WEEKLY
    - Stato: {price_analysis['text']}
    - Dettaglio: {price_analysis['detail']}

    TERM STRUCTURE
    - Stato d'uso: {term_usage_status}
    - Valore manuale: {term_structure}

    RESPONSO DETERMINISTICO SMART MONEY
    - Bias finale: {smart['final_bias']}
    - Lettura semplice: {smart['simple_title']}
    - Spiegazione semplice: {smart['simple_detail']}
    - Ultimo report: {smart['last_report']}
    - Struttura 3-6W: {smart['structure']}
    - Lettura controparte: {smart['counter_reading']}
    - Cosa fare: {smart['plain_action']}
    - Perché: {smart['explanation']}

    VINCOLI FINALI DELLA DASHBOARD
    - Il responso deterministico è il punto di partenza: non contraddirlo senza dichiarare chiaramente il limite dei dati.
    - Usa i valori calcolati dell'Alignment Map; non inventare screenshot, POC, supporti, resistenze o livelli tecnici.
    - Distingui nuovi Long, nuovi Short, short covering, liquidazione Long e flusso misto usando i delta disponibili.
    - Un COT Index estremo descrive la collocazione della Net Position nel proprio range storico, non necessariamente una posizione Net Long o Net Short e non un segnale automatico di inversione.
    - L'Alignment Map usato per il possibile cambio di regime è fisso a 156W e usa la stessa logica contrarian per tutte le famiglie: Trend basso + Controparte alta + Small basso per Bull; l'opposto per Bear.
    - Un Alignment 2/3 è soltanto parziale. Un 3/3 è un setup, non ancora un cambio di regime confermato.
    - "IN SVILUPPO" richiede 3/3 + nuovi Long/Short coerenti + Net Position nella stessa direzione. "CONFERMATO" richiede inoltre prezzo Weekly e struttura macro coerenti.
    - Non promuovere mai 2/3 o 3/3 grezzo a regime confermato. Usa lo stato deterministico già calcolato dalla dashboard.
    - Il Rapid Shift 6W segnala velocità del cambiamento della controparte, non un ingresso automatico.
    - L'OI Index 52W misura il livello di partecipazione nel range annuale e non indica da solo la direzione.
    - Rapid Shift 6W e OI Index 52W non modificano il responso deterministico né lo Score.
    - Considera la Term Structure soltanto secondo lo stato d'uso indicato.
    - Il COT non è un segnale immediato di ingresso.
    - In modalità domanda specifica, rispondi soltanto a quella domanda senza applicare formato, sezioni o output imposti da {AI_PROMPT_FILENAME}.
    """.strip()

        return prompt, request_mode


    operational_prompt, operational_prompt_source = load_operational_prompt()
    with st.expander(f"Prompt operativo preimpostato — {operational_prompt_source}", expanded=False):
        st.caption(
            "Il testo viene caricato automaticamente a ogni esecuzione. Per aggiornarlo modifica "
            f"{AI_PROMPT_FILENAME} nel repository, senza intervenire sul codice Python."
        )
        st.code(operational_prompt, language=None)

    ai_col1, ai_col2 = st.columns([1, 1])
    with ai_col1:
        ai_provider = st.selectbox("Provider AI", ["Google Gemini", "Groq"], key="ai_provider")
    with ai_col2:
        if ai_provider == "Google Gemini":
            ai_model = st.text_input(
                "Modello Gemini",
                value=secret_value("GEMINI_MODEL", "gemini-3.5-flash"),
                key="ai_model_gemini",
            )
        else:
            ai_model = st.text_input(
                "Modello Groq",
                value=secret_value("GROQ_MODEL", "openai/gpt-oss-120b"),
                key="ai_model_groq",
            )

    with st.form("cot_ai_form"):
        ai_question = st.text_area(
            "Domanda specifica — se compilata sostituisce PROMPT.TXT",
            placeholder=(
                "Lascia vuoto per ottenere l'analisi completa prevista da PROMPT.TXT. "
                "Scrivi una domanda per ignorare PROMPT.TXT e ricevere soltanto una risposta mirata, "
                "per esempio: concentrati solo sulla qualità dei flussi dell'ultimo report."
            ),
            help=(
                "Campo vuoto: viene applicato PROMPT.TXT. Campo compilato: PROMPT.TXT viene ignorato "
                "e l'AI risponde esclusivamente alla domanda inserita."
            ),
            height=105,
            key="cot_ai_question",
        )
        ai_submit = st.form_submit_button("Genera analisi con AI", type="primary", width="stretch")

    if ai_submit:
        prompt_ai, ai_request_mode = build_ai_prompt(ai_question, operational_prompt)
        with st.spinner(f"Interrogo {ai_provider}..."):
            try:
                if ai_provider == "Google Gemini":
                    from google import genai

                    api_key = secret_value("GEMINI_API_KEY", "")
                    if not api_key:
                        raise KeyError("GEMINI_API_KEY")
                    client = genai.Client(api_key=api_key)
                    response = client.models.generate_content(model=ai_model, contents=prompt_ai)
                    answer = response.text or "Nessuna risposta ricevuta da Gemini."
                else:
                    from groq import Groq

                    api_key = secret_value("GROQ_API_KEY", "")
                    if not api_key:
                        raise KeyError("GROQ_API_KEY")
                    client = Groq(api_key=api_key)
                    response = client.chat.completions.create(
                        model=ai_model,
                        messages=[
                            {
                                "role": "system",
                                "content": (
                                    "Sei un analista COT prudente. Rispetta rigorosamente i dati forniti, "
                                    "non inventare informazioni e non sostituire il responso deterministico. "
                                    "Segui la modalità dichiarata nel messaggio utente: quando è DOMANDA SPECIFICA, "
                                    "ignora il prompt operativo preimpostato e rispondi soltanto alla domanda."
                                ),
                            },
                            {"role": "user", "content": prompt_ai},
                        ],
                        temperature=0.2,
                    )
                    answer = response.choices[0].message.content or "Nessuna risposta ricevuta da Groq."

                st.session_state["cot_ai_answer"] = answer
                st.session_state["cot_ai_context"] = (
                    f"{spec.root}|{smart['report_date'].isoformat()}|{ai_provider}|{ai_model}|{ai_request_mode}"
                )
            except KeyError as exc:
                st.error(
                    f"Chiave API mancante nei Secrets di Streamlit: {exc}. "
                    "Configura GEMINI_API_KEY oppure GROQ_API_KEY."
                )
            except ModuleNotFoundError as exc:
                st.error(
                    f"Dipendenza AI non installata: {exc}. Aggiorna requirements.txt con google-genai e groq."
                )
            except Exception as exc:
                error_text = str(exc)
                if "503" in error_text or "UNAVAILABLE" in error_text.upper():
                    st.warning("Il provider AI è temporaneamente non disponibile. Riprova più tardi o seleziona l'altro provider.")
                else:
                    st.error(f"Errore durante la comunicazione con l'AI: {exc}")

    saved_ai_context = st.session_state.get("cot_ai_context", "")
    if st.session_state.get("cot_ai_answer") and saved_ai_context.startswith(current_ai_data_context + "|"):
        st.markdown("### Risposta dell'AI")
        st.write(st.session_state["cot_ai_answer"])
        st.caption("Contesto usato: " + saved_ai_context)


    # =============================================================================
    # DATI STORICI
    # =============================================================================
    with st.expander("Dati storici COT ed esportazione", expanded=False):
        st.caption("La risposta AI resta soltanto a video. Qui puoi scaricare esclusivamente lo storico numerico COT del mercato selezionato.")
        history_export = specific_df.copy()
        history_export["date"] = history_export["date"].dt.strftime("%Y-%m-%d")
        st.download_button(
            "Scarica storico COT CSV",
            data=history_export.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"cot_storico_{spec.root}.csv",
            mime="text/csv",
            width="stretch",
        )


    # =============================================================================
    # DIAGNOSTICA
    # =============================================================================
    if show_debug:
        st.header("Diagnostica")
        st.write("Campi specifici risolti")
        st.json(specific_fields)
        st.write("Ultima riga specifica normalizzata")
        st.dataframe(specific_df.tail(1), width="stretch", hide_index=True)

    st.divider()
    st.warning(
        "Il COT è un dato settimanale ritardato e descrive posizioni rilevate il martedì. "
        "Questa dashboard fornisce contesto e validazione, non un segnale automatico di ingresso."
    )


# =============================================================================
# NAVIGAZIONE PRINCIPALE
# =============================================================================
with st.sidebar:
    st.header("Navigazione")
    app_section = st.radio(
        "Sezione",
        ["Analisi singolo strumento", "COT Screener"],
        index=0,
        help="Le due sezioni sono indipendenti: la prima approfondisce un mercato, la seconda crea la classifica completa.",
    )
    st.divider()

if app_section == "Analisi singolo strumento":
    render_single_analysis()
else:
    render_screener()
